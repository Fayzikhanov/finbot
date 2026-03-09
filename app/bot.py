from __future__ import annotations

import asyncio
import calendar
import html
import logging
import math
import os
import random
import re
import tempfile
from datetime import date, datetime, time, timedelta, timezone
from urllib.parse import quote
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    MenuButtonCommands,
    MenuButtonWebApp,
    ReplyKeyboardMarkup,
    Update,
    User,
    WebAppInfo,
)
from telegram.constants import ParseMode
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ConversationHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

from app.ai import AIService, ParsedEvent
from app.config import Settings
from app.db import (
    DEFAULT_DAILY_REPORT_HOUR,
    DEFAULT_DAILY_REPORT_MINUTE,
    DEFAULT_REPORT_TIMEZONE,
    TRANSFER_IN_CATEGORY,
    TRANSFER_OUT_CATEGORY,
    Database,
    FinanceTransaction,
)
from app.i18n import (
    DEFAULT_LANGUAGE,
    SUPPORTED_LANGUAGES,
    language_flag_label,
    language_short_label,
    normalize_language,
    t,
)


logger = logging.getLogger(__name__)

MONTH_NAMES_RU = [
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
]

MONTH_NAMES_RU_GENITIVE = [
    "января",
    "февраля",
    "марта",
    "апреля",
    "мая",
    "июня",
    "июля",
    "августа",
    "сентября",
    "октября",
    "ноября",
    "декабря",
]

MONTH_NAMES_UZ = [
    "yanvar",
    "fevral",
    "mart",
    "aprel",
    "may",
    "iyun",
    "iyul",
    "avgust",
    "sentabr",
    "oktabr",
    "noyabr",
    "dekabr",
]

MONTH_NAMES_EN = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

NAME_PATTERN = re.compile(r"^[A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\-\s]{0,31}$")
NAME_EXPLICIT_PATTERNS = [
    re.compile(
        r"^\s*меня\s+зовут\s+([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^\s*зови\s+меня\s+([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^\s*(?:поменяй|измени)\s+мое?\s+имя\s+на\s+([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^\s*change\s+my\s+name\s+to\s+([A-Za-z][A-Za-z\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^\s*я\s+([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^\s*my\s+name\s+is\s+([A-Za-z][A-Za-z\-\s]{0,31})\s*$",
        re.IGNORECASE,
    ),
]

FINANCE_HINTS = (
    "купил",
    "купила",
    "потрат",
    "расход",
    "заплат",
    "оплат",
    "покупк",
    "кофе",
    "продукт",
    "получил",
    "получила",
    "зарплат",
    "зп",
    "доход",
    "income",
    "expense",
    "spent",
    "salary",
)

NON_NAME_WORDS = {
    "да",
    "нет",
    "ок",
    "ага",
    "привет",
    "здравствуйте",
    "добрый",
    "вечер",
    "утро",
    "я",
    "а",
    "ты",
    "он",
    "она",
    "мы",
    "вы",
    "они",
    "кто",
    "что",
}

EDIT_STATE_KEY = "pending_transaction_edit"
REPORT_STATE_KEY = "report_state_by_chat"
REPORT_FILTER_INPUT_KEY = "pending_report_filter_input"
REPORT_SINGLE_DATE_CONFIRM_KEY = "pending_report_single_date_confirm"
TRANSFER_PENDING_KEY = "pending_transfer_input"
ONBOARDING_STATE_KEY = "group_onboarding_state"
PENDING_GUESS_KEY = "pending_guess_event"
ASSISTANT_SETTINGS_TRIGGER_MAP_KEY = "assistant_settings_trigger_map"
PENDING_LANGUAGE_CHANGE_KEY = "pending_language_change"
PENDING_START_AFTER_LANGUAGE_KEY = "pending_start_after_language"
PERSONAL_ONBOARDING_STEP_KEY = "personal_onboarding_step"
ONBOARDING_CLEANUP_IDS_KEY = "onboarding_cleanup_msg_ids"
MAIN_MENU_REPORT_TEXT = "📊 Отчёты"
MAIN_MENU_SETTINGS_TEXT = "⚙️ Настройки Ассистента"
MAIN_MENU_APP_TEXT = "📱 Приложение"
SUPPORT_MENU, SUPPORT_MESSAGE, SUPPORT_BUG = range(3)
PROFILE_MENU, PROFILE_NAME, PROFILE_PHONE, PROFILE_EMAIL, PROFILE_BIRTHDATE = range(10, 15)
BOT_REVIEW_MENU, BOT_REVIEW_RATING, BOT_REVIEW_COMMENT = range(20, 23)
PROFILE_WIZARD_KEY = "profile_wizard_state"
PROFILE_VIEW_REF_KEY = "profile_view_ref"
BOT_REVIEW_PENDING_KEY = "bot_review_pending"
PROFILE_PHONE_RE = re.compile(r"^\+\d{7,15}$")
PROFILE_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
REPORT_TIME_INPUT_RE = re.compile(r"^(?:[01]?\d|2[0-3]):[0-5]\d$")
SCHEDULED_REPORT_TICK_SECONDS = 30
SCHEDULED_REPORT_BOOT_DELAY_SECONDS = 5

NAME_GREETING_VARIANTS = (
    "Рад знакомству, {name}! ✅",
    "Отлично, {name}, запомнил 👍",
    "Приятно познакомиться, {name}. Поехали дальше!",
    "Супер, {name}! Буду так обращаться.",
    "Класс, {name}. Записал 👌",
)

CATEGORY_LABELS_RU = {
    "home_rent": "Жильё и дом / Аренда",
    "home_mortgage": "Жильё и дом / Ипотека",
    "home_utilities": "Жильё и дом / Коммунальные услуги",
    "home_electricity": "Жильё и дом / Электричество",
    "home_gas": "Жильё и дом / Газ",
    "home_water": "Жильё и дом / Вода",
    "home_heating": "Жильё и дом / Отопление",
    "home_internet": "Жильё и дом / Интернет",
    "home_tv": "Жильё и дом / ТВ",
    "home_repair": "Жильё и дом / Ремонт",
    "home_furniture": "Жильё и дом / Мебель",
    "home_appliances": "Жильё и дом / Бытовая техника",
    "home_goods": "Жильё и дом / Товары для дома",
    "groceries_products": "Продукты и быт / Продукты",
    "groceries_household": "Продукты и быт / Бытовая химия",
    "groceries_water": "Продукты и быт / Вода",
    "groceries_delivery": "Продукты и быт / Доставка продуктов",
    "cafe_cafe": "Кафе и рестораны / Кафе",
    "cafe_restaurant": "Кафе и рестораны / Ресторан",
    "cafe_fastfood": "Кафе и рестораны / Фастфуд",
    "cafe_coffee": "Кафе и рестораны / Кофе",
    "cafe_streetfood": "Кафе и рестораны / Стритфуд",
    "cafe_delivery": "Кафе и рестораны / Доставка еды",
    "cafe_bar": "Кафе и рестораны / Бар",
    "cafe_hookah": "Кафе и рестораны / Кальянная",
    "transport_fuel": "Транспорт / Бензин",
    "transport_gas": "Транспорт / Газ",
    "transport_taxi": "Транспорт / Такси",
    "transport_metro": "Транспорт / Метро",
    "transport_bus": "Транспорт / Автобус",
    "transport_parking": "Транспорт / Парковка",
    "transport_fine": "Транспорт / Штраф",
    "transport_repair": "Транспорт / Ремонт авто",
    "transport_insurance": "Транспорт / Страховка авто",
    "work_subscriptions": "Работа и бизнес / Подписки",
    "work_software": "Работа и бизнес / Софт",
    "work_ads": "Работа и бизнес / Реклама",
    "work_hosting": "Работа и бизнес / Хостинг",
    "work_domain": "Работа и бизнес / Домен",
    "work_courses": "Работа и бизнес / Курсы",
    "work_taxes": "Работа и бизнес / Налоги",
    "work_tools": "Работа и бизнес / Инструменты",
    "education_school": "Образование / Школа",
    "education_kindergarten": "Образование / Садик",
    "education_tutor": "Образование / Репетитор",
    "education_online_courses": "Образование / Онлайн-курсы",
    "education_books": "Образование / Книги",
    "health_pharmacy": "Здоровье / Аптека",
    "health_doctor": "Здоровье / Врач",
    "health_tests": "Здоровье / Анализы",
    "health_dentist": "Здоровье / Стоматолог",
    "health_gym": "Здоровье / Спортзал",
    "health_supplements": "Здоровье / БАДы",
    "fashion_clothes": "Одежда и уход / Одежда",
    "fashion_shoes": "Одежда и уход / Обувь",
    "fashion_cosmetics": "Одежда и уход / Косметика",
    "fashion_hairdresser": "Одежда и уход / Парикмахер",
    "fashion_manicure": "Одежда и уход / Маникюр",
    "fashion_accessories": "Одежда и уход / Аксессуары",
    "kids_toys": "Дети / Игрушки",
    "kids_clothes": "Дети / Детская одежда",
    "kids_diapers": "Дети / Подгузники",
    "kids_food": "Дети / Детское питание",
    "kids_entertainment": "Дети / Развлечения",
    "leisure_subscriptions": "Развлечения / Подписки",
    "leisure_games": "Развлечения / Игры",
    "leisure_cinema": "Развлечения / Кино",
    "leisure_travel": "Развлечения / Путешествия",
    "leisure_hotel": "Развлечения / Отель",
    "leisure_flights": "Развлечения / Авиабилеты",
    "leisure_gifts": "Развлечения / Подарки",
    "leisure_holidays": "Развлечения / Праздники",
    "finance_credit": "Финансы / Кредит",
    "finance_installment": "Финансы / Рассрочка",
    "finance_interest": "Финансы / Проценты",
    "finance_bank_fees": "Финансы / Банковская комиссия",
    "finance_transfers": "Финансы / Переводы",
    "finance_investments": "Финансы / Инвестиции",
    "finance_savings": "Финансы / Сбережения",
    "finance_debt_given": "Финансы / Долг дал",
    "finance_debt_received": "Финансы / Долг получил",
    "pets_food": "Животные / Корм",
    "pets_vet": "Животные / Ветеринар",
    "expense_other": "Прочие расходы",
    "salary": "Зарплата",
    "bonus": "Бонус/премия",
    "windfall": "Выигрыш/находка",
    "profit": "Прибыль",
    "cashback": "Кэшбэк",
    "gift": "Подарок",
    "income_other": "Прочие доходы",
    "transfer_internal_out": "Переводы / Отправлено",
    "transfer_internal_in": "Переводы / Получено",
}


def _display_name(update: Update) -> str:
    user = update.effective_user
    if not user:
        return "Unknown"
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    if full_name:
        return full_name
    return user.username or str(user.id)


def _user_handle(user: User) -> str:
    if user.username:
        return f"@{user.username}"
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    return full_name or str(user.id)


def _member_handle(member: dict[str, object]) -> str:
    username = member.get("username")
    full_name = str(member.get("full_name") or "Участник")
    return f"@{username}" if username else full_name


def _fmt_money(amount: float, currency: str, *, signed: bool = False) -> str:
    rounded = int(round(amount))
    sign = ""
    if signed and rounded > 0:
        sign = "+"
    abs_amount = abs(rounded)
    amount_str = f"{abs_amount:,}".replace(",", " ")

    currency_u = (currency or "").upper()
    if currency_u == "UZS":
        suffix = "сум"
    elif currency_u == "RUB":
        suffix = "₽"
    elif currency_u == "USD":
        suffix = "$"
    elif currency_u == "EUR":
        suffix = "€"
    else:
        suffix = currency_u or "сум"

    if rounded < 0 and not signed:
        return f"-{amount_str} {suffix}".strip()
    if rounded < 0 and signed:
        return f"-{amount_str} {suffix}".strip()
    return f"{sign}{amount_str} {suffix}".strip()


def _miniapp_button_text(workspace: object) -> str:
    workspace_type = ""
    if isinstance(workspace, dict):
        workspace_type = str(workspace.get("type") or "").strip().lower()
    else:
        workspace_type = str(getattr(workspace, "type", "") or "").strip().lower()

    if workspace_type == "personal":
        return "📱 Открыть личный дашборд"
    if workspace_type == "family":
        return "📱 Открыть семейный дашборд"
    return "📱 Открыть дашборд"


def build_miniapp_button(workspace: object, base_url: str) -> InlineKeyboardMarkup:
    button_text = _miniapp_button_text(workspace)

    return InlineKeyboardMarkup(
        [[InlineKeyboardButton(text=button_text, web_app=WebAppInfo(url=str(base_url).strip()))]]
    )


def build_miniapp_link_button(workspace: object, url: str) -> InlineKeyboardMarkup:
    button_text = _miniapp_button_text(workspace)
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton(text=button_text, url=str(url).strip())]]
    )


def _extract_amount_value(raw_text: str) -> float | None:
    amount_re = re.compile(r"(?<!\d)(\d{1,3}(?:[ \u00A0,]\d{3})+|\d+(?:[.,]\d+)?)")
    match = amount_re.search(raw_text)
    if not match:
        return None

    raw = match.group(1).replace(" ", "").replace("\u00A0", "").replace(",", "")
    try:
        amount = float(raw)
    except ValueError:
        return None

    # Support phrases like "1 млн", "1 миллион", "2 тыс", "1.5 млн"
    tail = (raw_text[match.end() : match.end() + 32] or "").lower().strip()
    multiplier = 1.0
    if re.match(r"^(млн|миллион|миллиона|миллионов)\b", tail):
        multiplier = 1_000_000.0
    elif re.match(r"^(млрд|миллиард|миллиарда|миллиардов)\b", tail):
        multiplier = 1_000_000_000.0
    elif re.match(r"^(тыс|тысяч|тысяча|тысячи|k)\b", tail):
        multiplier = 1_000.0

    amount *= multiplier
    if amount <= 0:
        return None
    return amount


def _has_amount_and_text(raw_text: str) -> bool:
    text = (raw_text or "").strip()
    if not text:
        return False
    has_digit = any(ch.isdigit() for ch in text)
    has_letter = bool(re.search(r"[A-Za-zА-Яа-яЁё]", text))
    return has_digit and has_letter


def _looks_like_math_expression(raw_text: str) -> bool:
    text = (raw_text or "").strip().lower()
    if not text:
        return False
    if re.search(r"\d+\s*[\+\-\*/]\s*\d+", text):
        return True
    if "сколько будет" in text:
        return True
    if "?" in text and any(ch.isdigit() for ch in text):
        return True
    return False


def _ru_category_label(category: str, kind: str | None = None) -> str:
    raw = (category or "").strip()
    if not raw:
        return "Категория не указана"

    key = raw.lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "other": "expense_other" if kind == "expense" else "income_other",
        "misc": "expense_other" if kind == "expense" else "income_other",
        "miscellaneous": "expense_other" if kind == "expense" else "income_other",
        "food": "groceries_products",
        "transport": "transport_taxi",
        "home": "home_utilities",
        "internet": "home_internet",
        "health": "health_doctor",
        "shopping": "fashion_clothes",
        "entertainment": "leisure_games",
        "winnings": "windfall",
        "winning": "windfall",
        "win": "windfall",
        "found_money": "windfall",
        "trading_profit": "profit",
        "profit_income": "profit",
        "pnl": "profit",
        "p&l": "profit",
    }
    key = aliases.get(key, key)

    label = CATEGORY_LABELS_RU.get(key)
    if label:
        return label

    # If AI returned a custom category, keep it readable.
    pretty = raw.replace("_", " ").replace("-", " ").strip()
    return pretty[:1].upper() + pretty[1:] if pretty else raw


def _current_month_period() -> tuple[datetime, datetime, str]:
    now = datetime.now()
    period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if now.month == 12:
        period_end = period_start.replace(year=now.year + 1, month=1)
    else:
        period_end = period_start.replace(month=now.month + 1)
    label = f"{MONTH_NAMES_RU[now.month - 1]} {now.year}"
    return period_start, period_end, label


def _today_period() -> tuple[datetime, datetime, str]:
    now = datetime.now()
    period_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    period_end = period_start + timedelta(days=1)
    label = now.strftime("%d/%m/%Y")
    return period_start, period_end, label


def _today_human_label() -> str:
    now = datetime.now()
    return f"{now.day} {MONTH_NAMES_RU_GENITIVE[now.month - 1]} {now.year}"


class FamilyFinanceBot:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.db = Database(settings.database_path)
        self.db.initialize()

        self.ai = AIService(
            api_key=settings.openai_api_key,
            extraction_model=settings.openai_model,
            transcribe_model=settings.openai_transcribe_model,
            default_currency=settings.default_currency,
        )

        self.name_prompted: set[tuple[int, int]] = set()
        self._can_read_all_groups: bool | None = None
        self._privacy_warned_chats: set[int] = set()
        self._bot_username: str | None = None
        self._scheduled_report_task: asyncio.Task | None = None
        self._scheduled_report_tick_in_progress = False

        self.application: Application = (
            ApplicationBuilder()
            .token(settings.telegram_bot_token)
            .connect_timeout(30)
            .read_timeout(30)
            .write_timeout(30)
            .pool_timeout(30)
            .post_init(self._post_init)
            .post_shutdown(self._post_shutdown)
            .build()
        )
        self._register_handlers()

    async def _post_init(self, application: Application) -> None:
        await self._sync_chat_menu_button(application)
        if self._scheduled_report_task is None or self._scheduled_report_task.done():
            self._scheduled_report_task = asyncio.create_task(
                self._scheduled_report_loop(),
                name="scheduled-report-loop",
            )

    async def _post_shutdown(self, application: Application) -> None:
        task = self._scheduled_report_task
        self._scheduled_report_task = None
        if task is None or task.done():
            return
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass

    async def _sync_chat_menu_button(self, application: Application) -> None:
        base_url = self._miniapp_base_url()
        try:
            if base_url:
                await application.bot.set_chat_menu_button(
                    menu_button=MenuButtonWebApp(
                        text="Открыть приложение",
                        web_app=WebAppInfo(url=base_url),
                    )
                )
            else:
                await application.bot.set_chat_menu_button(menu_button=MenuButtonCommands())
        except Exception:
            logger.exception("Failed to sync Telegram chat menu button")

    def _user_language(self, user_id: int | None) -> str:
        if not user_id:
            return DEFAULT_LANGUAGE
        try:
            profile = self.db.get_user_profile(int(user_id))
        except Exception:
            logger.exception("Failed to load user language: user_id=%s", user_id)
            return DEFAULT_LANGUAGE
        return normalize_language(str(profile.get("language") or DEFAULT_LANGUAGE))

    def _user_language_selected(self, user_id: int | None) -> bool:
        if not user_id:
            return False
        try:
            profile = self.db.get_user_profile(int(user_id))
        except Exception:
            logger.exception("Failed to read language_selected: user_id=%s", user_id)
            return False
        return bool(profile.get("language_selected"))

    def _update_user_language(self, user: User | None, language: str) -> str:
        if not user:
            return normalize_language(language)
        lang = normalize_language(language)
        try:
            self.db.register_user(
                telegram_user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                last_name=user.last_name,
            )
            self.db.set_user_language(user.id, lang)
        except Exception:
            logger.exception("Failed to update user language: user_id=%s lang=%s", user.id, lang)
            raise
        return lang

    @staticmethod
    def _main_menu_report_label(lang: str) -> str:
        return t("main_menu_reports", lang)

    @staticmethod
    def _main_menu_settings_label(lang: str) -> str:
        return t("main_menu_settings", lang)

    @staticmethod
    def _main_menu_app_label(lang: str) -> str:
        return t("main_menu_app", lang)

    @staticmethod
    def _language_picker_keyboard(
        *,
        callback_prefix: str,
        ui_lang: str,
        current_lang: str | None = None,
        back_callback: str | None = None,
    ) -> InlineKeyboardMarkup:
        current = normalize_language(current_lang)
        rows: list[list[InlineKeyboardButton]] = []
        for code in SUPPORTED_LANGUAGES:
            label = language_flag_label(code)
            if code == current and callback_prefix != "lang:init":
                label = f"{label} ✅"
            rows.append(
                [InlineKeyboardButton(label, callback_data=f"{callback_prefix}:{code}")]
            )
        if back_callback:
            rows.append([InlineKeyboardButton(t("back", ui_lang), callback_data=back_callback)])
        return InlineKeyboardMarkup(rows)

    @staticmethod
    def _language_confirm_keyboard(*, ui_lang: str, target_lang: str) -> InlineKeyboardMarkup:
        target = normalize_language(target_lang)
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton(t("confirm_yes", ui_lang), callback_data=f"lang:confirm:{target}")],
                [InlineKeyboardButton(t("confirm_no", ui_lang), callback_data="lang:cancel")],
            ]
        )

    @staticmethod
    def _assistant_language_screen_text(*, ui_lang: str, current_lang: str) -> str:
        return (
            f"{t('assistant_language_title', ui_lang)}\n\n"
            f"{t('language_current', ui_lang, language=language_short_label(current_lang))}\n"
            f"{t('assistant_language_note', ui_lang)}"
        )

    @staticmethod
    def _set_pending_start_after_language(
        context: ContextTypes.DEFAULT_TYPE,
        *,
        chat_id: int,
        chat_type: str,
        start_payload: str,
    ) -> None:
        context.user_data[PENDING_START_AFTER_LANGUAGE_KEY] = {
            "chat_id": int(chat_id),
            "chat_type": str(chat_type or ""),
            "start_payload": str(start_payload or ""),
        }

    @staticmethod
    def _pop_pending_start_after_language(
        context: ContextTypes.DEFAULT_TYPE,
    ) -> dict[str, object] | None:
        raw = context.user_data.pop(PENDING_START_AFTER_LANGUAGE_KEY, None)
        return raw if isinstance(raw, dict) else None

    @staticmethod
    def _store_pending_language_change(
        context: ContextTypes.DEFAULT_TYPE,
        *,
        source: str,
        target_lang: str,
        return_callback: str,
    ) -> None:
        context.user_data[PENDING_LANGUAGE_CHANGE_KEY] = {
            "source": str(source or ""),
            "target_lang": normalize_language(target_lang),
            "return_callback": str(return_callback or "ast:language"),
        }

    @staticmethod
    def _get_pending_language_change(
        context: ContextTypes.DEFAULT_TYPE,
    ) -> dict[str, object] | None:
        raw = context.user_data.get(PENDING_LANGUAGE_CHANGE_KEY)
        return raw if isinstance(raw, dict) else None

    @staticmethod
    def _clear_pending_language_change(context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop(PENDING_LANGUAGE_CHANGE_KEY, None)

    def _register_handlers(self) -> None:
        self.application.add_handler(CommandHandler("start", self.start_command))
        self.application.add_handler(CommandHandler("help", self.help_command))
        self.application.add_handler(CommandHandler("admin", self.admin_command))
        self.application.add_handler(self._support_conversation_handler())
        self.application.add_handler(CommandHandler("profile", self.profile_command))
        self.application.add_handler(self._profile_conversation_handler())
        self.application.add_handler(self._bot_review_conversation_handler())
        self.application.add_handler(
            CallbackQueryHandler(self.profile_menu_callback, pattern=r"^profile_")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.assistant_settings_callback_handler, pattern=r"^ast:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.language_callback_handler, pattern=r"^lang:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.start_mode_callback_handler, pattern=r"^mode_(?:personal|family)$")
        )
        self.application.add_handler(CommandHandler("app", self.app_command))
        self.application.add_handler(CommandHandler("miniapp", self.app_command))
        self.application.add_handler(CommandHandler("name", self.name_command))
        self.application.add_handler(CommandHandler("register", self.register_command))
        self.application.add_handler(CommandHandler("whoami", self.whoami_command))
        self.application.add_handler(CommandHandler("members", self.members_command))
        self.application.add_handler(CommandHandler("stats", self.stats_command))
        self.application.add_handler(CommandHandler("mystats", self.mystats_command))
        self.application.add_handler(
            CommandHandler("familystats", self.familystats_command)
        )
        self.application.add_handler(CommandHandler("reporttime", self.report_time_command))
        self.application.add_handler(CommandHandler("report", self.report_command))
        self.application.add_handler(
            CallbackQueryHandler(self.report_callback_handler, pattern=r"^rp:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.onboarding_callback_handler, pattern=r"^onb:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.main_menu_callback_handler, pattern=r"^mn:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.guess_callback_handler, pattern=r"^gc:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.transfer_callback_handler, pattern=r"^tf:")
        )
        self.application.add_handler(
            CallbackQueryHandler(self.transaction_callback_handler, pattern=r"^tx:")
        )

        self.application.add_handler(
            MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, self.new_members_handler)
        )
        self.application.add_handler(MessageHandler(filters.VOICE, self.handle_voice))
        self.application.add_handler(
            MessageHandler(filters.PHOTO | filters.Document.ALL, self.handle_photo)
        )
        self.application.add_handler(
            MessageHandler(filters.TEXT & (~filters.COMMAND), self.handle_text)
        )
        self.application.add_error_handler(self.on_error)

    def run(self) -> None:
        self.application.run_polling(drop_pending_updates=True)

    @staticmethod
    def _safe_zoneinfo(timezone_name: str):
        raw_name = str(timezone_name or DEFAULT_REPORT_TIMEZONE).strip() or DEFAULT_REPORT_TIMEZONE
        try:
            return ZoneInfo(raw_name)
        except (ZoneInfoNotFoundError, ValueError):
            logger.warning("Unknown timezone for scheduled report, fallback to %s: %s", DEFAULT_REPORT_TIMEZONE, raw_name)
        try:
            return ZoneInfo(DEFAULT_REPORT_TIMEZONE)
        except (ZoneInfoNotFoundError, ValueError):
            return timezone(timedelta(hours=5))

    async def _scheduled_report_loop(self) -> None:
        try:
            await asyncio.sleep(SCHEDULED_REPORT_BOOT_DELAY_SECONDS)
            while True:
                try:
                    await self._run_scheduled_report_tick()
                except asyncio.CancelledError:
                    raise
                except Exception:
                    logger.exception("Scheduled report tick failed")
                await asyncio.sleep(SCHEDULED_REPORT_TICK_SECONDS)
        except asyncio.CancelledError:
            logger.info("Scheduled report loop stopped")
            raise

    async def _run_scheduled_report_tick(self) -> None:
        if self._scheduled_report_tick_in_progress:
            return
        self._scheduled_report_tick_in_progress = True
        try:
            now_utc = datetime.now(timezone.utc)
            await self._dispatch_scheduled_reports(report_type="daily", now_utc=now_utc)
            await self._dispatch_scheduled_reports(report_type="weekly", now_utc=now_utc)
            await self._dispatch_scheduled_reports(report_type="monthly", now_utc=now_utc)
        finally:
            self._scheduled_report_tick_in_progress = False

    async def _dispatch_scheduled_reports(
        self,
        *,
        report_type: str,
        now_utc: datetime,
    ) -> None:
        targets = self.db.list_scheduled_report_targets(report_type=report_type)
        for target in targets:
            if not bool(target.get("enabled")):
                continue
            try:
                await self._maybe_send_scheduled_report(
                    target=target,
                    report_type=report_type,
                    now_utc=now_utc,
                )
            except asyncio.CancelledError:
                raise
            except Exception:
                logger.exception(
                    "Failed to dispatch scheduled report: workspace_id=%s report_type=%s",
                    target.get("workspace_id"),
                    report_type,
                )

    def _scheduled_report_due_context(
        self,
        *,
        target: dict[str, object],
        report_type: str,
        now_utc: datetime,
    ) -> dict[str, object] | None:
        chat_id = int(target.get("chat_id") or 0)
        workspace_id = int(target.get("workspace_id") or 0)
        if chat_id == 0 or workspace_id <= 0:
            return None

        tzinfo = self._safe_zoneinfo(str(target.get("timezone") or DEFAULT_REPORT_TIMEZONE))
        local_now = now_utc.astimezone(tzinfo)
        send_hour = max(0, min(23, int(target.get("send_hour") or DEFAULT_DAILY_REPORT_HOUR)))
        send_minute = max(0, min(59, int(target.get("send_minute") or DEFAULT_DAILY_REPORT_MINUTE)))
        scheduled_local = local_now.replace(
            hour=send_hour,
            minute=send_minute,
            second=0,
            microsecond=0,
        )
        if local_now < scheduled_local:
            return None

        period_date = local_now.date()
        period_key = ""
        period_label = ""
        period_start_local: datetime
        period_end_local: datetime

        if report_type == "daily":
            period_key = period_date.isoformat()
            period_start_local = datetime.combine(period_date, time.min, tzinfo=tzinfo)
            period_end_local = datetime.combine(period_date, time(send_hour, send_minute), tzinfo=tzinfo)
            period_label = self._scheduled_report_date_label(period_date, lang="ru")
        elif report_type == "weekly":
            scheduled_weekday = target.get("weekday")
            try:
                due_weekday = int(scheduled_weekday) if scheduled_weekday is not None else 6
            except (TypeError, ValueError):
                due_weekday = 6
            due_weekday = max(0, min(6, due_weekday))
            if local_now.weekday() != due_weekday:
                return None
            week_start_date = period_date - timedelta(days=period_date.weekday())
            iso_year, iso_week, _ = period_date.isocalendar()
            period_key = f"{int(iso_year):04d}-W{int(iso_week):02d}"
            period_start_local = datetime.combine(week_start_date, time.min, tzinfo=tzinfo)
            period_end_local = datetime.combine(period_date, time(send_hour, send_minute), tzinfo=tzinfo)
            period_label = f"{week_start_date:%d.%m.%Y} — {period_date:%d.%m.%Y}"
        elif report_type == "monthly":
            configured_monthday = target.get("monthday")
            try:
                monthday = int(configured_monthday) if configured_monthday is not None else 31
            except (TypeError, ValueError):
                monthday = 31
            monthday = max(1, min(31, monthday))
            last_day = calendar.monthrange(period_date.year, period_date.month)[1]
            due_monthday = min(monthday, last_day)
            if period_date.day != due_monthday:
                return None
            period_key = f"{period_date.year:04d}-{period_date.month:02d}"
            month_start = period_date.replace(day=1)
            period_start_local = datetime.combine(month_start, time.min, tzinfo=tzinfo)
            period_end_local = datetime.combine(period_date, time(send_hour, send_minute), tzinfo=tzinfo)
            period_label = f"{self._report_month_name(period_date.month, lang='ru')} {period_date.year}"
        else:
            return None

        if self.db.has_scheduled_report_delivery(
            workspace_id=workspace_id,
            report_type=report_type,
            period_key=period_key,
        ):
            return None

        return {
            "period_key": period_key,
            "period_date": period_date,
            "period_label": period_label,
            "local_now": local_now,
            "scheduled_local": scheduled_local,
            "period_start_utc": period_start_local.astimezone(timezone.utc).replace(tzinfo=None),
            "period_end_utc": period_end_local.astimezone(timezone.utc).replace(tzinfo=None),
        }

    async def _maybe_send_scheduled_report(
        self,
        *,
        target: dict[str, object],
        report_type: str,
        now_utc: datetime,
    ) -> None:
        due = self._scheduled_report_due_context(target=target, report_type=report_type, now_utc=now_utc)
        if due is None:
            return

        message_text = self._build_scheduled_report_message(
            target=target,
            report_type=report_type,
            due=due,
        )
        if not message_text:
            return

        chat_id = int(target.get("chat_id") or 0)
        workspace_id = int(target.get("workspace_id") or 0)
        period_key = str(due.get("period_key") or "")

        await self.application.bot.send_message(chat_id=chat_id, text=message_text)

        recorded = self.db.record_scheduled_report_delivery(
            workspace_id=workspace_id,
            chat_id=chat_id,
            report_type=report_type,
            period_key=period_key,
            sent_at_utc=now_utc.replace(tzinfo=None),
        )
        if not recorded:
            logger.info(
                "Scheduled report delivery was already recorded: workspace_id=%s report_type=%s period_key=%s",
                workspace_id,
                report_type,
                period_key,
            )

    def _scheduled_report_owner_user_id(self, target: dict[str, object]) -> int | None:
        for key in ("owner_user_id", "created_by"):
            raw = target.get(key)
            if raw is None:
                continue
            try:
                value = int(raw)
            except (TypeError, ValueError):
                continue
            if value > 0:
                return value
        workspace_type = str(target.get("workspace_type") or "").strip().lower()
        chat_id = int(target.get("chat_id") or 0)
        if workspace_type == "personal" and chat_id > 0:
            return chat_id
        return None

    def _scheduled_report_workspace_language(
        self,
        *,
        workspace_id: int,
        owner_user_id: int | None,
    ) -> str:
        if not owner_user_id or owner_user_id <= 0:
            return DEFAULT_LANGUAGE
        try:
            return normalize_language(
                self.db.get_effective_workspace_language(
                    workspace_id=int(workspace_id),
                    telegram_user_id=int(owner_user_id),
                )
            )
        except Exception:
            logger.exception(
                "Failed to load workspace language for scheduled report: workspace_id=%s user_id=%s",
                workspace_id,
                owner_user_id,
            )
            return DEFAULT_LANGUAGE

    def _scheduled_report_preferred_currency(self, owner_user_id: int | None) -> str:
        if owner_user_id and owner_user_id > 0:
            try:
                profile = self.db.get_user_profile(int(owner_user_id))
                currency = str(profile.get("currency") or "").strip().upper()
                if currency:
                    return currency
            except Exception:
                logger.exception("Failed to load preferred currency for scheduled report: user_id=%s", owner_user_id)
        return str(self.settings.default_currency or "UZS").strip().upper() or "UZS"

    def _scheduled_report_date_label(self, target_date: date, *, lang: str = DEFAULT_LANGUAGE) -> str:
        ui_lang = normalize_language(lang)
        if ui_lang == "en":
            month_name = self._report_month_name(target_date.month, lang=ui_lang)
            return f"{month_name} {target_date.day}, {target_date.year}"
        month_name = self._report_month_name(target_date.month, lang=ui_lang, genitive=True)
        return f"{target_date.day} {month_name} {target_date.year}"

    @staticmethod
    def _scheduled_breakdown_has_entries(
        breakdown: dict[str, dict[str, dict[str, object]]],
    ) -> bool:
        for kind in ("expense", "income"):
            by_currency = breakdown.get(kind, {})
            if not isinstance(by_currency, dict):
                continue
            for bucket in by_currency.values():
                if not isinstance(bucket, dict):
                    continue
                if float(bucket.get("total", 0.0)) > 0:
                    return True
        return False

    def _scheduled_breakdown_snapshot(
        self,
        breakdown: dict[str, dict[str, dict[str, object]]],
        *,
        preferred_currency: str,
    ) -> dict[str, object]:
        currencies: list[str] = []
        for kind in ("expense", "income"):
            by_currency = breakdown.get(kind, {})
            if not isinstance(by_currency, dict):
                continue
            for currency_code, bucket in by_currency.items():
                if not isinstance(bucket, dict):
                    continue
                if float(bucket.get("total", 0.0)) <= 0:
                    continue
                code = str(currency_code or "").upper()
                if code and code not in currencies:
                    currencies.append(code)

        preferred = str(preferred_currency or "").strip().upper()
        selected_currency = ""
        if preferred and preferred in currencies:
            selected_currency = preferred
        elif currencies:
            selected_currency = sorted(currencies)[0]
        else:
            selected_currency = preferred or (str(self.settings.default_currency or "UZS").strip().upper() or "UZS")

        expense_bucket = breakdown.get("expense", {}).get(selected_currency, {})
        income_bucket = breakdown.get("income", {}).get(selected_currency, {})
        if not isinstance(expense_bucket, dict):
            expense_bucket = {}
        if not isinstance(income_bucket, dict):
            income_bucket = {}

        expense_categories = expense_bucket.get("categories", [])
        income_categories = income_bucket.get("categories", [])
        if not isinstance(expense_categories, list):
            expense_categories = []
        if not isinstance(income_categories, list):
            income_categories = []

        return {
            "selected_currency": selected_currency,
            "expense_total": float(expense_bucket.get("total", 0.0)),
            "income_total": float(income_bucket.get("total", 0.0)),
            "expense_categories": expense_categories,
            "income_categories": income_categories,
            "has_multiple_currencies": len(currencies) > 1,
            "all_currencies": sorted(currencies),
        }

    def _scheduled_family_member_lines(
        self,
        user_summaries: list[dict[str, object]],
        *,
        currency: str,
    ) -> list[str]:
        lines: list[str] = []
        currency_code = str(currency or "").upper()
        for item in user_summaries:
            name = str(item.get("name") or "Участник").strip() or "Участник"
            totals = item.get("totals")
            if not isinstance(totals, dict):
                totals = {}
            stat = totals.get(currency_code, {})
            if not isinstance(stat, dict):
                stat = {}
            expense = float(stat.get("expense", 0.0))
            income = float(stat.get("income", 0.0))
            # Keep the list concise and relevant to the selected currency.
            if expense <= 0 and income <= 0:
                continue
            lines.append(
                f"• {name}: расходы {_fmt_money(expense, currency_code)} / доходы {_fmt_money(income, currency_code)}"
            )
        return lines

    def _build_scheduled_report_message(
        self,
        *,
        target: dict[str, object],
        report_type: str,
        due: dict[str, object],
    ) -> str | None:
        if report_type not in {"daily", "weekly", "monthly"}:
            return None

        workspace_type = str(target.get("workspace_type") or "").strip().lower()
        if workspace_type not in {"personal", "family"}:
            return None

        workspace_id = int(target.get("workspace_id") or 0)
        chat_id = int(target.get("chat_id") or 0)
        if workspace_id <= 0 or chat_id == 0:
            return None

        owner_user_id = self._scheduled_report_owner_user_id(target)
        preferred_currency = self._scheduled_report_preferred_currency(owner_user_id)

        period_start_utc = due.get("period_start_utc")
        period_end_utc = due.get("period_end_utc")
        period_date = due.get("period_date")
        if not isinstance(period_start_utc, datetime) or not isinstance(period_end_utc, datetime):
            return None
        if not isinstance(period_date, date):
            return None

        target_user_id = owner_user_id if workspace_type == "personal" else None
        transactions = self.db.get_period_transactions(
            chat_id,
            period_start_utc,
            period_end_utc,
            workspace_id=workspace_id,
            telegram_user_id=target_user_id,
        )
        breakdown = self._build_breakdown_from_transactions(transactions)
        has_entries = self._scheduled_breakdown_has_entries(breakdown)
        period_label_raw = str(due.get("period_label") or "").strip()
        period_label = period_label_raw or self._scheduled_report_date_label(period_date, lang="ru")

        if report_type == "daily":
            personal_title = f"📊 Ежедневный отчет за {period_label}"
            family_title = f"👨‍👩‍👧‍👦 Семейный отчет за {period_label}"
            personal_no_data_lines = [
                f"🔔 Ежедневный отчет за {period_label}",
                "",
                "Сегодня вы не добавили ни расходов, ни доходов.",
                "Похоже, забыли внести записи за день.",
                "",
                "Добавьте расходы/доходы сейчас, чтобы отчет был полным ✅",
            ]
            family_no_data_lines = [
                f"🔔 Семейный отчет за {period_label}",
                "",
                "Сегодня в группе не добавили ни расходов, ни доходов.",
                "Похоже, записи за день забыли внести.",
                "",
                "Добавьте расходы/доходы, чтобы семейный учет был актуальным ✅",
            ]
            personal_expense_line = "💸 Расходы за сегодня: {amount}"
            personal_income_line = "💰 Доходы за сегодня: {amount}"
            family_expense_line = "💸 Общие расходы за сегодня: {amount}"
            family_income_line = "💰 Общие доходы за сегодня: {amount}"
            reminder_line = "Если что-то забыли — добавьте сейчас, пока день не закончился ✅"
            wish_line = None
        elif report_type == "weekly":
            personal_title = f"📅 Еженедельный отчет за {period_label}"
            family_title = f"👨‍👩‍👧‍👦 Семейный еженедельный отчет за {period_label}"
            personal_no_data_lines = [
                f"🔔 Еженедельный отчет за {period_label}",
                "",
                "На этой неделе вы не добавили ни расходов, ни доходов.",
                "Похоже, забыли внести записи.",
                "",
                "Добавьте расходы/доходы сегодня, чтобы отчет за неделю был полным ✅",
                "",
                "Удачи на следующей неделе 🍀",
            ]
            family_no_data_lines = [
                f"🔔 Семейный еженедельный отчет за {period_label}",
                "",
                "На этой неделе в группе не добавили ни расходов, ни доходов.",
                "Похоже, записи забыли внести.",
                "",
                "Добавьте расходы/доходы сегодня, чтобы недельный учет был актуальным ✅",
                "",
                "Удачи на следующей неделе 🍀",
            ]
            personal_expense_line = "💸 Расходы за неделю: {amount}"
            personal_income_line = "💰 Доходы за неделю: {amount}"
            family_expense_line = "💸 Общие расходы за неделю: {amount}"
            family_income_line = "💰 Общие доходы за неделю: {amount}"
            reminder_line = "Если что-то забыли за эту неделю — добавьте сегодня, пока день не закончился ✅"
            wish_line = "Удачи на следующей неделе 🍀"
        else:  # monthly
            personal_title = f"📆 Ежемесячный отчет за {period_label}"
            family_title = f"👨‍👩‍👧‍👦 Семейный ежемесячный отчет за {period_label}"
            personal_no_data_lines = [
                f"🔔 Ежемесячный отчет за {period_label}",
                "",
                "В этом месяце вы не добавили ни расходов, ни доходов.",
                "Похоже, забыли внести записи.",
                "",
                "Добавьте расходы/доходы сегодня, чтобы отчет за месяц был полным ✅",
                "",
                "Удачи в следующем месяце 🍀",
            ]
            family_no_data_lines = [
                f"🔔 Семейный ежемесячный отчет за {period_label}",
                "",
                "В этом месяце в группе не добавили ни расходов, ни доходов.",
                "Похоже, записи забыли внести.",
                "",
                "Добавьте расходы/доходы сегодня, чтобы месячный учет был актуальным ✅",
                "",
                "Удачи в следующем месяце 🍀",
            ]
            personal_expense_line = "💸 Расходы за месяц: {amount}"
            personal_income_line = "💰 Доходы за месяц: {amount}"
            family_expense_line = "💸 Общие расходы за месяц: {amount}"
            family_income_line = "💰 Общие доходы за месяц: {amount}"
            reminder_line = "Если что-то забыли за этот месяц — добавьте сегодня, пока день не закончился ✅"
            wish_line = "Удачи в следующем месяце 🍀"

        if not has_entries:
            if workspace_type == "personal":
                return "\n".join(personal_no_data_lines)
            return "\n".join(family_no_data_lines)

        snapshot = self._scheduled_breakdown_snapshot(
            breakdown,
            preferred_currency=preferred_currency,
        )
        currency = str(snapshot.get("selected_currency") or preferred_currency or "UZS")
        expense_total = float(snapshot.get("expense_total", 0.0))
        income_total = float(snapshot.get("income_total", 0.0))
        expense_categories = list(snapshot.get("expense_categories") or [])
        income_categories = list(snapshot.get("income_categories") or [])
        has_multiple_currencies = bool(snapshot.get("has_multiple_currencies"))

        if workspace_type == "personal":
            lines = [
                personal_title,
                "",
                personal_expense_line.format(amount=_fmt_money(expense_total, currency)),
                personal_income_line.format(amount=_fmt_money(income_total, currency)),
            ]

            if expense_categories:
                lines.extend(["", "Расходы по категориям:"])
                for category, amount in expense_categories[:3]:
                    _, label = self._report_category_view(str(category), "expense")
                    lines.append(f"• {label}: {_fmt_money(float(amount), currency)}")

            if income_categories:
                lines.extend(["", "Доходы по категориям:"])
                for category, amount in income_categories[:2]:
                    _, label = self._report_category_view(str(category), "income")
                    lines.append(f"• {label}: {_fmt_money(float(amount), currency)}")

            if has_multiple_currencies:
                lines.extend(["", f"ℹ️ Есть записи и в других валютах (показана {currency})."])
            lines.extend(["", reminder_line])
            if wish_line:
                lines.extend(["", wish_line])
            return "\n".join(lines)

        user_summaries = self.db.get_group_user_summaries(
            chat_id,
            period_start_utc,
            period_end_utc,
            workspace_id=workspace_id,
        )
        member_lines = self._scheduled_family_member_lines(user_summaries, currency=currency)

        lines = [
            family_title,
            "",
            family_expense_line.format(amount=_fmt_money(expense_total, currency)),
            family_income_line.format(amount=_fmt_money(income_total, currency)),
        ]
        if member_lines:
            lines.extend(["", "По участникам:"])
            lines.extend(member_lines)
        if has_multiple_currencies:
            lines.extend(["", f"ℹ️ Есть записи и в других валютах (показана {currency})."])
        lines.extend(["", reminder_line])
        if wish_line:
            lines.extend(["", wish_line])
        return "\n".join(lines)

    @staticmethod
    def _format_hhmm(hour: int, minute: int) -> str:
        return f"{max(0, min(23, int(hour))):02d}:{max(0, min(59, int(minute))):02d}"

    async def report_time_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return

        try:
            self.db.register_user(
                telegram_user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                last_name=user.last_name,
            )
        except Exception:
            logger.exception("Failed to register user before /reporttime: user_id=%s", user.id)

        resolved = self.db.resolve_workspace(user.id, chat.id, chat.type)
        if resolved is None:
            if chat.type == "private":
                await message.reply_text("Сначала запустите бота в этом чате через /start.")
            else:
                await message.reply_text("Сначала активируйте бота в этой группе через /start.")
            return

        workspace_id = int(resolved.get("workspace_id") or 0)
        workspace_type = str(resolved.get("type") or "").strip().lower()
        member_role = str(resolved.get("member_role") or "").strip().lower()
        if workspace_type == "family" and member_role != "owner":
            await message.reply_text("Менять время ежедневного отчета в группе может только владелец семейного пространства.")
            return

        current = self.db.get_workspace_report_schedule(
            workspace_id=workspace_id,
            report_type="daily",
            ensure_default=True,
        )
        args = [str(arg).strip() for arg in (context.args or []) if str(arg).strip()]
        if not args:
            state_label = "включен" if bool(current.get("enabled")) else "выключен"
            current_time = self._format_hhmm(
                int(current.get("send_hour") or DEFAULT_DAILY_REPORT_HOUR),
                int(current.get("send_minute") or DEFAULT_DAILY_REPORT_MINUTE),
            )
            await message.reply_text(
                "Ежедневный отчет сейчас "
                f"{state_label}: {current_time} ({DEFAULT_REPORT_TIMEZONE}).\n"
                "Отчет считается с 00:00 до времени отправки.\n"
                "Изменить время: /reporttime HH:MM\n"
                "Выключить: /reporttime off\n"
                "Включить обратно: /reporttime on"
            )
            return

        raw_value = args[0].lower()
        if raw_value in {"off", "disable", "выкл"}:
            updated = self.db.upsert_workspace_report_schedule(
                workspace_id=workspace_id,
                report_type="daily",
                enabled=False,
                timezone=str(current.get("timezone") or DEFAULT_REPORT_TIMEZONE),
                send_hour=int(current.get("send_hour") or DEFAULT_DAILY_REPORT_HOUR),
                send_minute=int(current.get("send_minute") or DEFAULT_DAILY_REPORT_MINUTE),
                weekday=current.get("weekday") if isinstance(current.get("weekday"), int) else None,
                monthday=current.get("monthday") if isinstance(current.get("monthday"), int) else None,
            )
            await message.reply_text(
                "Ежедневный отчет выключен.\n"
                f"Сохраненное время: {self._format_hhmm(int(updated.get('send_hour') or 0), int(updated.get('send_minute') or 0))} ({DEFAULT_REPORT_TIMEZONE})."
            )
            return

        if raw_value in {"on", "enable", "вкл"}:
            updated = self.db.upsert_workspace_report_schedule(
                workspace_id=workspace_id,
                report_type="daily",
                enabled=True,
                timezone=str(current.get("timezone") or DEFAULT_REPORT_TIMEZONE),
                send_hour=int(current.get("send_hour") or DEFAULT_DAILY_REPORT_HOUR),
                send_minute=int(current.get("send_minute") or DEFAULT_DAILY_REPORT_MINUTE),
                weekday=current.get("weekday") if isinstance(current.get("weekday"), int) else None,
                monthday=current.get("monthday") if isinstance(current.get("monthday"), int) else None,
            )
            await message.reply_text(
                "Ежедневный отчет включен.\n"
                f"Время отправки: {self._format_hhmm(int(updated.get('send_hour') or 0), int(updated.get('send_minute') or 0))} ({DEFAULT_REPORT_TIMEZONE})."
            )
            return

        if not REPORT_TIME_INPUT_RE.fullmatch(raw_value):
            await message.reply_text(
                "Неверный формат времени.\n"
                "Используйте: /reporttime HH:MM (например: /reporttime 21:00)\n"
                "Также можно: /reporttime off или /reporttime on"
            )
            return

        hour_s, minute_s = raw_value.split(":", 1)
        updated = self.db.upsert_workspace_report_schedule(
            workspace_id=workspace_id,
            report_type="daily",
            enabled=True,
            timezone=DEFAULT_REPORT_TIMEZONE,
            send_hour=int(hour_s),
            send_minute=int(minute_s),
            weekday=current.get("weekday") if isinstance(current.get("weekday"), int) else None,
            monthday=current.get("monthday") if isinstance(current.get("monthday"), int) else None,
        )
        time_value = self._format_hhmm(
            int(updated.get("send_hour") or DEFAULT_DAILY_REPORT_HOUR),
            int(updated.get("send_minute") or DEFAULT_DAILY_REPORT_MINUTE),
        )
        await message.reply_text(
            "Время ежедневного отчета обновлено.\n"
            f"Теперь отправка будет в {time_value} ({DEFAULT_REPORT_TIMEZONE}).\n"
            "Отчет считается с 00:00 до времени отправки."
        )

    def _support_conversation_handler(self) -> ConversationHandler:
        return ConversationHandler(
            entry_points=[
                CommandHandler("support", self.support_command),
                CallbackQueryHandler(self.support_menu_callback, pattern=r"^sup:(message|bug)$"),
                CallbackQueryHandler(
                    self.support_direct_callback,
                    pattern=r"^(profile_support_message|profile_support_bug|ast:support_message|ast:support_bug)$",
                ),
            ],
            states={
                SUPPORT_MENU: [
                    CallbackQueryHandler(self.support_menu_callback, pattern=r"^sup:")
                ],
                SUPPORT_MESSAGE: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.support_message_handler,
                    ),
                    CallbackQueryHandler(self.support_menu_callback, pattern=r"^sup:")
                ],
                SUPPORT_BUG: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.support_bug_text_handler,
                    ),
                    MessageHandler(filters.PHOTO, self.support_bug_photo_handler),
                    CallbackQueryHandler(self.support_menu_callback, pattern=r"^sup:")
                ],
            },
            fallbacks=[CallbackQueryHandler(self.support_menu_callback, pattern=r"^sup:")],
            allow_reentry=True,
            per_chat=True,
            per_user=True,
        )

    def _bot_review_conversation_handler(self) -> ConversationHandler:
        return ConversationHandler(
            entry_points=[CallbackQueryHandler(self.bot_review_start_callback, pattern=r"^ast:rate$")],
            states={
                BOT_REVIEW_MENU: [
                    CallbackQueryHandler(
                        self.bot_review_menu_callback,
                        pattern=r"^rv:(?:change|comment|cancel)$",
                    )
                ],
                BOT_REVIEW_RATING: [
                    CallbackQueryHandler(self.bot_review_rating_callback, pattern=r"^rv:(?:rate:[1-5]|cancel)$")
                ],
                BOT_REVIEW_COMMENT: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.bot_review_comment_handler,
                    ),
                    CallbackQueryHandler(
                        self.bot_review_comment_callback,
                        pattern=r"^rv:(?:skip|cancel)$",
                    ),
                ],
            },
            fallbacks=[
                CallbackQueryHandler(self.bot_review_cancel_callback, pattern=r"^rv:cancel$")
            ],
            allow_reentry=True,
            per_chat=True,
            per_user=True,
        )

    def _profile_conversation_handler(self) -> ConversationHandler:
        return ConversationHandler(
            entry_points=[
                CallbackQueryHandler(self.profile_edit_start_callback, pattern=r"^edit_profile$")
            ],
            states={
                PROFILE_NAME: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.profile_name_handler,
                    )
                ],
                PROFILE_PHONE: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.profile_phone_handler,
                    )
                ],
                PROFILE_EMAIL: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.profile_email_handler,
                    )
                ],
                PROFILE_BIRTHDATE: [
                    MessageHandler(
                        filters.TEXT & (~filters.COMMAND),
                        self.profile_birth_handler,
                    )
                ],
            },
            fallbacks=[
                CallbackQueryHandler(self.profile_edit_cancel_callback, pattern=r"^profile_cancel$"),
                CallbackQueryHandler(self.profile_wizard_interrupt_callback, pattern=r"^ast:"),
            ],
            allow_reentry=True,
            per_chat=True,
            per_user=True,
        )

    @staticmethod
    def _profile_menu_keyboard(
        *,
        lang: str = DEFAULT_LANGUAGE,
        exit_callback: str = "profile_close",
        exit_text: str | None = None,
    ) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton(t("assistant_settings_profile", ui_lang), callback_data="edit_profile")],
                [
                    InlineKeyboardButton(
                        {
                            "ru": "⚙️ Настройки",
                            "uz": "⚙️ Sozlamalar",
                            "en": "⚙️ Settings",
                        }.get(ui_lang, "⚙️ Settings"),
                        callback_data="profile_settings",
                    )
                ],
                [InlineKeyboardButton(t("assistant_settings_support", ui_lang), callback_data="profile_support")],
                [InlineKeyboardButton(exit_text or t("close", ui_lang), callback_data=exit_callback)],
            ]
        )

    @staticmethod
    def _profile_wizard_cancel_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton(t("cancel", lang), callback_data="profile_cancel")]]
        )

    @staticmethod
    def _profile_back_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton(t("back", lang), callback_data="profile_back")]]
        )

    def _profile_settings_keyboard(
        self,
        *,
        lang: str = DEFAULT_LANGUAGE,
        current_language: str = DEFAULT_LANGUAGE,
        current_currency: str = "UZS",
    ) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        t(
                            "profile_settings_language_row",
                            ui_lang,
                            language=language_short_label(current_language),
                        ),
                        callback_data="profile_settings_lang",
                    )
                ],
                [
                    InlineKeyboardButton(
                        t(
                            "profile_settings_currency_row",
                            ui_lang,
                            currency=str(current_currency or "UZS").upper(),
                        ),
                        callback_data="profile_settings_currency",
                    )
                ],
                [InlineKeyboardButton(t("back", ui_lang), callback_data="profile_back")],
            ]
        )

    @staticmethod
    def _profile_support_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton(t("profile_support_dev", lang), callback_data="sup:message")],
                [InlineKeyboardButton(t("profile_support_bug", lang), callback_data="sup:bug")],
                [InlineKeyboardButton(t("back", lang), callback_data="profile_back")],
            ]
        )

    @staticmethod
    def _profile_progress_dots(completed: int, total: int) -> str:
        completed_safe = max(0, min(completed, total))
        return ("◉" * completed_safe) + ("◯" * max(0, total - completed_safe))

    @staticmethod
    def _profile_completion_from_profile(profile: dict[str, object]) -> tuple[int, int]:
        tracked_fields = ("display_name", "phone", "email", "birth_date", "currency")
        completed = 0
        for field in tracked_fields:
            value = profile.get(field)
            if isinstance(value, str):
                if value.strip():
                    completed += 1
            elif value is not None:
                completed += 1
        return completed, len(tracked_fields)

    @staticmethod
    def _profile_display_birth(value: str | None) -> str:
        if not value:
            return "—"
        try:
            dt = datetime.strptime(value, "%Y-%m-%d")
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            return value

    def _format_profile_menu_text(
        self,
        profile: dict,
        completed: int,
        total: int,
        *,
        lang: str = DEFAULT_LANGUAGE,
    ) -> str:
        ui_lang = normalize_language(lang)
        currency = str(profile.get("currency") or "UZS").upper()
        if ui_lang == "uz":
            return (
                "👤 Profil\n\n"
                f"Ism: {profile.get('display_name') or '—'}\n"
                f"Telefon: {profile.get('phone') or '—'}\n"
                f"Email: {profile.get('email') or '—'}\n"
                f"Tug'ilgan sana: {self._profile_display_birth(profile.get('birth_date'))}\n\n"
                f"Valyuta: {currency}\n\n"
                f"To'ldirilgan: {completed}/{total}\n"
                f"{self._profile_progress_dots(completed, total)}"
            )
        if ui_lang == "en":
            return (
                "👤 Profile\n\n"
                f"Name: {profile.get('display_name') or '—'}\n"
                f"Phone: {profile.get('phone') or '—'}\n"
                f"Email: {profile.get('email') or '—'}\n"
                f"Birth date: {self._profile_display_birth(profile.get('birth_date'))}\n\n"
                f"Currency: {currency}\n\n"
                f"Completed: {completed}/{total}\n"
                f"{self._profile_progress_dots(completed, total)}"
            )
        return (
            "👤 Профиль\n\n"
            f"Имя: {profile.get('display_name') or '—'}\n"
            f"Телефон: {profile.get('phone') or '—'}\n"
            f"Email: {profile.get('email') or '—'}\n"
            f"Дата рождения: {self._profile_display_birth(profile.get('birth_date'))}\n\n"
            f"Валюта: {currency}\n\n"
            f"Заполнено: {completed}/{total}\n"
            f"{self._profile_progress_dots(completed, total)}"
        )

    @staticmethod
    def _assistant_settings_text(lang: str = DEFAULT_LANGUAGE) -> str:
        return t("assistant_settings_title", lang)

    @staticmethod
    def _assistant_settings_keyboard(
        *, lang: str = DEFAULT_LANGUAGE, back_callback: str | None = None
    ) -> InlineKeyboardMarkup:
        rows = [
            [InlineKeyboardButton(t("assistant_settings_profile", lang), callback_data="ast:profile")],
            [InlineKeyboardButton(t("assistant_settings_language", lang), callback_data="ast:language")],
            [InlineKeyboardButton(t("assistant_settings_currency", lang), callback_data="ast:currency")],
            [InlineKeyboardButton(t("assistant_settings_rate", lang), callback_data="ast:rate")],
            [InlineKeyboardButton(t("assistant_settings_support", lang), callback_data="ast:support")],
        ]
        rows.append(
            [InlineKeyboardButton(t("back", lang), callback_data=back_callback)]
            if back_callback
            else [InlineKeyboardButton(t("close", lang), callback_data="ast:close")]
        )
        return InlineKeyboardMarkup(rows)

    @staticmethod
    def _assistant_settings_back_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton(t("back_to_settings", lang), callback_data="ast:menu")]]
        )

    @staticmethod
    def _bot_review_rating_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("⭐", callback_data="rv:rate:1")],
                [InlineKeyboardButton("⭐⭐", callback_data="rv:rate:2")],
                [InlineKeyboardButton("⭐⭐⭐", callback_data="rv:rate:3")],
                [InlineKeyboardButton("⭐⭐⭐⭐", callback_data="rv:rate:4")],
                [InlineKeyboardButton("⭐⭐⭐⭐⭐", callback_data="rv:rate:5")],
                [InlineKeyboardButton("❌ Отмена", callback_data="rv:cancel")],
            ]
        )

    @staticmethod
    def _bot_review_comment_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("➡️ Пропустить", callback_data="rv:skip")],
                [InlineKeyboardButton("❌ Отмена", callback_data="rv:cancel")],
            ]
        )

    @staticmethod
    def _bot_review_existing_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("Изменить оценку", callback_data="rv:change")],
                [InlineKeyboardButton("Оставить комментарий", callback_data="rv:comment")],
                [InlineKeyboardButton("❌ Отмена", callback_data="rv:cancel")],
            ]
        )

    @staticmethod
    def _assistant_support_menu_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton(t("profile_support_dev", lang), callback_data="sup:message")],
                [InlineKeyboardButton(t("profile_support_bug", lang), callback_data="sup:bug")],
                [InlineKeyboardButton(t("back_to_settings", lang), callback_data="ast:menu")],
            ]
        )

    def _profile_defaults_from_user(self, user: User | None) -> tuple[str | None, str | None, str | None]:
        if not user:
            return (None, None, None)
        return (user.username, user.first_name, user.last_name)

    @staticmethod
    def _assistant_settings_trigger_map(
        context: ContextTypes.DEFAULT_TYPE,
    ) -> dict[str, int]:
        raw = context.chat_data.get(ASSISTANT_SETTINGS_TRIGGER_MAP_KEY)
        if not isinstance(raw, dict):
            raw = {}
            context.chat_data[ASSISTANT_SETTINGS_TRIGGER_MAP_KEY] = raw
        return raw

    def _remember_assistant_settings_trigger(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        settings_message_id: int,
        user_message_id: int,
    ) -> None:
        if settings_message_id <= 0 or user_message_id <= 0:
            return
        self._assistant_settings_trigger_map(context)[str(settings_message_id)] = user_message_id

    def _pop_assistant_settings_trigger(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        settings_message_id: int,
    ) -> int | None:
        if settings_message_id <= 0:
            return None
        raw = self._assistant_settings_trigger_map(context).pop(str(settings_message_id), None)
        try:
            return int(raw) if raw is not None else None
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _looks_like_button_label_text(text: str) -> bool:
        normalized = (text or "").strip()
        blocked = {
            "⚙️ Настройки",
            "📩 Поддержка",
            "➡️ Пропустить",
        }
        for lang in SUPPORTED_LANGUAGES:
            blocked.update(
                {
                    t("main_menu_reports", lang),
                    t("main_menu_settings", lang),
                    t("main_menu_app", lang),
                    t("assistant_settings_profile", lang),
                    t("assistant_settings_support", lang),
                    t("assistant_settings_title", lang).split("\n", 1)[0],
                    t("assistant_settings_rate", lang),
                    t("close", lang),
                    t("cancel", lang),
                    t("back_to_settings", lang),
                    t("back", lang),
                }
            )
        return normalized in blocked

    async def _refresh_profile_view_message(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        user: User,
    ) -> None:
        ref = context.user_data.get(PROFILE_VIEW_REF_KEY)
        if not isinstance(ref, dict):
            return
        chat_id = int(ref.get("chat_id") or 0)
        message_id = int(ref.get("message_id") or 0)
        if chat_id == 0 or message_id == 0:
            return
        exit_callback = str(ref.get("exit_callback") or "profile_close")
        from_assistant_settings = bool(ref.get("from_assistant_settings"))
        lang = self._user_language(user.id)
        exit_text = (
            t("back_to_settings", lang) if from_assistant_settings else t("close", lang)
        )
        profile = self.db.get_user_profile(user.id)
        completed, total = self._profile_completion_from_profile(profile)
        text = self._format_profile_menu_text(profile, completed, total, lang=lang)
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=text,
                reply_markup=self._profile_menu_keyboard(
                    lang=lang,
                    exit_callback=exit_callback,
                    exit_text=exit_text,
                ),
            )
        except Exception:
            logger.exception(
                "Failed to refresh profile view message: chat_id=%s message_id=%s user_id=%s",
                chat_id,
                message_id,
                user.id,
            )

    async def _show_profile_menu_message(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        edit_query_message: bool = False,
        from_assistant_settings: bool | None = None,
    ) -> int:
        message = update.effective_message
        query = update.callback_query
        user = update.effective_user
        if not user:
            return ConversationHandler.END

        self._touch_user_safely(user.id)
        lang = self._user_language(user.id)
        try:
            self.db.register_user(
                telegram_user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                last_name=user.last_name,
            )
            profile = self.db.get_user_profile(user.id)
            completed, total = self._profile_completion_from_profile(profile)
        except Exception:
            logger.exception("Failed to load profile menu: user_id=%s", user.id)
            if message:
                await message.reply_text(
                    {
                        "ru": "Не удалось открыть профиль.",
                        "uz": "Profilni ochib bo'lmadi.",
                        "en": "Failed to open profile.",
                    }.get(lang, "Failed to open profile.")
                )
            return ConversationHandler.END

        text = self._format_profile_menu_text(profile, completed, total, lang=lang)
        prior_ref = context.user_data.get(PROFILE_VIEW_REF_KEY)
        use_assistant_footer = False
        if isinstance(prior_ref, dict) and from_assistant_settings is None:
            use_assistant_footer = bool(prior_ref.get("from_assistant_settings"))
        elif from_assistant_settings is not None:
            use_assistant_footer = bool(from_assistant_settings)

        exit_callback = "ast:menu" if use_assistant_footer else "profile_close"
        exit_text = t("back_to_settings", lang) if use_assistant_footer else t("close", lang)
        keyboard = self._profile_menu_keyboard(
            lang=lang,
            exit_callback=exit_callback,
            exit_text=exit_text,
        )
        context.user_data.pop(PROFILE_WIZARD_KEY, None)

        if edit_query_message and query and query.message:
            try:
                await query.edit_message_text(text, reply_markup=keyboard)
                context.user_data[PROFILE_VIEW_REF_KEY] = {
                    "chat_id": query.message.chat_id,
                    "message_id": query.message.message_id,
                    "from_assistant_settings": use_assistant_footer,
                    "exit_callback": exit_callback,
                    "exit_text": exit_text,
                }
            except Exception:
                logger.exception("Failed to edit profile menu message")
                sent = await query.message.reply_text(text, reply_markup=keyboard)
                context.user_data[PROFILE_VIEW_REF_KEY] = {
                    "chat_id": sent.chat_id,
                    "message_id": sent.message_id,
                    "from_assistant_settings": use_assistant_footer,
                    "exit_callback": exit_callback,
                    "exit_text": exit_text,
                }
        elif message:
            sent = await message.reply_text(text, reply_markup=keyboard)
            context.user_data[PROFILE_VIEW_REF_KEY] = {
                "chat_id": sent.chat_id,
                "message_id": sent.message_id,
                "from_assistant_settings": use_assistant_footer,
                "exit_callback": exit_callback,
                "exit_text": exit_text,
            }
        return PROFILE_MENU

    @staticmethod
    def _support_menu_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "📝 Написать разработчику", callback_data="sup:message"
                    )
                ],
                [
                    InlineKeyboardButton(
                        "🐞 Сообщить об ошибке", callback_data="sup:bug"
                    )
                ],
                [InlineKeyboardButton("❌ Отмена", callback_data="sup:cancel")],
            ]
        )

    def _support_admin_chat_id(self) -> int:
        return int(getattr(self.settings, "admin_chat_id", 0) or 0)

    @staticmethod
    def _safe_telegram_caption(text: str, *, limit: int = 1024) -> str:
        normalized = (text or "").strip()
        if len(normalized) <= limit:
            return normalized
        return normalized[: limit - 1].rstrip() + "…"

    def _touch_user_safely(self, user_id: int | None) -> None:
        if not user_id:
            return
        try:
            self.db.touch_user(int(user_id))
        except Exception:
            logger.exception("Failed to touch user activity: user_id=%s", user_id)

    def track_user_activity(self, update: Update) -> None:
        user = update.effective_user
        if not user or user.is_bot:
            return
        try:
            self.db.register_user(
                telegram_user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                last_name=user.last_name,
            )
        except Exception:
            logger.exception("Failed to track user activity: user_id=%s", user.id)

    @staticmethod
    async def _safe_query_answer(
        query,
        text: str | None = None,
        *,
        show_alert: bool = False,
    ) -> None:
        try:
            await query.answer(text=text, show_alert=show_alert)
        except Exception:
            logger.exception("Failed to answer callback query")

    async def on_error(
        self, update: object, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        logger.exception("Unhandled error: %s", context.error)
        if isinstance(update, Update) and update.effective_message:
            await update.effective_message.reply_text(
                "Сейчас что-то пошло не так. Давай попробуем еще раз."
            )

    @staticmethod
    def _local_tzinfo():
        return datetime.now().astimezone().tzinfo or timezone.utc

    @classmethod
    def _local_naive_to_utc_naive(cls, value: datetime) -> datetime:
        local_dt = value.replace(tzinfo=cls._local_tzinfo())
        return local_dt.astimezone(timezone.utc).replace(tzinfo=None)

    @staticmethod
    def _get_group_onboarding_state(
        context: ContextTypes.DEFAULT_TYPE,
    ) -> dict[str, object]:
        raw = context.chat_data.get(ONBOARDING_STATE_KEY)
        if not isinstance(raw, dict):
            raw = {}
            context.chat_data[ONBOARDING_STATE_KEY] = raw

        raw.setdefault("admin_prompt_message_id", None)
        raw.setdefault("admin_message_ids", [])
        raw.setdefault("admin_confirmed", False)
        raw.setdefault("intro_sent", False)
        raw.setdefault("final_sent", False)
        return raw

    @staticmethod
    def _track_onboarding_message_id(state: dict[str, object], message_id: int) -> None:
        if message_id <= 0:
            return
        raw_ids = state.get("admin_message_ids")
        ids = raw_ids if isinstance(raw_ids, list) else []
        if message_id not in ids:
            ids.append(message_id)
        state["admin_message_ids"] = ids

    async def _get_bot_username(self, context: ContextTypes.DEFAULT_TYPE) -> str:
        if self._bot_username:
            return self._bot_username
        try:
            me = await context.bot.get_me()
            self._bot_username = str(me.username or "").strip() or "Bot"
        except Exception:
            logger.exception("Failed to fetch bot username for start link")
            self._bot_username = "Bot"
        return self._bot_username

    async def _build_add_group_url(self, context: ContextTypes.DEFAULT_TYPE) -> str:
        username = await self._get_bot_username(context)
        return (
            f"https://t.me/{username}?startgroup=true"
            "&admin=delete_messages+restrict_members+pin_messages"
        )

    async def _build_group_miniapp_direct_link(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
    ) -> str:
        username = await self._get_bot_username(context)
        start_param = quote(f"chat_{int(chat_id)}", safe="")
        app_short_name = str(getattr(self.settings, "telegram_miniapp_short_name", "") or "").strip()
        if app_short_name:
            return f"https://t.me/{username}/{quote(app_short_name, safe='')}?startapp={start_param}"
        return f"https://t.me/{username}?startapp={start_param}"

    def _miniapp_base_url(self) -> str | None:
        base = (self.settings.miniapp_base_url or "").strip()
        if not base or not base.lower().startswith("https://"):
            return None
        return base

    async def _send_miniapp_workspace_message(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
        workspace: dict[str, object],
        text: str,
        lang: str,
    ) -> None:
        base_url = self._miniapp_base_url()
        if not base_url:
            await context.bot.send_message(chat_id=chat_id, text=text)
            await context.bot.send_message(chat_id=chat_id, text=t("miniapp_not_configured", lang))
            return
        workspace_type = str(workspace.get("type") or "").strip().lower()
        # Telegram Bot API rejects InlineKeyboardButton(web_app=...) in group chats
        # with Button_type_invalid. Use Telegram Direct Link Mini App instead.
        if chat_id < 0 and workspace_type == "family":
            direct_link = await self._build_group_miniapp_direct_link(
                context=context,
                chat_id=chat_id,
            )
            await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=build_miniapp_link_button(workspace, direct_link),
            )
            return
        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=build_miniapp_button(workspace, base_url),
            )
        except BadRequest as exc:
            if "Button_type_invalid" in str(exc):
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=(
                        f"{text}\n\n"
                        "Telegram отклонил web_app-кнопку для этого чата "
                        "(Button_type_invalid)."
                    ),
                )
                return
            raise

    @staticmethod
    def _private_mode_picker_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        t("start_mode_personal_button", ui_lang),
                        callback_data="mode_personal",
                    )
                ],
                [
                    InlineKeyboardButton(
                        t("start_mode_family_button", ui_lang),
                        callback_data="mode_family",
                    )
                ],
            ]
        )

    @staticmethod
    def _private_start_keyboard(add_group_url: str, *, lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        {
                            "ru": "➕ Добавить в группу",
                            "uz": "➕ Guruhga qo'shish",
                            "en": "➕ Add to Group",
                        }.get(ui_lang, "➕ Add to Group"),
                        url=add_group_url,
                    )
                ],
                [
                    InlineKeyboardButton(
                        {
                            "ru": "ℹ️ Как это работает",
                            "uz": "ℹ️ Qanday ishlaydi",
                            "en": "ℹ️ How It Works",
                        }.get(ui_lang, "ℹ️ How It Works"),
                        callback_data="onb:how_private",
                    )
                ],
            ]
        )

    @staticmethod
    def _private_how_keyboard(add_group_url: str, *, lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        return InlineKeyboardMarkup(
            [[
                InlineKeyboardButton(
                    {
                        "ru": "➕ Добавить в группу",
                        "uz": "➕ Guruhga qo'shish",
                        "en": "➕ Add to Group",
                    }.get(ui_lang, "➕ Add to Group"),
                    url=add_group_url,
                )
            ]]
        )

    def _workspace_exists(self, chat_id: int) -> bool:
        try:
            return self.db.workspace_exists(chat_id)
        except Exception:
            logger.exception("Failed to check workspace existence: chat_id=%s", chat_id)
            return False

    async def _send_private_mode_picker(
        self,
        update: Update,
        *,
        context: ContextTypes.DEFAULT_TYPE | None = None,
        lang: str = DEFAULT_LANGUAGE,
        text: str | None = None,
    ) -> None:
        message = update.effective_message
        chat = update.effective_chat
        if not message or not chat or chat.type != "private":
            return
        sent = await message.reply_text(
            text or t("start_private_mode_pick", lang),
            reply_markup=self._private_mode_picker_keyboard(lang=lang),
        )
        if context:
            self._track_onboarding_message(context, sent.message_id)

    async def _ensure_private_mode_selected(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> bool:
        chat = update.effective_chat
        user = update.effective_user
        message = update.effective_message
        if not chat:
            return True
        if self._workspace_exists(chat.id):
            return True

        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE
        if chat.type == "private":
            query = update.callback_query
            if query:
                await self._safe_query_answer(
                    query,
                    t("private_mode_required", lang),
                    show_alert=True,
                )
            await self._send_private_mode_picker(
                update,
                context=context,
                lang=lang,
                text=t("private_mode_required", lang),
            )
            return False

        if chat.type in {"group", "supergroup"}:
            query = update.callback_query
            if query:
                await self._safe_query_answer(
                    query,
                    t("group_start_required", lang),
                    show_alert=True,
                )
            elif message:
                await message.reply_text(t("group_start_required", lang))
            return False

        return True

    @staticmethod
    def _get_personal_onboarding_step(context: ContextTypes.DEFAULT_TYPE) -> int:
        return int(context.user_data.get(PERSONAL_ONBOARDING_STEP_KEY) or 0)

    @staticmethod
    def _set_personal_onboarding_step(context: ContextTypes.DEFAULT_TYPE, step: int) -> None:
        context.user_data[PERSONAL_ONBOARDING_STEP_KEY] = step

    @staticmethod
    def _clear_personal_onboarding(context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop(PERSONAL_ONBOARDING_STEP_KEY, None)
        context.user_data.pop(ONBOARDING_CLEANUP_IDS_KEY, None)

    def _track_onboarding_message(self, context: ContextTypes.DEFAULT_TYPE, message_id: int) -> None:
        ids = context.user_data.get(ONBOARDING_CLEANUP_IDS_KEY)
        if not isinstance(ids, list):
            ids = []
            context.user_data[ONBOARDING_CLEANUP_IDS_KEY] = ids
        ids.append(message_id)

    async def _cleanup_onboarding_messages(self, chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
        ids = context.user_data.pop(ONBOARDING_CLEANUP_IDS_KEY, None)
        if not isinstance(ids, list):
            return
        for mid in ids:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=int(mid))
            except Exception:
                pass

    async def _send_personal_onboarding_step1(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        self._set_personal_onboarding_step(context, 1)
        await context.bot.send_message(
            chat_id=chat_id,
            text=t("ponb_step1", lang),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton(t("ponb_step1_btn", lang), callback_data="onb:p_step2")]]
            ),
        )

    async def _send_personal_onboarding_step2(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        self._set_personal_onboarding_step(context, 2)
        await context.bot.send_message(chat_id=chat_id, text=t("ponb_step2", lang))

    async def _send_personal_onboarding_step3(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        self._set_personal_onboarding_step(context, 3)
        await context.bot.send_message(
            chat_id=chat_id,
            text=t("ponb_step3", lang),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton(t("ponb_skip_btn", lang), callback_data="onb:p_skip_voice")]]
            ),
        )

    async def _send_personal_onboarding_step4(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        self._set_personal_onboarding_step(context, 4)
        await context.bot.send_message(
            chat_id=chat_id,
            text=t("ponb_step4", lang),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton(
                    t("ponb_show_report_btn", lang),
                    callback_data="onb:p_show_report",
                )]]
            ),
        )

    async def _send_personal_onboarding_report(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        state = self._default_report_state()
        payload = self._build_report_payload(chat_id, state, lang=lang)
        report_text = self._build_report_text(
            period_title=str(payload["period_title"]),
            scope_title=str(payload["scope_title"]),
            breakdown=dict(payload["breakdown"]),
            detailed=False,
            transactions=list(payload["transactions"]),
            balance_title=str(payload["balance_title"]),
            show_transfer_summary=False,
            lang=lang,
        )
        await context.bot.send_message(
            chat_id=chat_id,
            text=report_text,
            parse_mode=ParseMode.HTML,
        )
        await self._send_personal_onboarding_step5(chat_id, context, lang)

    async def _send_personal_onboarding_step5(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str
    ) -> None:
        self._set_personal_onboarding_step(context, 5)
        await context.bot.send_message(
            chat_id=chat_id,
            text=t("ponb_step5", lang),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("20:00", callback_data="onb:p_time:20")],
                [InlineKeyboardButton("21:00", callback_data="onb:p_time:21")],
                [InlineKeyboardButton("22:00", callback_data="onb:p_time:22")],
                [InlineKeyboardButton("23:00", callback_data="onb:p_time:23")],
            ]),
        )

    async def _finish_personal_onboarding(
        self, chat_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE, lang: str,
        *, send_hour: int,
    ) -> None:
        resolved = self.db.resolve_workspace(user_id, chat_id, "private")
        if resolved:
            workspace_id = int(resolved.get("workspace_id") or 0)
            if workspace_id:
                for rtype in ("daily", "weekly", "monthly"):
                    minute = 0 if rtype == "daily" else 1
                    self.db.upsert_workspace_report_schedule(
                        workspace_id=workspace_id,
                        report_type=rtype,
                        enabled=True,
                        timezone=DEFAULT_REPORT_TIMEZONE,
                        send_hour=send_hour,
                        send_minute=minute,
                    )
        try:
            self.db.delete_all_transactions(chat_id)
        except Exception:
            logger.exception("Failed to clear onboarding test transactions: chat_id=%s", chat_id)
        self._clear_personal_onboarding(context)
        time_text = t("ponb_step6", lang).format(time=f"{send_hour:02d}:00")
        complete_text = t("ponb_complete", lang)
        base_url = self._miniapp_base_url()
        if base_url:
            await context.bot.send_message(
                chat_id=chat_id,
                text=f"{time_text}\n\n{complete_text}",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton(
                        t("ponb_open_app_btn", lang),
                        web_app=WebAppInfo(url=base_url),
                    )]]
                ),
            )
        else:
            await context.bot.send_message(chat_id=chat_id, text=f"{time_text}\n\n{complete_text}")

    async def _send_personal_mode_activated(
        self,
        *,
        chat_id: int,
        user_id: int,
        context: ContextTypes.DEFAULT_TYPE,
        lang: str,
    ) -> None:
        workspace = self.db.resolve_workspace(user_id, chat_id, "private")
        if not workspace:
            workspace_by_chat = self.db.get_workspace_by_chat_id(chat_id)
            if not workspace_by_chat:
                await context.bot.send_message(chat_id=chat_id, text=t("private_mode_required", lang))
                return
            workspace = workspace_by_chat
        await self._send_miniapp_workspace_message(
            context=context,
            chat_id=chat_id,
            workspace=workspace,
            text="Вы работаете в личном бюджете.",
            lang=lang,
        )

    async def _send_group_family_mode_activated(
        self,
        *,
        chat_id: int,
        context: ContextTypes.DEFAULT_TYPE,
        lang: str,
    ) -> None:
        workspace = self.db.get_workspace_by_chat_id(chat_id)
        if not workspace:
            await context.bot.send_message(chat_id=chat_id, text=t("group_start_required", lang))
            return
        workspace_title = str(workspace.get("title") or "").strip() or "Семейный бюджет"
        await self._send_miniapp_workspace_message(
            context=context,
            chat_id=chat_id,
            workspace=workspace,
            text=f"Вы работаете в семейном бюджете: {workspace_title}",
            lang=lang,
        )

    @staticmethod
    def _group_admin_ready_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[
                InlineKeyboardButton(
                    {
                        "ru": "✅ Готово",
                        "uz": "✅ Tayyor",
                        "en": "✅ Done",
                    }.get(normalize_language(lang), "✅ Done"),
                    callback_data="onb:admin_ready",
                )
            ]]
        )

    def _group_after_onboarding_keyboard(self, *, lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        self._main_menu_report_label(lang),
                        callback_data="mn:report",
                    )
                ],
                [
                    InlineKeyboardButton(
                        self._main_menu_settings_label(lang),
                        callback_data="mn:settings",
                    )
                ],
                [
                    InlineKeyboardButton(
                        self._main_menu_app_label(lang),
                        callback_data="mn:app",
                    )
                ],
            ]
        )

    def _group_main_reply_keyboard(self, *, lang: str = DEFAULT_LANGUAGE) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            [
                [
                    KeyboardButton(self._main_menu_report_label(lang)),
                    KeyboardButton(self._main_menu_settings_label(lang)),
                ],
                [KeyboardButton(self._main_menu_app_label(lang))],
            ],
            resize_keyboard=True,
        )

    @staticmethod
    def _group_back_keyboard(lang: str = DEFAULT_LANGUAGE) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton(t("back", lang), callback_data="onb:back")]]
        )

    def _period_input_cancel_keyboard(
        self, lang: str = DEFAULT_LANGUAGE
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton(t("cancel", lang), callback_data="rp:pcancel")]]
        )

    async def _is_bot_admin_in_group(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE
    ) -> bool:
        try:
            member = await context.bot.get_chat_member(chat_id=chat_id, user_id=context.bot.id)
        except Exception:
            logger.exception("Failed to check bot admin rights in chat %s", chat_id)
            return False
        status = str(getattr(member, "status", "")).lower()
        return status in {"administrator", "creator"}

    async def _human_member_count(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE
    ) -> int | None:
        try:
            count = int(await context.bot.get_chat_member_count(chat_id=chat_id))
        except Exception:
            logger.exception("Failed to get chat member count for %s", chat_id)
            return None
        # Exclude bot itself.
        return max(0, count - 1)

    def _missing_name_handles(self, chat_id: int, *, limit: int = 2) -> list[str]:
        result: list[str] = []
        for member in self.db.list_members(chat_id):
            if member.get("custom_name"):
                continue
            result.append(_member_handle(member))
            if len(result) >= limit:
                break
        return result

    def _choose_intro_member_handles(
        self,
        *,
        chat_id: int,
        starter_user_id: int | None,
    ) -> list[str]:
        members = self.db.list_members(chat_id)
        id_to_handle: dict[int, str] = {}
        for member in members:
            user_id = int(member.get("telegram_user_id") or 0)
            if user_id <= 0:
                continue
            id_to_handle[user_id] = _member_handle(member)

        ordered_ids: list[int] = []
        if isinstance(starter_user_id, int) and starter_user_id > 0:
            ordered_ids.append(starter_user_id)
        for user_id in id_to_handle.keys():
            if user_id not in ordered_ids:
                ordered_ids.append(user_id)
            if len(ordered_ids) >= 2:
                break

        return [id_to_handle[user_id] for user_id in ordered_ids if user_id in id_to_handle][:2]

    async def _auto_set_member_names(
        self,
        chat_id: int,
        context: ContextTypes.DEFAULT_TYPE,
    ) -> None:
        await self._sync_members_from_admins(chat_id, context)
        for member in self.db.list_members(chat_id):
            if member.get("custom_name"):
                continue
            user_id = int(member.get("telegram_user_id") or 0)
            if user_id <= 0:
                continue
            full_name = str(member.get("full_name") or "").strip()
            first_name = full_name.split()[0] if full_name else None
            if first_name:
                self.db.set_custom_name(chat_id, user_id, first_name)

    async def _send_group_family_onboarding(
        self,
        *,
        chat_id: int,
        context: ContextTypes.DEFAULT_TYPE,
        lang: str = DEFAULT_LANGUAGE,
    ) -> None:
        ui_lang = normalize_language(lang)
        members = self.db.list_members(chat_id)
        names = [str(m.get("custom_name") or "").strip() for m in members if str(m.get("custom_name") or "").strip()]
        if len(names) >= 2:
            greeting = f"{names[0]} и {names[1]}"
        elif len(names) == 1:
            greeting = names[0]
        else:
            greeting = ""

        if ui_lang == "uz":
            hello = f"Salom, {greeting}! 👋" if greeting else "Salom! 👋"
            msg1 = (
                f"{hello}\nMen oilaviy moliya yordamchisiman. Qanday ishlashimni ko'rsataman:\n\n"
                "💳 *Shaxsiy xarajatlar*\n"
                "Har bir oila a'zosining o'z xarajatlari — qahva, tushlik, transport.\n"
                "Shunchaki chatga yozing: «qahva 15 000» yoki «taksi 25 000»\n\n"
                "🏠 *Oilaviy xarajatlar*\n"
                "Umumiy xarajatlar — oziq-ovqat, kommunal, ijaraga.\n"
                "Yozing: «oziq-ovqat 200 000» yoki «kommunal 150 000»\n\n"
                "🔄 *Oila ichidagi o'tkazmalar*\n"
                f"{'Agar ' + names[0] + ' ' + names[1] + 'ga' if len(names) >= 2 else 'Agar er xotinga'}"
                " pul o'tkazsa — bu xarajat emas, shunchaki cho'ntakdan cho'ntakka.\n"
                "Yozing: «o'tkazdim 300 000»"
            )
        elif ui_lang == "en":
            hello = f"Hello, {greeting}! 👋" if greeting else "Hello! 👋"
            msg1 = (
                f"{hello}\nI'm your family finance assistant. Here's how I work:\n\n"
                "💳 *Personal expenses*\n"
                "Each family member's own spending — coffee, lunch, transport.\n"
                "Just type in chat: \"coffee 15,000\" or \"taxi 25,000\"\n\n"
                "🏠 *Family expenses*\n"
                "Shared costs — groceries, utilities, rent.\n"
                "Type: \"groceries 200,000\" or \"utilities 150,000\"\n\n"
                "🔄 *Family transfers*\n"
                f"{'If ' + names[0] + ' sends money to ' + names[1] if len(names) >= 2 else 'If husband sends money to wife'}"
                " — that's not an expense, just pocket to pocket.\n"
                "Type: \"transferred 300,000\""
            )
        else:
            hello = f"Привет, {greeting}! 👋" if greeting else "Привет! 👋"
            msg1 = (
                f"{hello}\nЯ ваш семейный финансовый ассистент. Покажу как работаю:\n\n"
                "💳 *Личные расходы*\n"
                "Свои траты каждого — кофе, обед, такси.\n"
                "Просто пишите в чат: «кофе 15 000» или «такси 25 000»\n\n"
                "🏠 *Семейные расходы*\n"
                "Общие траты — продукты, коммуналка, аренда.\n"
                "Пишите: «продукты 200 000» или «коммуналка 150 000»\n\n"
                "🔄 *Переводы внутри семьи*\n"
                f"{'Если ' + names[0] + ' переводит ' + names[1] if len(names) >= 2 else 'Если муж переводит жене'}"
                " — это не расход, просто деньги из одного кармана в другой.\n"
                "Пишите: «перевёл 300 000»"
            )

        ui_try = {"ru": "🚀 Попробуем!", "uz": "🚀 Sinab ko'ramiz!", "en": "🚀 Let's try!"}.get(ui_lang, "🚀 Let's try!")
        await context.bot.send_message(
            chat_id=chat_id,
            text=msg1,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(ui_try, callback_data="onb:try_it"),
            ]]),
        )

    async def _send_group_onboarding_final(
        self,
        *,
        chat_id: int,
        context: ContextTypes.DEFAULT_TYPE,
        lang: str = DEFAULT_LANGUAGE,
    ) -> None:
        await context.bot.send_message(
            chat_id=chat_id,
            text=self._group_onboarding_final_text(lang=lang),
            reply_markup=self._group_after_onboarding_keyboard(lang=lang),
        )
        await context.bot.send_message(
            chat_id=chat_id,
            text={
                "ru": "Внизу появились быстрые кнопки: Отчёты, Настройки Ассистента и Приложение.",
                "uz": "Pastda tezkor tugmalar paydo bo'ldi: Hisobotlar, Assistent sozlamalari va Ilova.",
                "en": "Quick buttons appeared below: Reports, Assistant Settings, and App.",
            }.get(normalize_language(lang), "Quick buttons appeared below."),
            reply_markup=self._group_main_reply_keyboard(lang=lang),
        )
        await self._send_group_family_mode_activated(
            chat_id=chat_id,
            context=context,
            lang=lang,
        )

    @staticmethod
    def _group_onboarding_final_text(*, lang: str = DEFAULT_LANGUAGE) -> str:
        ui_lang = normalize_language(lang)
        if ui_lang == "uz":
            return (
                "Tayyor ✅ Endi chatga xarajat va daromadlarni yozsangiz bo'ldi.\n"
                "Misollar:\n"
                "• «qahva 20 000»\n"
                "• «mahsulotlar 150 000»\n"
                "• «maosh 5 000 000»\n\n"
                "Bir-biringizga pul o'tkazsangiz — shunday yozing:\n"
                "• «turmush o'rtog'imga 300 000 o'tkazdim» "
                "(bu oilaviy xarajat emas, ichki o'tkazma)"
            )
        if ui_lang == "en":
            return (
                "Done ✅ Now just send expenses and income messages in the chat.\n"
                "Examples:\n"
                "• “coffee 20,000”\n"
                "• “groceries 150,000”\n"
                "• “salary 5,000,000”\n\n"
                "If you transfer money to each other, write:\n"
                "• “transferred 300,000 to my spouse” "
                "(this is not a family expense, it is an internal transfer)"
            )
        return (
            "Готово ✅ Теперь просто пишите в чат расходы и доходы.\n"
            "Примеры:\n"
            "• «кофе 20 000»\n"
            "• «продукты 150 000»\n"
            "• «зарплата 5 000 000»\n\n"
            "Если переводите деньги друг другу — пишите:\n"
            "• «перевёл жене 300 000» "
            "(это не расход семьи, а перевод внутри)"
        )

    @staticmethod
    def _demo_report_text() -> str:
        return (
            "📅 17 февраля\n\n"
            "💰 Баланс дня\n"
            "Доходы: 5 000 000 сум\n"
            "Расходы: 3 200 000 сум\n"
            "Итого: +1 800 000 сум\n\n"
            "👨‍👩‍👧 Семейные расходы\n"
            "🏠 Коммунальные — 2 000 000 сум\n"
            "🛒 Продукты — 800 000 сум\n"
            "🍽 Кафе — 400 000 сум\n\n"
            "👤 Личные расходы\n"
            "Ромео — 2 100 000 сум\n"
            "Джульетта — 1 100 000 сум"
        )

    @staticmethod
    def _settings_placeholder_text() -> str:
        return (
            "⚙️ Настройки Ассистента\n"
            "Скоро добавим: язык, валюту и другие параметры."
        )

    def _category_group_view(self, category: str, kind: str) -> tuple[str, str]:
        emoji, _ = self._report_category_view(category, kind)
        label = _ru_category_label(category, kind)
        group = label.split(" / ", 1)[0] if " / " in label else label
        return emoji, group

    def _looks_like_guessable_amount_message(self, text: str) -> bool:
        text_l = " ".join((text or "").lower().split())
        if not _has_amount_and_text(text_l):
            return False
        if re.search(r"\d+\s*[\+\-\*/]\s*\d+", text_l):
            return False
        if "?" in text_l:
            return False
        if any(
            marker in text_l
            for marker in (
                "купил",
                "купила",
                "оплат",
                "заплат",
                "потрат",
                "получил",
                "получила",
                "доход",
                "расход",
                "перев",
            )
        ):
            return False
        words = [w for w in text_l.split() if w]
        if len(words) > 5:
            return False
        return True

    @staticmethod
    def _guess_kind_from_text(text_l: str) -> str:
        income_markers = (
            "получил",
            "получила",
            "получили",
            "зарплат",
            "зп",
            "доход",
            "бонус",
            "преми",
            "выигр",
            "побед",
            "нашел",
            "нашёл",
            "прибыл",
            "profit",
            "cashback",
            "gift",
        )
        if any(marker in text_l for marker in income_markers):
            return "income"
        return "expense"

    async def _build_guess_event(
        self,
        *,
        text: str,
    ) -> ParsedEvent | None:
        amount = _extract_amount_value(text)
        if amount is None:
            return None

        text_l = " ".join((text or "").lower().replace("ё", "е").split())
        kind = self._guess_kind_from_text(text_l)
        currency = self._detect_currency_from_text(text_l, self.settings.default_currency)
        category = self.ai.guess_place_category(text_l, kind)
        if category is None:
            try:
                category = await self.ai.infer_category(
                    kind=kind,
                    description=text.strip()[:120],
                    amount=amount,
                    currency=currency,
                    is_family=False,
                )
            except Exception:
                logger.exception("Failed to infer category for guess flow")
                category = "expense_other" if kind == "expense" else "income_other"

        if kind == "expense" and category == "expense_other":
            if any(
                token in text_l
                for token in (
                    "яблок",
                    "банан",
                    "груш",
                    "фрукт",
                    "овощ",
                    "хлеб",
                    "молок",
                    "мяс",
                    "куриц",
                    "рыб",
                    "сыр",
                    "яйц",
                    "картош",
                    "помидор",
                    "огур",
                )
            ):
                category = "groceries_products"

        is_family = kind == "expense" and (
            category.startswith("groceries_")
            or category.startswith("home_")
            or category.startswith("kids_")
            or category in {"pets_food", "pets_vet"}
        )
        return ParsedEvent(
            kind=kind,
            amount=amount,
            currency=currency,
            category=category,
            description=text.strip()[:120],
            is_family=is_family,
            confidence=0.6,
        )

    @staticmethod
    def _guess_confirm_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("✅ Да", callback_data="gc:yes")],
                [InlineKeyboardButton("✏️ Изменить категорию", callback_data="gc:change")],
                [InlineKeyboardButton("❌ Отмена", callback_data="gc:cancel")],
            ]
        )

    @staticmethod
    def _guess_category_picker_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("🍽 Кафе и рестораны", callback_data="gc:cat:cafe_fastfood"),
                    InlineKeyboardButton("🛒 Продукты и быт", callback_data="gc:cat:groceries_products"),
                ],
                [
                    InlineKeyboardButton("🏠 Жильё и дом", callback_data="gc:cat:home_internet"),
                    InlineKeyboardButton("🚗 Транспорт", callback_data="gc:cat:transport_taxi"),
                ],
                [InlineKeyboardButton("📦 Прочие расходы", callback_data="gc:cat:expense_other")],
                [InlineKeyboardButton("↩️ Назад", callback_data="gc:back")],
            ]
        )

    async def _ask_guess_confirmation(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        event: ParsedEvent,
        source_text: str,
    ) -> None:
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return

        emoji, group = self._category_group_view(event.category, event.kind)
        kind_text = "доход" if event.kind == "income" else "расход"
        prompt_text = (
            f"Похоже, это {kind_text} на {_fmt_money(event.amount, event.currency)}.\n"
            "Я правильно понимаю?\n"
            f"Категория: {emoji} {group}"
        )
        prompt = await message.reply_text(
            prompt_text,
            reply_markup=self._guess_confirm_keyboard(),
        )
        context.user_data[PENDING_GUESS_KEY] = {
            "chat_id": chat.id,
            "user_id": user.id,
            "source_message_id": message.message_id,
            "prompt_message_id": prompt.message_id,
            "source_text": source_text,
            "event": {
                "kind": event.kind,
                "amount": event.amount,
                "currency": event.currency,
                "category": event.category,
                "description": event.description,
                "is_family": event.is_family,
            },
        }

    def _default_report_state(self) -> dict[str, object]:
        return {
            "scope": "all",
            "selected_user_id": None,
            "detailed": False,
            "period_mode": "day",
            "custom_date": None,
            "range_from": None,
            "range_to": None,
        }

    def _get_report_state(
        self, context: ContextTypes.DEFAULT_TYPE, chat_id: int
    ) -> dict[str, object]:
        raw_storage = context.user_data.get(REPORT_STATE_KEY)
        if not isinstance(raw_storage, dict):
            raw_storage = {}
            context.user_data[REPORT_STATE_KEY] = raw_storage

        key = str(chat_id)
        raw_state = raw_storage.get(key)
        if not isinstance(raw_state, dict):
            raw_state = self._default_report_state()
            raw_storage[key] = raw_state

        # Ensure all keys exist after updates.
        defaults = self._default_report_state()
        for field, default_value in defaults.items():
            raw_state.setdefault(field, default_value)
        return raw_state

    @staticmethod
    def _parse_user_date_token(token: str) -> date | None:
        cleaned = token.strip()
        if not cleaned:
            return None
        cleaned = re.sub(r"[.,/]", "-", cleaned)
        cleaned = re.sub(r"-{2,}", "-", cleaned)
        for fmt in (
            "%d-%m-%Y",
            "%d-%m-%y",
            "%Y-%m-%d",
        ):
            try:
                parsed = datetime.strptime(cleaned, fmt)
                return parsed.date()
            except ValueError:
                continue
        return None

    def _parse_user_date_range(self, text: str) -> tuple[date, date] | None:
        tokens = re.findall(
            r"\d{4}[./,\-]\d{1,2}[./,\-]\d{1,2}|\d{1,2}[./,\-]\d{1,2}[./,\-]\d{2,4}",
            text,
        )
        if len(tokens) < 2:
            return None
        left = self._parse_user_date_token(tokens[0])
        right = self._parse_user_date_token(tokens[1])
        if left is None or right is None:
            return None
        if left > right:
            left, right = right, left
        return left, right

    def _looks_like_date_or_range_text(self, text: str) -> bool:
        cleaned = (text or "").strip()
        if not cleaned:
            return False
        if self._parse_user_date_token(cleaned) is not None:
            return True
        if self._parse_user_date_range(cleaned) is not None:
            return True
        return False

    @staticmethod
    def _looks_like_report_dump_text(text: str) -> bool:
        text_l = (text or "").lower()
        if "━━━━━━━━" in text:
            return True
        report_markers = (
            "баланс дня",
            "баланс периода",
            "доходы:",
            "расходы:",
            "итого:",
            "посмотреть детали",
            "семейные",
        )
        hits = sum(1 for marker in report_markers if marker in text_l)
        return hits >= 3

    @staticmethod
    def _report_text(key: str, lang: str = DEFAULT_LANGUAGE, **kwargs: object) -> str:
        code = normalize_language(lang)
        texts: dict[str, dict[str, str]] = {
            "today": {"ru": "Сегодня", "uz": "Bugun", "en": "Today"},
            "week_prefix": {"ru": "Неделя", "uz": "Hafta", "en": "Week"},
            "period_prefix": {"ru": "Период", "uz": "Davr", "en": "Period"},
            "balance_day": {"ru": "💰 Баланс дня", "uz": "💰 Kun balansi", "en": "💰 Day Balance"},
            "balance_period": {
                "ru": "💰 Баланс периода",
                "uz": "💰 Davr balansi",
                "en": "💰 Period Balance",
            },
            "scope_family": {"ru": "Семейные", "uz": "Oilaviy", "en": "Family"},
            "scope_all": {"ru": "Всё", "uz": "Hammasi", "en": "All"},
            "toggle_show_details": {
                "ru": "🔎 Показать детали",
                "uz": "🔎 Tafsilotlarni ko'rsatish",
                "en": "🔎 Show Details",
            },
            "toggle_hide_details": {
                "ru": "↩️ Скрыть детали",
                "uz": "↩️ Tafsilotlarni yashirish",
                "en": "↩️ Hide Details",
            },
            "pick_period": {
                "ru": "📅 Выбрать дату/период",
                "uz": "📅 Sana/davr tanlash",
                "en": "📅 Choose Date/Period",
            },
            "download_excel": {
                "ru": "📥 Скачать Excel",
                "uz": "📥 Excel yuklab olish",
                "en": "📥 Download Excel",
            },
            "truncated_report_note": {
                "ru": "Отчёт длинный, показываю только часть. Уточните период при необходимости.",
                "uz": "Hisobot uzun, faqat bir qismini ko'rsatdim. Kerak bo'lsa davrni aniqlashtiring.",
                "en": "The report is long, so I’m showing only part of it. Narrow the period if needed.",
            },
            "transfers_sent": {
                "ru": "⬅ Переводы отправлено: {amount}",
                "uz": "⬅ Yuborilgan o'tkazmalar: {amount}",
                "en": "⬅ Transfers sent: {amount}",
            },
            "transfers_received": {
                "ru": "➡ Переводы получено: {amount}",
                "uz": "➡ Olingan o'tkazmalar: {amount}",
                "en": "➡ Transfers received: {amount}",
            },
            "income": {"ru": "Доходы", "uz": "Daromadlar", "en": "Income"},
            "expense": {"ru": "Расходы", "uz": "Xarajatlar", "en": "Expenses"},
            "total": {"ru": "Итого", "uz": "Jami", "en": "Total"},
            "section_expense": {"ru": "📈 Расходы", "uz": "📈 Xarajatlar", "en": "📈 Expenses"},
            "section_income": {"ru": "↘ Доходы", "uz": "↘ Daromadlar", "en": "↘ Income"},
            "no_ops_period": {
                "ru": "- За выбранный период пока операций нет.",
                "uz": "- Tanlangan davr uchun hozircha operatsiyalar yo'q.",
                "en": "- No transactions yet for the selected period.",
            },
            "more_categories": {
                "ru": "… и ещё {count} катег.",
                "uz": "… va yana {count} ta kategoriya",
                "en": "… and {count} more categories",
            },
            "period_cancelled": {
                "ru": "Хорошо, выбор периода отменил.",
                "uz": "Mayli, davr tanlovi bekor qilindi.",
                "en": "Okay, period selection was cancelled.",
            },
            "period_too_long": {
                "ru": "Период слишком длинный. Выберите до 366 дней.",
                "uz": "Davr juda uzun. 366 kungacha tanlang.",
                "en": "The period is too long. Please choose up to 366 days.",
            },
            "cant_parse_date": {
                "ru": "Не смог прочитать дату. Пример: 17.02.2026 или 2026-02-17.",
                "uz": "Sanani o'qib bo'lmadi. Misol: 17.02.2026 yoki 2026-02-17.",
                "en": "Couldn’t read the date. Example: 17.02.2026 or 2026-02-17.",
            },
            "single_date_confirm": {
                "ru": "Вы хотите получить отчёт за {date}?",
                "uz": "{date} uchun hisobotni olishni xohlaysizmi?",
                "en": "Do you want a report for {date}?",
            },
            "cant_parse_date_or_period": {
                "ru": "Не смог прочитать дату/период.\nПримеры:\n17.02.2026\n01.01.2026 - 15.01.2026",
                "uz": "Sana/davrni o'qib bo'lmadi.\nMisollar:\n17.02.2026\n01.01.2026 - 15.01.2026",
                "en": "Couldn’t read the date/period.\nExamples:\n17.02.2026\n01.01.2026 - 15.01.2026",
            },
            "single_date_confirm_outdated": {
                "ru": "Подтверждение уже неактуально.",
                "uz": "Tasdiqlash endi dolzarb emas.",
                "en": "This confirmation is no longer valid.",
            },
            "single_date_reenter": {
                "ru": "✍️ Ввести заново",
                "uz": "✍️ Qayta kiritish",
                "en": "✍️ Enter Again",
            },
            "reenter_period_prompt": {
                "ru": "Хорошо, напиши дату или период заново:\n17.02.2026\nили\n01.01.2026 - 15.01.2026",
                "uz": "Mayli, sana yoki davrni qayta yozing:\n17.02.2026\nyoki\n01.01.2026 - 15.01.2026",
                "en": "Okay, send the date or period again:\n17.02.2026\nor\n01.01.2026 - 15.01.2026",
            },
            "waiting_period_answer": {
                "ru": "Ок, жду период",
                "uz": "Xo'p, davrni kutyapman",
                "en": "Okay, waiting for the period",
            },
            "done": {"ru": "Готово", "uz": "Tayyor", "en": "Done"},
            "ask_date_or_period": {
                "ru": "Напиши дату: день.месяц.год, например: 17.02.2026\nИли сразу период: 01.01.2026 - 15.01.2026.",
                "uz": "Sanani yozing: kun.oy.yil, masalan: 17.02.2026\nYoki darrov davr: 01.01.2026 - 15.01.2026.",
                "en": "Send a date: day.month.year, for example: 17.02.2026\nOr a range right away: 01.01.2026 - 15.01.2026.",
            },
            "ask_period_only": {
                "ru": "Напиши период, например:\n01.01.2026 - 15.01.2026",
                "uz": "Davrni yozing, masalan:\n01.01.2026 - 15.01.2026",
                "en": "Send a period, for example:\n01.01.2026 - 15.01.2026",
            },
            "cancelled": {"ru": "Отменено", "uz": "Bekor qilindi", "en": "Cancelled"},
            "collecting_excel": {
                "ru": "Собираю Excel...",
                "uz": "Excel tayyorlanyapti...",
                "en": "Preparing Excel...",
            },
            "member_not_in_report_menu": {
                "ru": "Этот участник не в меню отчёта.",
                "uz": "Bu ishtirokchi hisobot menyusida yo'q.",
                "en": "This member is not in the report menu.",
            },
            "cant_open_report": {
                "ru": "Не получилось открыть отчёт.",
                "uz": "Hisobotni ochib bo'lmadi.",
                "en": "Couldn’t open the report.",
            },
            "opening_report": {
                "ru": "Открываю отчёт",
                "uz": "Hisobot ochilyapti",
                "en": "Opening report",
            },
            "member_fallback": {
                "ru": "Участник {id}",
                "uz": "Ishtirokchi {id}",
                "en": "Member {id}",
            },
            "excel_generate_failed": {
                "ru": "Не получилось собрать Excel-файл. Попробуйте ещё раз.",
                "uz": "Excel faylni tayyorlab bo'lmadi. Qayta urinib ko'ring.",
                "en": "Couldn’t generate the Excel file. Please try again.",
            },
            "excel_sheet_title": {"ru": "Отчёт", "uz": "Hisobot", "en": "Report"},
            "excel_title_prefix": {"ru": "Отчёт", "uz": "Hisobot", "en": "Report"},
            "excel_col_date": {"ru": "Дата", "uz": "Sana", "en": "Date"},
            "excel_col_time": {"ru": "Время", "uz": "Vaqt", "en": "Time"},
            "excel_col_type": {"ru": "Тип транзакции", "uz": "Tranzaksiya turi", "en": "Transaction Type"},
            "excel_col_category": {"ru": "Категория", "uz": "Kategoriya", "en": "Category"},
            "excel_col_amount": {"ru": "Сумма", "uz": "Summa", "en": "Amount"},
            "excel_col_currency": {"ru": "Валюта", "uz": "Valyuta", "en": "Currency"},
            "excel_col_description": {"ru": "Описание", "uz": "Tavsif", "en": "Description"},
            "excel_type_transfer": {"ru": "Перевод", "uz": "O'tkazma", "en": "Transfer"},
            "excel_type_income": {"ru": "Доход", "uz": "Daromad", "en": "Income"},
            "excel_type_expense": {"ru": "Расход", "uz": "Xarajat", "en": "Expense"},
            "excel_family_transfers": {
                "ru": "🔁 Семейные переводы",
                "uz": "🔁 Oilaviy o'tkazmalar",
                "en": "🔁 Family Transfers",
            },
            "excel_total_income": {"ru": "Итого доходы", "uz": "Jami daromad", "en": "Total income"},
            "excel_total_expense": {"ru": "Итого расходы", "uz": "Jami xarajat", "en": "Total expenses"},
            "excel_total_balance": {"ru": "Итог (баланс)", "uz": "Jami (balans)", "en": "Total (balance)"},
            "excel_total_transfers": {"ru": "Итого переводы", "uz": "Jami o'tkazmalar", "en": "Total transfers"},
        }
        by_lang = texts.get(key)
        if not by_lang:
            return key
        template = by_lang.get(code) or by_lang.get(DEFAULT_LANGUAGE) or key
        if kwargs:
            try:
                return template.format(**kwargs)
            except Exception:
                return template
        return template

    @staticmethod
    def _report_month_name(month: int, *, lang: str = DEFAULT_LANGUAGE, genitive: bool = False) -> str:
        code = normalize_language(lang)
        idx = max(1, min(12, int(month))) - 1
        if code == "uz":
            return MONTH_NAMES_UZ[idx]
        if code == "en":
            return MONTH_NAMES_EN[idx]
        return MONTH_NAMES_RU_GENITIVE[idx] if genitive else MONTH_NAMES_RU[idx]

    def _resolve_report_period(
        self, state: dict[str, object], *, lang: str = DEFAULT_LANGUAGE
    ) -> tuple[datetime, datetime, str, str]:
        ui_lang = normalize_language(lang)
        mode = str(state.get("period_mode") or "day")
        now = datetime.now()
        if mode == "week":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(
                days=now.weekday()
            )
            end = start + timedelta(days=7)
            end_inclusive = end - timedelta(days=1)
            title = (
                f"{self._report_text('week_prefix', ui_lang)}: "
                f"{start:%d.%m.%Y} — {end_inclusive:%d.%m.%Y}"
            )
            return start, end, title, self._report_text("balance_period", ui_lang)
        if mode == "month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if start.month == 12:
                end = start.replace(year=start.year + 1, month=1)
            else:
                end = start.replace(month=start.month + 1)
            if ui_lang == "en":
                title = f"{self._report_month_name(start.month, lang=ui_lang)} {start.year}"
            elif ui_lang == "uz":
                title = f"{self._report_month_name(start.month, lang=ui_lang).capitalize()} {start.year}"
            else:
                title = f"{MONTH_NAMES_RU[start.month - 1].capitalize()} {start.year}"
            return start, end, title, self._report_text("balance_period", ui_lang)
        if mode == "date":
            picked = state.get("custom_date")
            target_date = picked if isinstance(picked, date) else now.date()
            start = datetime.combine(target_date, datetime.min.time())
            end = start + timedelta(days=1)
            if ui_lang == "ru":
                title = (
                    f"{target_date.day} "
                    f"{self._report_month_name(target_date.month, lang=ui_lang, genitive=True)} "
                    f"{target_date.year}"
                )
            elif ui_lang == "en":
                title = f"{self._report_month_name(target_date.month, lang=ui_lang)} {target_date.day}, {target_date.year}"
            else:
                title = f"{target_date.day} {self._report_month_name(target_date.month, lang=ui_lang)} {target_date.year}"
            if target_date == now.date():
                title += f" — {self._report_text('today', ui_lang)}"
            return start, end, title, self._report_text("balance_day", ui_lang)
        if mode == "range":
            range_from = state.get("range_from")
            range_to = state.get("range_to")
            if isinstance(range_from, date) and isinstance(range_to, date):
                start_date = range_from if range_from <= range_to else range_to
                end_date = range_to if range_to >= range_from else range_from
                start = datetime.combine(start_date, datetime.min.time())
                end = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
                title = (
                    f"{self._report_text('period_prefix', ui_lang)}: "
                    f"{start_date:%d.%m.%Y} — {end_date:%d.%m.%Y}"
                )
                return start, end, title, self._report_text("balance_period", ui_lang)

        # day (default)
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        if ui_lang == "ru":
            title = (
                f"{now.day} {self._report_month_name(now.month, lang=ui_lang, genitive=True)} "
                f"{now.year} — {self._report_text('today', ui_lang)}"
            )
        elif ui_lang == "en":
            title = (
                f"{self._report_month_name(now.month, lang=ui_lang)} {now.day}, {now.year} "
                f"— {self._report_text('today', ui_lang)}"
            )
        else:
            title = (
                f"{now.day} {self._report_month_name(now.month, lang=ui_lang)} {now.year} "
                f"— {self._report_text('today', ui_lang)}"
            )
        return start, end, title, self._report_text("balance_day", ui_lang)

    def _build_report_payload(
        self, chat_id: int, state: dict[str, object], *, lang: str = DEFAULT_LANGUAGE
    ) -> dict[str, object]:
        ui_lang = normalize_language(lang)
        start_local, end_local, period_title, balance_title = self._resolve_report_period(
            state, lang=ui_lang
        )
        start_query = self._local_naive_to_utc_naive(start_local)
        end_query = self._local_naive_to_utc_naive(end_local)

        scope = str(state.get("scope") or "all")
        selected_user_id = state.get("selected_user_id")
        if isinstance(selected_user_id, str) and selected_user_id.isdigit():
            selected_user_id = int(selected_user_id)
        if not isinstance(selected_user_id, int):
            selected_user_id = None

        if scope == "family":
            scope_title = f"👨‍👩‍👧‍👦 {self._report_text('scope_family', ui_lang)}"
            transactions = self.db.get_period_transactions(
                chat_id, start_query, end_query, is_family=True
            )
        elif scope == "user" and selected_user_id is not None:
            member_items = dict(self._report_member_items(chat_id, lang=ui_lang))
            if selected_user_id not in member_items:
                scope = "all"
                selected_user_id = None
                scope_title = f"📊 {self._report_text('scope_all', ui_lang)}"
                transactions = self.db.get_period_transactions(chat_id, start_query, end_query)
            else:
                scope_title = f"👤 {member_items[selected_user_id]}"
                transactions = self.db.get_period_transactions(
                    chat_id,
                    start_query,
                    end_query,
                    telegram_user_id=selected_user_id,
                )
        else:
            scope = "all"
            selected_user_id = None
            scope_title = f"📊 {self._report_text('scope_all', ui_lang)}"
            transactions = self.db.get_period_transactions(chat_id, start_query, end_query)

        breakdown = self._build_breakdown_from_transactions(transactions)
        return {
            "scope": scope,
            "selected_user_id": selected_user_id,
            "period_mode": str(state.get("period_mode") or "day"),
            "period_title": period_title,
            "balance_title": balance_title,
            "scope_title": scope_title,
            "transactions": transactions,
            "breakdown": breakdown,
        }

    def _report_single_date_confirm_keyboard(
        self, lang: str = DEFAULT_LANGUAGE
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(t("confirm_yes", lang), callback_data="rp:p1d:y"),
                    InlineKeyboardButton(
                        self._report_text("single_date_reenter", lang),
                        callback_data="rp:p1d:n",
                    ),
                ]
            ]
        )

    async def _try_apply_pending_report_filter(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: str,
    ) -> bool:
        pending = context.user_data.get(REPORT_FILTER_INPUT_KEY)
        if not isinstance(pending, dict):
            return False

        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat:
            return False
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE

        pending_chat_id = int(pending.get("chat_id") or 0)
        if pending_chat_id and pending_chat_id != chat.id:
            return False

        mode = str(pending.get("mode") or "")
        # If user continues typing while single-date confirmation is open,
        # treat new input as a fresh attempt.
        if mode in {"range", "date_or_range"}:
            context.user_data.pop(REPORT_SINGLE_DATE_CONFIRM_KEY, None)
        text_l = text.strip().lower()
        if text_l in {"отмена", "cancel"}:
            context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
            context.user_data.pop(REPORT_SINGLE_DATE_CONFIRM_KEY, None)
            await message.reply_text(self._report_text("period_cancelled", lang))
            return True

        state = self._get_report_state(context, chat.id)
        if mode == "date":
            parsed_range = self._parse_user_date_range(text)
            if parsed_range is not None:
                range_from, range_to = parsed_range
                if (range_to - range_from).days > 366:
                    await message.reply_text(self._report_text("period_too_long", lang))
                    return True
                state["period_mode"] = "range"
                state["range_from"] = range_from
                state["range_to"] = range_to
                state["custom_date"] = None
            else:
                target = self._parse_user_date_token(text)
                if target is None:
                    await message.reply_text(self._report_text("cant_parse_date", lang))
                    return True
                state["period_mode"] = "date"
                state["custom_date"] = target
                state["range_from"] = None
                state["range_to"] = None
        elif mode in {"range", "date_or_range"}:
            parsed = self._parse_user_date_range(text)
            if parsed is None:
                single_date = self._parse_user_date_token(text)
                if single_date is not None:
                    context.user_data[REPORT_SINGLE_DATE_CONFIRM_KEY] = {
                        "chat_id": chat.id,
                        "date": single_date,
                    }
                    await message.reply_text(
                        self._report_text(
                            "single_date_confirm",
                            lang,
                            date=f"{single_date:%d.%m.%Y}",
                        ),
                        reply_markup=self._report_single_date_confirm_keyboard(lang),
                    )
                    return True
                await message.reply_text(self._report_text("cant_parse_date_or_period", lang))
                return True
            range_from, range_to = parsed
            if (range_to - range_from).days > 366:
                await message.reply_text(self._report_text("period_too_long", lang))
                return True
            state["period_mode"] = "range"
            state["range_from"] = range_from
            state["range_to"] = range_to
            state["custom_date"] = None
        else:
            context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
            return False

        context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
        payload = self._build_report_payload(chat.id, state, lang=lang)
        detailed = bool(state.get("detailed", False))
        report_text = self._build_report_text(
            period_title=str(payload["period_title"]),
            scope_title=str(payload["scope_title"]),
            breakdown=dict(payload["breakdown"]),
            detailed=detailed,
            transactions=list(payload["transactions"]),
            balance_title=str(payload["balance_title"]),
            show_transfer_summary=str(payload["scope"]) == "user",
            lang=lang,
        )
        sent = await message.reply_text(
            report_text,
            parse_mode=ParseMode.HTML,
            reply_markup=self._report_menu_keyboard(
                chat.id,
                selected_scope=str(payload["scope"]),
                selected_user_id=(
                    int(payload["selected_user_id"])
                    if isinstance(payload.get("selected_user_id"), int)
                    else None
                ),
                detailed=detailed,
                period_mode=str(payload["period_mode"]),
                lang=lang,
            ),
        )
        state["report_message_id"] = sent.message_id
        return True

    @staticmethod
    def _detect_currency_from_text(text: str, default_currency: str) -> str:
        text_l = (text or "").lower()
        if any(token in text_l for token in ("uzs", "сум", "сўм", "so'm", "so‘m", "sum")):
            return "UZS"
        if any(token in text_l for token in ("usd", "$", "доллар")):
            return "USD"
        if any(token in text_l for token in ("eur", "€", "евро")):
            return "EUR"
        if any(token in text_l for token in ("rub", "₽", "руб")):
            return "RUB"
        return default_currency

    @staticmethod
    def _is_transfer_intent(text: str) -> bool:
        text_l = (text or "").lower()
        transfer_markers = (
            "перевел",
            "перевёл",
            "перевод",
            "перевести",
            "скинул",
            "скинула",
            "отправил",
            "отправила",
            "кинул",
            "кинула",
            "перекинул",
            "перекинула",
        )
        return any(marker in text_l for marker in transfer_markers)

    @staticmethod
    def _normalize_match_text(text: str) -> str:
        return " ".join((text or "").lower().replace("ё", "е").split())

    def _transfer_recipient_candidates(
        self, chat_id: int, sender_user_id: int
    ) -> list[dict[str, object]]:
        candidates: list[dict[str, object]] = []
        for member in self.db.list_members(chat_id):
            user_id = int(member.get("telegram_user_id") or 0)
            if user_id <= 0 or user_id == sender_user_id:
                continue
            candidates.append(member)
        return candidates

    def _resolve_transfer_recipient(
        self,
        *,
        chat_id: int,
        sender_user_id: int,
        text: str,
    ) -> int | None:
        candidates = self._transfer_recipient_candidates(chat_id, sender_user_id)
        if not candidates:
            return None

        text_norm = self._normalize_match_text(text)
        username_matches = re.findall(r"@([a-zA-Z0-9_]{3,})", text)
        if username_matches:
            usernames = {u.lower() for u in username_matches}
            for member in candidates:
                username = str(member.get("username") or "").lower()
                if username and username in usernames:
                    return int(member["telegram_user_id"])

        matched_ids: set[int] = set()
        for member in candidates:
            user_id = int(member["telegram_user_id"])
            aliases: set[str] = set()
            custom_name = str(member.get("custom_name") or "").strip()
            full_name = str(member.get("full_name") or "").strip()
            username = str(member.get("username") or "").strip()
            for raw in (custom_name, full_name):
                if raw:
                    aliases.add(raw)
                    first = raw.split()[0]
                    if first:
                        aliases.add(first)
            if username:
                aliases.add(username)
            for alias in aliases:
                alias_norm = self._normalize_match_text(alias)
                if not alias_norm:
                    continue
                if re.search(rf"(?<!\w){re.escape(alias_norm)}(?!\w)", text_norm):
                    matched_ids.add(user_id)
                    break

        if len(matched_ids) == 1:
            return next(iter(matched_ids))

        relation_markers = (
            "жене",
            "жене ",
            "супруге",
            "мужу",
            "супругу",
            "wife",
            "husband",
        )
        if len(candidates) == 1 and any(marker in text_norm for marker in relation_markers):
            return int(candidates[0]["telegram_user_id"])

        return None

    def _transfer_recipient_keyboard(
        self,
        *,
        chat_id: int,
        sender_user_id: int,
    ) -> InlineKeyboardMarkup:
        rows: list[list[InlineKeyboardButton]] = []
        for member in self._transfer_recipient_candidates(chat_id, sender_user_id):
            user_id = int(member.get("telegram_user_id") or 0)
            display_name = (
                str(member.get("custom_name") or "").strip()
                or str(member.get("full_name") or "").strip()
                or (f"@{member.get('username')}" if member.get("username") else f"Участник {user_id}")
            )
            rows.append(
                [InlineKeyboardButton(f"👤 {display_name}", callback_data=f"tf:to:{user_id}")]
            )
        rows.append([InlineKeyboardButton("❌ Отменить", callback_data="tf:cancel")])
        return InlineKeyboardMarkup(rows)

    def _record_internal_transfer(
        self,
        *,
        chat_id: int,
        sender_user_id: int,
        recipient_user_id: int,
        amount: float,
        currency: str,
        source_text: str,
        source_message_id: int,
    ) -> tuple[str, str, int, int]:
        sender_name = (
            self.db.get_member_display_name(chat_id, sender_user_id) or f"Участник {sender_user_id}"
        )
        recipient_name = (
            self.db.get_member_display_name(chat_id, recipient_user_id)
            or f"Участник {recipient_user_id}"
        )

        sender_tx = FinanceTransaction(
            chat_id=chat_id,
            telegram_user_id=sender_user_id,
            member_name=sender_name,
            kind="expense",
            amount=amount,
            currency=currency,
            category=TRANSFER_OUT_CATEGORY,
            description=f"Перевод {recipient_name}",
            is_family=False,
            source_type="text",
            original_text=source_text,
            transcript=None,
            message_id=source_message_id,
        )
        recipient_tx = FinanceTransaction(
            chat_id=chat_id,
            telegram_user_id=recipient_user_id,
            member_name=recipient_name,
            kind="income",
            amount=amount,
            currency=currency,
            category=TRANSFER_IN_CATEGORY,
            description=f"Перевод от {sender_name}",
            is_family=False,
            source_type="text",
            original_text=source_text,
            transcript=None,
            message_id=source_message_id,
        )
        sender_tx_id = self.db.add_transaction(sender_tx)
        recipient_tx_id = self.db.add_transaction(recipient_tx)
        if sender_tx_id:
            try:
                self.db.mark_user_activated(sender_user_id)
            except Exception:
                logger.exception(
                    "Failed to mark sender as activated after transfer: user_id=%s",
                    sender_user_id,
                )
        return sender_name, recipient_name, sender_tx_id, recipient_tx_id

    async def _try_apply_pending_transfer_input(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: str,
    ) -> bool:
        pending = context.user_data.get(TRANSFER_PENDING_KEY)
        if not isinstance(pending, dict):
            return False

        chat = update.effective_chat
        user = update.effective_user
        message = update.effective_message
        if not chat or not user or not message:
            return False

        if int(pending.get("chat_id") or 0) != chat.id or int(
            pending.get("sender_user_id") or 0
        ) != user.id:
            return False

        text_l = text.strip().lower()
        if text_l in {"отмена", "cancel"}:
            context.user_data.pop(TRANSFER_PENDING_KEY, None)
            await message.reply_text("Хорошо, перевод отменил.")
            return True

        recipient_user_id = self._resolve_transfer_recipient(
            chat_id=chat.id,
            sender_user_id=user.id,
            text=text,
        )
        if recipient_user_id is None:
            await message.reply_text(
                "Кому перевести?",
                reply_markup=self._transfer_recipient_keyboard(
                    chat_id=chat.id,
                    sender_user_id=user.id,
                ),
            )
            return True

        amount = float(pending.get("amount") or 0.0)
        currency = str(pending.get("currency") or self.settings.default_currency)
        source_text = str(pending.get("source_text") or text)
        source_message_id = int(pending.get("source_message_id") or message.message_id)
        sender_name, recipient_name, sender_tx_id, _recipient_tx_id = self._record_internal_transfer(
            chat_id=chat.id,
            sender_user_id=user.id,
            recipient_user_id=recipient_user_id,
            amount=amount,
            currency=currency,
            source_text=source_text,
            source_message_id=source_message_id,
        )
        context.user_data.pop(TRANSFER_PENDING_KEY, None)
        await message.reply_text(
            self._format_transfer_saved_text(
                amount=amount,
                currency=currency,
                sender_name=sender_name,
                recipient_name=recipient_name,
            ),
            reply_markup=self._transfer_actions_keyboard(sender_tx_id, user.id),
        )
        return True

    async def _try_handle_transfer_message(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: str,
    ) -> bool:
        chat = update.effective_chat
        user = update.effective_user
        message = update.effective_message
        if not chat or not user or not message:
            return False

        if not self._is_transfer_intent(text):
            return False

        amount = _extract_amount_value(text)
        if amount is None:
            await message.reply_text("Не вижу сумму перевода. Напиши, пожалуйста, сумму.")
            return True

        currency = self._detect_currency_from_text(text, self.settings.default_currency)
        recipient_user_id = self._resolve_transfer_recipient(
            chat_id=chat.id,
            sender_user_id=user.id,
            text=text,
        )

        if recipient_user_id is None:
            candidates = self._transfer_recipient_candidates(chat.id, user.id)
            if not candidates:
                await message.reply_text("Пока не вижу второго участника для перевода.")
                return True
            context.user_data[TRANSFER_PENDING_KEY] = {
                "chat_id": chat.id,
                "sender_user_id": user.id,
                "amount": amount,
                "currency": currency,
                "source_text": text,
                "source_message_id": message.message_id,
            }
            await message.reply_text(
                "Кому перевести?",
                reply_markup=self._transfer_recipient_keyboard(
                    chat_id=chat.id,
                    sender_user_id=user.id,
                ),
            )
            return True

        sender_name, recipient_name, sender_tx_id, _recipient_tx_id = self._record_internal_transfer(
            chat_id=chat.id,
            sender_user_id=user.id,
            recipient_user_id=recipient_user_id,
            amount=amount,
            currency=currency,
            source_text=text,
            source_message_id=message.message_id,
        )
        await message.reply_text(
            self._format_transfer_saved_text(
                amount=amount,
                currency=currency,
                sender_name=sender_name,
                recipient_name=recipient_name,
            ),
            reply_markup=self._transfer_actions_keyboard(sender_tx_id, user.id),
        )
        return True

    async def transfer_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            return

        data = query.data or ""
        if not data.startswith("tf:"):
            await query.answer()
            return

        pending = context.user_data.get(TRANSFER_PENDING_KEY)
        if not isinstance(pending, dict):
            await query.answer("Перевод уже обработан.", show_alert=True)
            return

        if int(pending.get("chat_id") or 0) != chat.id or int(
            pending.get("sender_user_id") or 0
        ) != user.id:
            await query.answer("Этот выбор не для вас.", show_alert=True)
            return

        parts = data.split(":")
        action = parts[1] if len(parts) > 1 else ""
        if action == "cancel":
            context.user_data.pop(TRANSFER_PENDING_KEY, None)
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    logger.exception("Failed to delete transfer selector message")
            await query.answer("Отменено")
            return

        if action != "to" or len(parts) < 3:
            await query.answer()
            return

        try:
            recipient_user_id = int(parts[2])
        except ValueError:
            await query.answer("Неверный получатель.", show_alert=True)
            return
        if recipient_user_id == user.id:
            await query.answer("Нельзя переводить самому себе.", show_alert=True)
            return

        amount = float(pending.get("amount") or 0.0)
        currency = str(pending.get("currency") or self.settings.default_currency)
        source_text = str(pending.get("source_text") or "Перевод внутри семьи")
        source_message_id = int(pending.get("source_message_id") or 0)
        sender_name, recipient_name, sender_tx_id, _recipient_tx_id = self._record_internal_transfer(
            chat_id=chat.id,
            sender_user_id=user.id,
            recipient_user_id=recipient_user_id,
            amount=amount,
            currency=currency,
            source_text=source_text,
            source_message_id=source_message_id,
        )
        context.user_data.pop(TRANSFER_PENDING_KEY, None)
        if query.message:
            try:
                await query.message.delete()
            except Exception:
                logger.exception("Failed to delete transfer selector after save")
        await context.bot.send_message(
            chat_id=chat.id,
            text=self._format_transfer_saved_text(
                amount=amount,
                currency=currency,
                sender_name=sender_name,
                recipient_name=recipient_name,
            ),
            reply_markup=self._transfer_actions_keyboard(sender_tx_id, user.id),
        )
        await query.answer()

    async def _remember_member(self, update: Update) -> None:
        chat = update.effective_chat
        user = update.effective_user
        if not chat or not user:
            return
        if chat.type == "private" and not self._workspace_exists(chat.id):
            return
        self.db.register_member(
            chat_id=chat.id,
            telegram_user_id=user.id,
            username=user.username,
            full_name=_display_name(update),
        )

    async def _get_can_read_all_group_messages(
        self, context: ContextTypes.DEFAULT_TYPE
    ) -> bool | None:
        if self._can_read_all_groups is not None:
            return self._can_read_all_groups
        try:
            me = await context.bot.get_me()
            value = getattr(me, "can_read_all_group_messages", None)
            if isinstance(value, bool):
                self._can_read_all_groups = value
            else:
                self._can_read_all_groups = None
        except Exception:
            logger.exception("Failed to query bot privacy mode")
            self._can_read_all_groups = None
        return self._can_read_all_groups

    async def _warn_if_privacy_mode_blocks_messages(
        self,
        chat_id: int,
        chat_type: str | None,
        message,
        context: ContextTypes.DEFAULT_TYPE,
    ) -> None:
        if chat_type not in {"group", "supergroup"}:
            return
        if chat_id in self._privacy_warned_chats:
            return
        can_read = await self._get_can_read_all_group_messages(context)
        if can_read is False:
            self._privacy_warned_chats.add(chat_id)
            await message.reply_text(
                "Я вижу в группе только команды и ответы на мои сообщения. "
                "Чтобы я ловил обычные траты/доходы, выключите privacy mode в BotFather: "
                "/setprivacy -> Disable."
            )

    async def _sync_members_from_admins(
        self, chat_id: int, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        try:
            admins = await context.bot.get_chat_administrators(chat_id)
        except Exception:
            logger.exception("Failed to load chat administrators for onboarding")
            return

        for admin in admins:
            user = admin.user
            if user.is_bot:
                continue
            full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or (
                user.username or str(user.id)
            )
            self.db.register_member(
                chat_id=chat_id,
                telegram_user_id=user.id,
                username=user.username,
                full_name=full_name,
            )

    def _intro_text(self, chat_id: int) -> str:
        lines = [
            "Привет! Я Семейный финансовый ассистент.",
            "Давайте познакомимся, чтобы я вел статистику по именам.",
        ]

        members = self.db.list_members(chat_id)
        missing_names = [m for m in members if not m.get("custom_name")]

        if missing_names:
            lines.append("")
            for member in missing_names:
                lines.append(f"{_member_handle(member)}, как к вам обращаться?")
            lines.append("Можно просто написать имя одним сообщением.")
        else:
            lines.append("")
            lines.append("Супер, все имена уже знаю.")

        lines.extend(
            [
                "",
                "Дальше просто пишите расходы/доходы текстом или голосом.",
                "Примеры:",
                "- Купил кофе за 20000",
                "- Получил зарплату 120000",
                "",
                "Команды:",
                "/stats - сводка по участникам за месяц",
                "/mystats - ваша личная статистика",
                "/familystats - семейные расходы",
                "/report - ежедневный отчет по семье",
                "/members - кто уже представился",
            ]
        )
        return "\n".join(lines)

    async def _prompt_name_if_needed(self, update: Update, context: ContextTypes.DEFAULT_TYPE | None = None) -> None:
        chat = update.effective_chat
        user = update.effective_user
        message = update.effective_message
        if not chat or not user or not message:
            return

        if chat.type == "private":
            return

        if context and self._get_personal_onboarding_step(context) > 0:
            return

        custom_name = self.db.get_custom_name(chat.id, user.id)
        if custom_name:
            return

        key = (chat.id, user.id)
        if key in self.name_prompted:
            return

        self.name_prompted.add(key)
        await message.reply_text(
            f"{_user_handle(user)}, подскажи, как к тебе обращаться? Можно просто написать имя."
        )

    async def start_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        await self._start_command_impl(update, context)

    async def _start_command_impl(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        *,
        skip_language_gate: bool = False,
        start_payload_override: str | None = None,
    ) -> None:
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat:
            return
        await self._remember_member(update)

        user_lang = DEFAULT_LANGUAGE
        if user:
            try:
                self.db.register_user(
                    telegram_user_id=user.id,
                    username=user.username,
                    first_name=user.first_name,
                    last_name=user.last_name,
                )
            except Exception:
                logger.exception("Failed to register user on /start: user_id=%s", user.id)
            user_lang = self._user_language(user.id)

            if not skip_language_gate and not self._user_language_selected(user.id):
                pending_payload = ""
                if chat.type == "private":
                    if start_payload_override is not None:
                        pending_payload = start_payload_override
                    elif context.args:
                        pending_payload = str(context.args[0] or "")
                self._set_pending_start_after_language(
                    context,
                    chat_id=chat.id,
                    chat_type=chat.type,
                    start_payload=pending_payload,
                )
                await message.reply_text(
                    t("language_prompt_combined", user_lang),
                    reply_markup=self._language_picker_keyboard(
                        callback_prefix="lang:init",
                        ui_lang=user_lang,
                        current_lang=user_lang,
                    ),
                )
                return

        lang = self._user_language(user.id) if user else user_lang

        if chat.type == "private":
            resolved_workspace = self.db.resolve_workspace(user.id, chat.id, "private") if user else None
            if resolved_workspace and str(resolved_workspace.get("type") or "") == "personal" and user:
                await self._send_personal_mode_activated(
                    chat_id=chat.id,
                    user_id=user.id,
                    context=context,
                    lang=lang,
                )
                return

            await self._send_private_mode_picker(update, context=context, lang=lang)
            return

        if chat.type not in {"group", "supergroup"}:
            return

        group_workspace = self.db.get_workspace_by_chat_id(chat.id)
        if group_workspace is None:
            try:
                workspace_id = self.db.create_workspace(
                    workspace_type="family",
                    title=str(getattr(chat, "title", "") or "").strip() or "Family workspace",
                    created_by=user.id if user and not user.is_bot else None,
                )
                self.db.bind_workspace_chat(workspace_id=workspace_id, telegram_chat_id=chat.id)
                if user and not user.is_bot:
                    self.db.add_workspace_member(
                        workspace_id=workspace_id,
                        telegram_user_id=user.id,
                        role="owner",
                    )
                    self.db.register_member(
                        chat_id=chat.id,
                        telegram_user_id=user.id,
                        username=user.username,
                        full_name=_display_name(update),
                    )
            except Exception:
                logger.exception(
                    "Failed to create family workspace on group /start: chat_id=%s user_id=%s",
                    chat.id,
                    user.id if user else None,
                )

        onboarding = self._get_group_onboarding_state(context)
        await self._sync_members_from_admins(chat.id, context)
        is_admin = await self._is_bot_admin_in_group(chat.id, context)
        if not is_admin:
            existing_admin_prompt = onboarding.get("admin_prompt_message_id")
            if isinstance(existing_admin_prompt, int) and existing_admin_prompt > 0:
                return
            sent = await message.reply_text(
                t("start_group_admin_request", lang),
                reply_markup=self._group_admin_ready_keyboard(lang),
            )
            onboarding["admin_prompt_message_id"] = sent.message_id
            self._track_onboarding_message_id(onboarding, sent.message_id)
            onboarding["admin_confirmed"] = False
            return

        if bool(onboarding.get("final_sent")):
            await message.reply_text(
                t("quick_menu_restored", lang),
                reply_markup=self._group_main_reply_keyboard(lang=lang),
            )
            await self._send_group_family_mode_activated(
                chat_id=chat.id,
                context=context,
                lang=lang,
            )
            return
        onboarding["admin_confirmed"] = True
        onboarding["admin_prompt_message_id"] = None
        await self._auto_set_member_names(chat.id, context)
        await self._send_group_family_onboarding(
            chat_id=chat.id,
            context=context,
            lang=lang,
        )
        onboarding["intro_sent"] = True
        onboarding["onboarding_step"] = "try_it"

    async def start_mode_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            return

        data = str(query.data or "")
        if data not in {"mode_personal", "mode_family"}:
            await self._safe_query_answer(query)
            return

        lang = self._user_language(user.id)
        if chat.type != "private":
            await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return

        if data == "mode_personal":
            try:
                resolved = self.db.resolve_workspace(user.id, chat.id, "private")
                if resolved is None:
                    workspace_id = self.db.create_workspace(
                        workspace_type="personal",
                        title=(user.full_name or user.username or f"User {user.id}"),
                        created_by=user.id,
                    )
                    self.db.bind_workspace_chat(workspace_id=workspace_id, telegram_chat_id=chat.id)
                    self.db.add_workspace_member(
                        workspace_id=workspace_id,
                        telegram_user_id=user.id,
                        role="owner",
                    )
                    for rtype in ("daily", "weekly", "monthly"):
                        self.db.upsert_workspace_report_schedule(
                            workspace_id=workspace_id,
                            report_type=rtype,
                            enabled=False,
                            timezone=DEFAULT_REPORT_TIMEZONE,
                            send_hour=DEFAULT_DAILY_REPORT_HOUR,
                            send_minute=DEFAULT_DAILY_REPORT_MINUTE,
                        )
                else:
                    self.db.add_workspace_member(
                        workspace_id=int(resolved.get("workspace_id") or 0),
                        telegram_user_id=user.id,
                        role="owner",
                    )
                self.db.register_member(
                    chat_id=chat.id,
                    telegram_user_id=user.id,
                    username=user.username,
                    full_name=_display_name(update),
                )
            except Exception:
                logger.exception(
                    "Failed to create personal workspace on mode selection: chat_id=%s user_id=%s",
                    chat.id,
                    user.id,
                )
                await self._safe_query_answer(query, t("error_generic", lang), show_alert=True)
                return

            await self._safe_query_answer(query)
            await self._cleanup_onboarding_messages(chat.id, context)
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    try:
                        await query.message.edit_reply_markup(reply_markup=None)
                    except Exception:
                        pass
            await self._send_personal_onboarding_step1(chat.id, context, lang)
            return

        await self._safe_query_answer(query)
        await self._cleanup_onboarding_messages(chat.id, context)
        if query.message:
            try:
                await query.message.delete()
            except Exception:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
        add_group_url = await self._build_add_group_url(context)
        ui_lang = normalize_language(lang)
        await context.bot.send_message(
            chat_id=chat.id,
            text=t("start_family_mode_group_hint", lang),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    {"ru": "➕ Добавить в группу",
                     "uz": "➕ Guruhga qo'shish",
                     "en": "➕ Add to Group"}.get(ui_lang, "➕ Add to Group"),
                    url=add_group_url,
                )],
            ]),
        )

    async def help_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        await self.start_command(update, context)

    async def admin_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        user = update.effective_user
        if not message or not user:
            return

        admin_id = int(getattr(self.settings, "admin_telegram_id", 0) or 0)
        if admin_id <= 0 or user.id != admin_id:
            return

        try:
            total_users = self.db.get_users_count()
            active_users_7d = self.db.get_active_users_count(days=7)
            activated_users = self.db.get_users_with_transactions_count()
            inactive_users = max(0, total_users - activated_users)
            total_transactions = self.db.get_transactions_count()
            transactions_7d = self.db.get_transactions_count_last_days(days=7)
            review_stats = self.db.get_bot_review_rating_stats()
        except Exception:
            logger.exception("Failed to load admin stats")
            await message.reply_text("Не удалось получить статистику.")
            return

        avg_current_rating = review_stats.get("avg_current_rating")
        rated_users = int(review_stats.get("rated_users") or 0)
        total_reviews = int(review_stats.get("total_reviews") or 0)
        changed_low_to_high_users = int(review_stats.get("changed_low_to_high_users") or 0)
        changed_high_to_low_users = int(review_stats.get("changed_high_to_low_users") or 0)
        if isinstance(avg_current_rating, (int, float)) and rated_users > 0:
            rating_line = (
                f"⭐ Рейтинг бота: {float(avg_current_rating):.2f}/5 "
                f"(текущие оценки, {rated_users} польз.)"
            )
        else:
            rating_line = "⭐ Рейтинг бота: пока нет оценок"

        text = (
            "📊 Статистика бота\n\n"
            f"👥 Всего пользователей: {total_users}\n"
            f"🔥 Активных (7 дней): {active_users_7d}\n"
            f"💰 Добавляли операции: {activated_users}\n"
            f"😴 Только старт: {inactive_users}\n\n"
            f"📈 Всего транзакций: {total_transactions}\n"
            f"📆 За 7 дней: {transactions_7d}\n\n"
            f"{rating_line}\n"
            f"📝 Всего отзывов: {total_reviews}\n"
            f"⬆️ Кол-во пользователей которые изменили отзыв с маленького на высокий: "
            f"{changed_low_to_high_users}\n"
            f"⬇️ Кол-во пользователей которые изменили отзыв с высокого на маленький: "
            f"{changed_high_to_low_users}"
        )
        await message.reply_text(text)

    @staticmethod
    def _bot_review_prompt_text() -> str:
        return "⭐ Оценить бота\n\nВыберите оценку:"

    @staticmethod
    def _bot_review_existing_text(rating: int) -> str:
        return f"⭐ Ваша оценка: {int(rating)}\nХотите изменить?"

    @staticmethod
    def _bot_review_comment_prompt_text() -> str:
        return (
            "Спасибо! Хотите оставить комментарий? "
            "Напишите сообщением или нажмите «Пропустить»."
        )

    @staticmethod
    def _format_tashkent_time_from_utc_text(value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return "Asia/Tashkent"
        try:
            utc_dt = datetime.strptime(raw, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            tashkent_tz = timezone(timedelta(hours=5), name="Asia/Tashkent")
            local_dt = utc_dt.astimezone(tashkent_tz)
            return f"{local_dt.strftime('%Y-%m-%d %H:%M:%S')} Asia/Tashkent"
        except ValueError:
            return raw

    def _bot_review_pending(self, context: ContextTypes.DEFAULT_TYPE) -> dict[str, object] | None:
        raw = context.user_data.get(BOT_REVIEW_PENDING_KEY)
        return raw if isinstance(raw, dict) else None

    def _bot_review_clear_pending(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop(BOT_REVIEW_PENDING_KEY, None)

    def _bot_review_store_pending(
        self,
        *,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
        telegram_user_id: int,
        prompt_chat_id: int,
        prompt_message_id: int,
        rating: int | None = None,
    ) -> None:
        pending = self._bot_review_pending(context) or {}
        pending.update(
            {
                "chat_id": int(chat_id),
                "telegram_user_id": int(telegram_user_id),
                "prompt_chat_id": int(prompt_chat_id),
                "prompt_message_id": int(prompt_message_id),
            }
        )
        if rating is None:
            pending.pop("rating", None)
        else:
            pending["rating"] = int(rating)
        context.user_data[BOT_REVIEW_PENDING_KEY] = pending

    async def _bot_review_finish_prompt_message(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: str,
        reply_markup: InlineKeyboardMarkup | None = None,
    ) -> None:
        pending = self._bot_review_pending(context) or {}
        prompt_chat_id = int(pending.get("prompt_chat_id") or 0)
        prompt_message_id = int(pending.get("prompt_message_id") or 0)
        if prompt_chat_id > 0 and prompt_message_id > 0:
            try:
                await context.bot.edit_message_text(
                    chat_id=prompt_chat_id,
                    message_id=prompt_message_id,
                    text=text,
                    reply_markup=reply_markup,
                )
                return
            except Exception:
                logger.exception(
                    "Failed to edit bot review prompt message: chat_id=%s message_id=%s",
                    prompt_chat_id,
                    prompt_message_id,
                )
        message = update.effective_message
        if message:
            await message.reply_text(text, reply_markup=reply_markup)

    def _bot_review_sender_name(self, update: Update) -> str:
        user = update.effective_user
        chat = update.effective_chat
        if user and chat and chat.type in {"group", "supergroup"}:
            try:
                custom_name = self.db.get_custom_name(chat.id, user.id)
            except Exception:
                logger.exception(
                    "Failed to read custom name for bot review notification: chat_id=%s user_id=%s",
                    chat.id,
                    user.id,
                )
                custom_name = None
            if isinstance(custom_name, str) and custom_name.strip():
                return custom_name.strip()
        return _display_name(update)

    async def _notify_admin_about_bot_review(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        review: dict[str, object],
        previous_review: dict[str, object] | None = None,
    ) -> None:
        admin_chat_id = self._support_admin_chat_id()
        if admin_chat_id == 0:
            logger.warning("Support admin chat is not configured; bot review notification skipped")
            return

        user_id = int(review.get("telegram_user_id") or 0)
        rating = int(review.get("rating") or 0)
        comment = str(review.get("comment") or "").strip() or "(без комментария)"
        created_at = str(review.get("created_at") or "").strip()
        created_at_label = self._format_tashkent_time_from_utc_text(created_at)
        prev_rating_raw = previous_review.get("rating") if isinstance(previous_review, dict) else None
        try:
            prev_rating = int(prev_rating_raw) if prev_rating_raw is not None else None
        except (TypeError, ValueError):
            prev_rating = None
        is_rating_changed = bool(prev_rating in {1, 2, 3, 4, 5} and prev_rating != rating)

        if is_rating_changed and prev_rating is not None:
            payload = (
                "#change_rate\n"
                "⭐ Изменили отзыв\n"
                f"👤 Имя: {self._bot_review_sender_name(update)}\n"
                f"🆔 ID: {user_id}\n"
                f"⭐ Была Оценка: {prev_rating}/5\n"
                f"⭐ Стала Оценка: {rating}/5\n"
                f"💬 Комментарий: {comment}\n"
                f"🕒 Время: {created_at_label}"
            )
        else:
            payload = (
                "#rate\n"
                "⭐ Новый отзыв о боте\n"
                f"👤 Имя: {self._bot_review_sender_name(update)}\n"
                f"🆔 ID: {user_id}\n"
                f"⭐ Оценка: {rating}/5\n"
                f"💬 Комментарий: {comment}\n"
                f"🕒 Время: {created_at_label}"
            )
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=payload)
        except Exception:
            logger.exception(
                "Failed to send bot review notification to admin chat: admin_chat_id=%s user_id=%s",
                admin_chat_id,
                user_id,
            )

    async def _save_bot_review_and_finish(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        comment: str | None,
    ) -> int:
        pending = self._bot_review_pending(context)
        user = update.effective_user
        chat = update.effective_chat
        if not pending or not user or not chat:
            self._bot_review_clear_pending(context)
            lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE
            await self._bot_review_finish_prompt_message(
                update=update,
                context=context,
                text="Ок, отменено.",
                reply_markup=self._assistant_settings_back_keyboard(lang),
            )
            return ConversationHandler.END

        try:
            rating = int(pending.get("rating") or 0)
        except (TypeError, ValueError):
            rating = 0
        if rating < 1 or rating > 5:
            self._bot_review_clear_pending(context)
            lang = self._user_language(user.id)
            await self._bot_review_finish_prompt_message(
                update=update,
                context=context,
                text="Ок, отменено.",
                reply_markup=self._assistant_settings_back_keyboard(lang),
            )
            return ConversationHandler.END

        previous_review: dict[str, object] | None = None
        try:
            previous_review = self.db.get_latest_bot_review(user.id)
        except Exception:
            logger.exception("Failed to load previous bot review before save: user_id=%s", user.id)

        try:
            review = self.db.add_bot_review(
                chat_id=int(pending.get("chat_id") or chat.id),
                telegram_user_id=user.id,
                rating=rating,
                comment=comment,
                created_at_utc=datetime.utcnow(),
            )
        except Exception:
            logger.exception("Failed to save bot review: user_id=%s chat_id=%s", user.id, chat.id)
            self._bot_review_clear_pending(context)
            lang = self._user_language(user.id)
            await self._bot_review_finish_prompt_message(
                update=update,
                context=context,
                text="Не удалось сохранить отзыв. Попробуйте позже.",
                reply_markup=self._assistant_settings_back_keyboard(lang),
            )
            return ConversationHandler.END

        await self._notify_admin_about_bot_review(
            update=update,
            context=context,
            review=review,
            previous_review=previous_review,
        )
        prompt_chat_id = int(pending.get("prompt_chat_id") or chat.id)
        prompt_message_id = int(pending.get("prompt_message_id") or 0)
        query = update.callback_query
        user_message = update.effective_message
        if user_message and query is None:
            try:
                await context.bot.delete_message(
                    chat_id=user_message.chat_id,
                    message_id=user_message.message_id,
                )
            except Exception:
                logger.exception(
                    "Failed to delete bot review user comment message: chat_id=%s message_id=%s",
                    user_message.chat_id,
                    user_message.message_id,
                )
        if prompt_chat_id > 0 and prompt_message_id > 0:
            try:
                await context.bot.delete_message(
                    chat_id=prompt_chat_id,
                    message_id=prompt_message_id,
                )
            except Exception:
                logger.exception(
                    "Failed to delete bot review prompt message: chat_id=%s message_id=%s",
                    prompt_chat_id,
                    prompt_message_id,
                )
        try:
            lang = self._user_language(user.id)
            await context.bot.send_message(
                chat_id=prompt_chat_id if prompt_chat_id > 0 else chat.id,
                text="✅ Спасибо за отзыв!",
                reply_markup=self._assistant_settings_back_keyboard(lang),
            )
        except Exception:
            logger.exception("Failed to send bot review thank-you message")
        self._bot_review_clear_pending(context)
        return ConversationHandler.END

    async def bot_review_start_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            return ConversationHandler.END

        await self._safe_query_answer(query)
        self._bot_review_clear_pending(context)
        latest_review: dict[str, object] | None = None
        try:
            latest_review = self.db.get_latest_bot_review(user.id)
        except Exception:
            logger.exception("Failed to load latest bot review: user_id=%s", user.id)

        has_existing_rating = latest_review is not None and int(latest_review.get("rating") or 0) in {1, 2, 3, 4, 5}
        prompt_text = (
            self._bot_review_existing_text(int(latest_review["rating"]))
            if has_existing_rating and latest_review is not None
            else self._bot_review_prompt_text()
        )
        prompt_keyboard = (
            self._bot_review_existing_keyboard()
            if has_existing_rating
            else self._bot_review_rating_keyboard()
        )
        try:
            if query.message:
                await query.edit_message_text(
                    prompt_text,
                    reply_markup=prompt_keyboard,
                )
                self._bot_review_store_pending(
                    context=context,
                    chat_id=chat.id,
                    telegram_user_id=user.id,
                    prompt_chat_id=query.message.chat_id,
                    prompt_message_id=query.message.message_id,
                    rating=(
                        int(latest_review["rating"])
                        if has_existing_rating and latest_review is not None
                        else None
                    ),
                )
                return BOT_REVIEW_MENU if has_existing_rating else BOT_REVIEW_RATING
        except Exception:
            logger.exception("Failed to render bot review start prompt in-place")

        if query.message:
            sent = await query.message.reply_text(
                prompt_text,
                reply_markup=prompt_keyboard,
            )
            self._bot_review_store_pending(
                context=context,
                chat_id=chat.id,
                telegram_user_id=user.id,
                prompt_chat_id=sent.chat_id,
                prompt_message_id=sent.message_id,
                rating=(
                    int(latest_review["rating"])
                    if has_existing_rating and latest_review is not None
                    else None
                ),
            )
            return BOT_REVIEW_MENU if has_existing_rating else BOT_REVIEW_RATING
        return ConversationHandler.END

    async def bot_review_menu_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            self._bot_review_clear_pending(context)
            return ConversationHandler.END

        data = query.data or ""
        if data == "rv:cancel":
            return await self.bot_review_cancel_callback(update, context)

        pending = self._bot_review_pending(context) or {}
        if data == "rv:change":
            await self._safe_query_answer(query)
            if query.message:
                self._bot_review_store_pending(
                    context=context,
                    chat_id=chat.id,
                    telegram_user_id=user.id,
                    prompt_chat_id=query.message.chat_id,
                    prompt_message_id=query.message.message_id,
                    rating=None,
                )
                try:
                    await query.edit_message_text(
                        self._bot_review_prompt_text(),
                        reply_markup=self._bot_review_rating_keyboard(),
                    )
                    return BOT_REVIEW_RATING
                except Exception:
                    logger.exception("Failed to switch bot review flow to rating selection")
            return BOT_REVIEW_MENU

        if data == "rv:comment":
            await self._safe_query_answer(query)
            try:
                rating = int(pending.get("rating") or 0)
            except (TypeError, ValueError):
                rating = 0
            if rating < 1 or rating > 5:
                try:
                    latest = self.db.get_latest_bot_review(user.id)
                    rating = int(latest.get("rating") or 0) if latest else 0
                except Exception:
                    logger.exception("Failed to load latest rating for comment-only review flow")
                    rating = 0
            if rating < 1 or rating > 5:
                self._bot_review_clear_pending(context)
                if query.message:
                    await query.edit_message_text(
                        "Не нашёл вашу текущую оценку. Выберите оценку заново.",
                        reply_markup=self._bot_review_rating_keyboard(),
                    )
                    self._bot_review_store_pending(
                        context=context,
                        chat_id=chat.id,
                        telegram_user_id=user.id,
                        prompt_chat_id=query.message.chat_id,
                        prompt_message_id=query.message.message_id,
                    )
                    return BOT_REVIEW_RATING
                return ConversationHandler.END

            if query.message:
                self._bot_review_store_pending(
                    context=context,
                    chat_id=chat.id,
                    telegram_user_id=user.id,
                    prompt_chat_id=query.message.chat_id,
                    prompt_message_id=query.message.message_id,
                    rating=rating,
                )
                try:
                    await query.edit_message_text(
                        self._bot_review_comment_prompt_text(),
                        reply_markup=self._bot_review_comment_keyboard(),
                    )
                    return BOT_REVIEW_COMMENT
                except Exception:
                    logger.exception("Failed to switch bot review flow to comment prompt")
            return BOT_REVIEW_MENU

        await self._safe_query_answer(query)
        return BOT_REVIEW_MENU

    async def bot_review_rating_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            self._bot_review_clear_pending(context)
            return ConversationHandler.END

        data = query.data or ""
        if data == "rv:cancel":
            return await self.bot_review_cancel_callback(update, context)

        match = re.match(r"^rv:rate:([1-5])$", data)
        if not match:
            await self._safe_query_answer(query)
            return BOT_REVIEW_RATING
        rating = int(match.group(1))

        await self._safe_query_answer(query)
        if query.message:
            self._bot_review_store_pending(
                context=context,
                chat_id=chat.id,
                telegram_user_id=user.id,
                prompt_chat_id=query.message.chat_id,
                prompt_message_id=query.message.message_id,
                rating=rating,
            )
            try:
                await query.edit_message_text(
                    self._bot_review_comment_prompt_text(),
                    reply_markup=self._bot_review_comment_keyboard(),
                )
                return BOT_REVIEW_COMMENT
            except Exception:
                logger.exception("Failed to render bot review comment prompt")
                await query.message.reply_text(
                    self._bot_review_comment_prompt_text(),
                    reply_markup=self._bot_review_comment_keyboard(),
                )
                return BOT_REVIEW_COMMENT
        self._bot_review_clear_pending(context)
        return ConversationHandler.END

    async def bot_review_comment_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        if not query:
            self._bot_review_clear_pending(context)
            return ConversationHandler.END

        data = query.data or ""
        if data == "rv:cancel":
            return await self.bot_review_cancel_callback(update, context)
        if data != "rv:skip":
            await self._safe_query_answer(query)
            return BOT_REVIEW_COMMENT

        await self._safe_query_answer(query)
        return await self._save_bot_review_and_finish(update=update, context=context, comment=None)

    async def bot_review_comment_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        message = update.effective_message
        if not message or not message.text:
            return BOT_REVIEW_COMMENT

        text = message.text.strip()
        if not text:
            await message.reply_text(
                "Напишите комментарий одним сообщением или нажмите «Пропустить»."
            )
            return BOT_REVIEW_COMMENT

        return await self._save_bot_review_and_finish(
            update=update,
            context=context,
            comment=text,
        )

    async def bot_review_cancel_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE
        if not query:
            self._bot_review_clear_pending(context)
            return ConversationHandler.END
        await self._safe_query_answer(query)
        self._bot_review_clear_pending(context)
        try:
            if query.message:
                await query.edit_message_text(
                    "Ок, отменено.",
                    reply_markup=self._assistant_settings_back_keyboard(lang),
                )
                return ConversationHandler.END
        except Exception:
            logger.exception("Failed to cancel bot review flow")
        if query.message:
            await query.message.reply_text(
                "Ок, отменено.",
                reply_markup=self._assistant_settings_back_keyboard(lang),
            )
        return ConversationHandler.END

    async def support_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        message = update.effective_message
        user = update.effective_user
        if not message:
            return ConversationHandler.END
        await message.reply_text(
            "Поддержка\n\nВыберите, что хотите отправить:",
            reply_markup=self._support_menu_keyboard(),
        )
        return SUPPORT_MENU

    async def support_menu_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query:
            return ConversationHandler.END

        data = query.data or ""
        await query.answer()

        if data == "sup:cancel":
            if query.message:
                try:
                    await query.edit_message_text("Поддержка отменена.")
                except Exception:
                    logger.exception("Failed to edit support cancel message")
            return ConversationHandler.END

        if data == "sup:message":
            if query.message:
                try:
                    await query.edit_message_text(
                        "📝 Напишите сообщение разработчику одним сообщением."
                    )
                except Exception:
                    logger.exception("Failed to show support message prompt")
            return SUPPORT_MESSAGE

        if data == "sup:bug":
            if query.message:
                try:
                    await query.edit_message_text(
                        "🐞 Опишите ошибку текстом или отправьте фото.\n"
                        "Если отправляете фото, добавьте описание в подписи."
                    )
                except Exception:
                    logger.exception("Failed to show support bug prompt")
            return SUPPORT_BUG

        return SUPPORT_MENU

    async def support_direct_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        if not query:
            return ConversationHandler.END

        data = query.data or ""
        await self._safe_query_answer(query)

        if data in {"profile_support_message", "ast:support_message"}:
            if query.message:
                try:
                    await query.edit_message_text(
                        "📝 Напишите сообщение разработчику одним сообщением."
                    )
                except Exception:
                    logger.exception("Failed to show support message prompt from legacy callback")
            return SUPPORT_MESSAGE

        if data in {"profile_support_bug", "ast:support_bug"}:
            if query.message:
                try:
                    await query.edit_message_text(
                        "🐞 Опишите ошибку текстом или отправьте фото.\n"
                        "Если отправляете фото, добавьте описание в подписи."
                    )
                except Exception:
                    logger.exception("Failed to show support bug prompt from legacy callback")
            return SUPPORT_BUG

        return ConversationHandler.END

    async def support_message_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return SUPPORT_MESSAGE

        self._touch_user_safely(user.id)
        admin_chat_id = self._support_admin_chat_id()
        if admin_chat_id == 0:
            await message.reply_text("Поддержка временно недоступна.")
            return ConversationHandler.END

        payload = (
            "#dev\n"
            "📩 Новое сообщение от пользователя\n\n"
            f"👤 Имя: {_display_name(update)}\n"
            f"🆔 ID: {user.id}\n"
            "💬 Текст:\n"
            f"{message.text.strip()}"
        )
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=payload)
        except Exception:
            logger.exception(
                "Failed to send support message to admin chat: admin_chat_id=%s user_id=%s",
                admin_chat_id,
                user.id,
            )
            await message.reply_text("Не удалось отправить сообщение. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text("Спасибо!")
        return ConversationHandler.END

    async def support_bug_text_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return SUPPORT_BUG

        self._touch_user_safely(user.id)
        admin_chat_id = self._support_admin_chat_id()
        if admin_chat_id == 0:
            await message.reply_text("Поддержка временно недоступна.")
            return ConversationHandler.END

        payload = (
            "#error\n"
            "🐞 Новый баг-репорт\n\n"
            f"👤 Имя: {_display_name(update)}\n"
            f"🆔 ID: {user.id}\n"
            "💬 Описание:\n"
            f"{message.text.strip()}"
        )
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=payload)
        except Exception:
            logger.exception(
                "Failed to send text bug report to admin chat: admin_chat_id=%s user_id=%s",
                admin_chat_id,
                user.id,
            )
            await message.reply_text("Не удалось отправить баг-репорт. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text("Спасибо! Баг-репорт отправил.")
        return ConversationHandler.END

    async def support_bug_photo_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.photo:
            return SUPPORT_BUG

        self._touch_user_safely(user.id)
        admin_chat_id = self._support_admin_chat_id()
        if admin_chat_id == 0:
            await message.reply_text("Поддержка временно недоступна.")
            return ConversationHandler.END

        description = (message.caption or "").strip() or "Фото без описания"
        caption = self._safe_telegram_caption(
            "#error\n"
            "🐞 Новый баг-репорт\n\n"
            f"👤 Имя: {_display_name(update)}\n"
            f"🆔 ID: {user.id}\n"
            "💬 Описание:\n"
            f"{description}"
        )
        try:
            await context.bot.send_photo(
                chat_id=admin_chat_id,
                photo=message.photo[-1].file_id,
                caption=caption,
            )
        except Exception:
            logger.exception(
                "Failed to send photo bug report to admin chat: admin_chat_id=%s user_id=%s",
                admin_chat_id,
                user.id,
            )
            await message.reply_text("Не удалось отправить баг-репорт. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text("Спасибо! Баг-репорт отправил.")
        return ConversationHandler.END

    async def profile_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        await self._show_profile_menu_message(
            update=update,
            context=context,
            from_assistant_settings=False,
        )

    async def profile_edit_start_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query or not user:
            return ConversationHandler.END

        print("PROFILE STATE:", "PROFILE_NAME")
        await self._safe_query_answer(query)

        if query.message:
            prior_ref = context.user_data.get(PROFILE_VIEW_REF_KEY)
            ref_payload = prior_ref.copy() if isinstance(prior_ref, dict) else {}
            ref_payload.update(
                {
                "chat_id": query.message.chat_id,
                "message_id": query.message.message_id,
                }
            )
            context.user_data[PROFILE_VIEW_REF_KEY] = ref_payload
        context.user_data[PROFILE_WIZARD_KEY] = {
            "started_at": datetime.utcnow().isoformat(timespec="seconds")
        }
        try:
            await query.edit_message_text(
                "Шаг 1/4 — Имя\n\nКак к вам обращаться?\n"
                "Напишите имя (можно псевдоним).",
                reply_markup=self._profile_wizard_cancel_keyboard(),
            )
        except Exception:
            logger.exception("Failed to start profile edit wizard")
            await self._safe_query_answer(query, "Ошибка запуска профиля", show_alert=True)
            return ConversationHandler.END
        return PROFILE_NAME

    async def profile_edit_cancel_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query:
            return ConversationHandler.END
        print("PROFILE STATE:", "CANCEL")
        await self._safe_query_answer(query)
        context.user_data.pop(PROFILE_WIZARD_KEY, None)
        if user:
            try:
                await self._refresh_profile_view_message(context=context, user=user)
            except Exception:
                logger.exception("Failed to restore profile after cancel")
        return ConversationHandler.END

    async def profile_wizard_interrupt_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        query = update.callback_query
        if not query:
            return ConversationHandler.END
        print("PROFILE STATE:", "INTERRUPT")
        context.user_data.pop(PROFILE_WIZARD_KEY, None)
        await self.assistant_settings_callback_handler(update, context)
        return ConversationHandler.END

    async def profile_menu_callback(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query:
            return ConversationHandler.END

        await self._safe_query_answer(query)
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE

        data = query.data or ""
        if data == "profile_close":
            context.user_data.pop(PROFILE_WIZARD_KEY, None)
            try:
                if query.message:
                    chat_id = query.message.chat_id
                    settings_message_id = query.message.message_id
                    trigger_message_id = self._pop_assistant_settings_trigger(
                        context=context,
                        settings_message_id=settings_message_id,
                    )
                    await query.message.delete()
                    if trigger_message_id:
                        try:
                            await context.bot.delete_message(
                                chat_id=chat_id,
                                message_id=trigger_message_id,
                            )
                        except Exception:
                            logger.exception(
                                "Failed to delete assistant settings trigger message on profile close: "
                                "chat_id=%s message_id=%s",
                                chat_id,
                                trigger_message_id,
                            )
                else:
                    await query.edit_message_text(t("profile_closed", lang))
            except Exception:
                logger.exception("Failed to close profile message")
                await self._safe_query_answer(
                    query,
                    {
                        "ru": "Ошибка при закрытии профиля",
                        "uz": "Profilni yopishda xatolik",
                        "en": "Error while closing profile",
                    }.get(lang, "Error while closing profile"),
                    show_alert=True,
                )
            return ConversationHandler.END

        if data == "profile_back":
            try:
                return await self._show_profile_menu_message(
                    update=update,
                    context=context,
                    edit_query_message=True,
                )
            except Exception:
                logger.exception("Failed to navigate back to profile menu")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
                return PROFILE_MENU

        if data == "profile_settings":
            try:
                current_currency = "UZS"
                current_language = lang
                if user:
                    try:
                        profile = self.db.get_user_profile(user.id)
                        current_currency = str(profile.get("currency") or "UZS").upper()
                        current_language = normalize_language(profile.get("language") or lang)
                    except Exception:
                        logger.exception("Failed to load profile settings values: user_id=%s", user.id)
                await query.edit_message_text(
                    t("profile_settings_title", lang),
                    reply_markup=self._profile_settings_keyboard(
                        lang=lang,
                        current_language=current_language,
                        current_currency=current_currency,
                    ),
                )
            except Exception:
                logger.exception("Failed to open profile settings screen")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        if data == "profile_support":
            try:
                await query.edit_message_text(
                    t("profile_support_title", lang),
                    reply_markup=self._profile_support_keyboard(lang),
                )
            except Exception:
                logger.exception("Failed to open profile support screen")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        if data == "profile_settings_lang":
            try:
                await query.edit_message_text(
                    self._assistant_language_screen_text(ui_lang=lang, current_lang=lang),
                    reply_markup=self._language_picker_keyboard(
                        callback_prefix="lang:pick_profile",
                        ui_lang=lang,
                        current_lang=lang,
                        back_callback="profile_settings",
                    ),
                )
            except Exception:
                logger.exception("Failed to open profile language picker")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        if data == "profile_settings_currency":
            try:
                await query.edit_message_text(
                    f"{t('assistant_currency_title', lang)}\n\n"
                    f"{t('assistant_currency_in_miniapp', lang)}",
                    reply_markup=self._profile_back_keyboard(lang),
                )
            except Exception:
                logger.exception("Failed to open profile currency settings item")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        if data in {"profile_support_message", "profile_support_bug"}:
            try:
                title = (
                    t("profile_support_dev", lang)
                    if data == "profile_support_message"
                    else t("profile_support_bug", lang)
                )
                await query.edit_message_text(
                    f"{title}\n\n"
                    + {
                        "ru": "Используйте команду /support и выберите нужный пункт.\nТак сообщение будет отправлено в канал поддержки.",
                        "uz": "/support buyrug'ini ishlating va kerakli bo'limni tanlang.\nShunda xabar yordam kanaliga yuboriladi.",
                        "en": "Use the /support command and choose the needed option.\nThis will send the message to the support channel.",
                    }.get(lang, "Use /support to send a message.")
                    ,
                    reply_markup=self._profile_back_keyboard(lang),
                )
            except Exception:
                logger.exception("Failed to open profile support item: %s", data)
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        if data == "profile_cancel":
            try:
                await query.edit_message_text(
                    {
                        "ru": "Редактирование профиля отменено.",
                        "uz": "Profilni tahrirlash bekor qilindi.",
                        "en": "Profile editing cancelled.",
                    }.get(lang, "Profile editing cancelled."),
                    reply_markup=self._profile_back_keyboard(lang),
                )
            except Exception:
                logger.exception("Failed to render profile cancel screen")
                await self._safe_query_answer(query, t("error_simple", lang), show_alert=True)
            return PROFILE_MENU

        return PROFILE_MENU

    async def language_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query or not user:
            return

        data = str(query.data or "")
        if not data.startswith("lang:"):
            await self._safe_query_answer(query)
            return
        await self._safe_query_answer(query)

        current_lang = self._user_language(user.id)
        parts = data.split(":")
        action = parts[1] if len(parts) > 1 else ""
        target_lang = normalize_language(parts[2] if len(parts) > 2 else current_lang)

        def _assistant_lang_picker() -> tuple[str, InlineKeyboardMarkup]:
            return (
                self._assistant_language_screen_text(ui_lang=current_lang, current_lang=current_lang),
                self._language_picker_keyboard(
                    callback_prefix="lang:pick_ast",
                    ui_lang=current_lang,
                    current_lang=current_lang,
                    back_callback="ast:menu",
                ),
            )

        if action == "init":
            try:
                new_lang = self._update_user_language(user, target_lang)
            except Exception:
                await self._safe_query_answer(query, t("error_generic", current_lang), show_alert=True)
                return

            changed_text = t(
                "language_changed",
                new_lang,
                language=language_short_label(new_lang),
            )
            pending_start = self._pop_pending_start_after_language(context)
            if query.message:
                try:
                    await query.edit_message_text(changed_text)
                    self._track_onboarding_message(context, query.message.message_id)
                except Exception:
                    logger.exception("Failed to edit init language selection message")
            if pending_start:
                start_payload = str(pending_start.get("start_payload") or "")
            else:
                start_payload = ""
            await self._start_command_impl(
                update,
                context,
                skip_language_gate=True,
                start_payload_override=start_payload,
            )
            return

        if action in {"pick_ast", "pick_profile"}:
            if target_lang == current_lang:
                if query.message:
                    try:
                        if action == "pick_profile":
                            await query.edit_message_text(
                                self._assistant_language_screen_text(
                                    ui_lang=current_lang,
                                    current_lang=current_lang,
                                ),
                                reply_markup=self._language_picker_keyboard(
                                    callback_prefix="lang:pick_profile",
                                    ui_lang=current_lang,
                                    current_lang=current_lang,
                                    back_callback="profile_settings",
                                ),
                            )
                        else:
                            text, kb = _assistant_lang_picker()
                            await query.edit_message_text(text, reply_markup=kb)
                    except Exception:
                        logger.exception("Failed to refresh language picker for same lang")
                return

            return_callback = "profile_settings_lang" if action == "pick_profile" else "ast:language"
            self._store_pending_language_change(
                context,
                source=action,
                target_lang=target_lang,
                return_callback=return_callback,
            )
            if query.message:
                try:
                    await query.edit_message_text(
                        f"{t('assistant_language_title', current_lang)}\n\n"
                        f"{t('language_change_confirm', current_lang, language=language_short_label(target_lang))}\n"
                        f"{t('language_change_confirm_note', current_lang)}",
                        reply_markup=self._language_confirm_keyboard(
                            ui_lang=current_lang,
                            target_lang=target_lang,
                        ),
                    )
                except Exception:
                    logger.exception("Failed to show language change confirmation")
                    await self._safe_query_answer(query, t("error_generic", current_lang), show_alert=True)
            return

        if action == "cancel":
            pending = self._get_pending_language_change(context)
            self._clear_pending_language_change(context)
            ui_lang = current_lang
            if query.message:
                try:
                    if pending and str(pending.get("source")) == "pick_profile":
                        await query.edit_message_text(
                            self._assistant_language_screen_text(
                                ui_lang=ui_lang,
                                current_lang=current_lang,
                            ),
                            reply_markup=self._language_picker_keyboard(
                                callback_prefix="lang:pick_profile",
                                ui_lang=ui_lang,
                                current_lang=current_lang,
                                back_callback="profile_settings",
                            ),
                        )
                    else:
                        text, kb = _assistant_lang_picker()
                        await query.edit_message_text(text, reply_markup=kb)
                except Exception:
                    logger.exception("Failed to cancel language change")
                    await self._safe_query_answer(query, t("error_generic", ui_lang), show_alert=True)
            return

        if action == "confirm":
            pending = self._get_pending_language_change(context)
            final_target = normalize_language(
                str((pending or {}).get("target_lang") or target_lang)
            )
            source = str((pending or {}).get("source") or "pick_ast")
            self._clear_pending_language_change(context)
            try:
                new_lang = self._update_user_language(user, final_target)
            except Exception:
                await self._safe_query_answer(query, t("error_generic", current_lang), show_alert=True)
                return

            # Refresh group reply keyboard so old RU quick buttons do not linger after language switch.
            msg_chat = query.message.chat if query.message else None
            if msg_chat and msg_chat.type in {"group", "supergroup"}:
                try:
                    await context.bot.send_message(
                        chat_id=msg_chat.id,
                        text={
                            "ru": "Язык меню обновлён. Быстрые кнопки ниже переключены.",
                            "uz": "Menyu tili yangilandi. Pastdagi tezkor tugmalar almashtirildi.",
                            "en": "Menu language updated. Quick buttons below were refreshed.",
                        }.get(new_lang, "Menu language updated."),
                        reply_markup=self._group_main_reply_keyboard(lang=new_lang),
                    )
                except Exception:
                    logger.exception("Failed to refresh group reply keyboard after language change")

            if not query.message:
                return
            try:
                if source == "pick_profile":
                    profile = self.db.get_user_profile(user.id)
                    currency = str(profile.get("currency") or "UZS").upper()
                    await query.edit_message_text(
                        t("profile_settings_title", new_lang),
                        reply_markup=self._profile_settings_keyboard(
                            lang=new_lang,
                            current_language=new_lang,
                            current_currency=currency,
                        ),
                    )
                else:
                    await query.edit_message_text(
                        self._assistant_settings_text(new_lang),
                        reply_markup=self._assistant_settings_keyboard(lang=new_lang),
                    )
                    await query.message.reply_text(
                        t(
                            "language_changed",
                            new_lang,
                            language=language_short_label(new_lang),
                        )
                    )
            except Exception:
                logger.exception("Failed to finalize language change")
                await self._safe_query_answer(query, t("error_generic", new_lang), show_alert=True)
            return

        await self._safe_query_answer(query, t("error_simple", current_lang), show_alert=True)

    async def assistant_settings_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        user = update.effective_user
        if not query:
            return

        data = query.data or ""
        if not data.startswith("ast:"):
            await self._safe_query_answer(query)
            return

        await self._safe_query_answer(query)
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE
        try:
            if data == "ast:close":
                if query.message:
                    chat_id = query.message.chat_id
                    settings_message_id = query.message.message_id
                    trigger_message_id = self._pop_assistant_settings_trigger(
                        context=context,
                        settings_message_id=settings_message_id,
                    )
                    await query.message.delete()
                    if trigger_message_id:
                        try:
                            await context.bot.delete_message(
                                chat_id=chat_id,
                                message_id=trigger_message_id,
                            )
                        except Exception:
                            logger.exception(
                                "Failed to delete assistant settings trigger message: "
                                "chat_id=%s message_id=%s",
                                chat_id,
                                trigger_message_id,
                            )
                return

            if data == "ast:menu":
                if query.message:
                    await query.edit_message_text(
                        self._assistant_settings_text(lang),
                        reply_markup=self._assistant_settings_keyboard(lang=lang),
                    )
                return

            if data == "ast:profile":
                await self._show_profile_menu_message(
                    update=update,
                    context=context,
                    edit_query_message=True,
                    from_assistant_settings=True,
                )
                return

            if data == "ast:language":
                if query.message:
                    await query.edit_message_text(
                        self._assistant_language_screen_text(ui_lang=lang, current_lang=lang),
                        reply_markup=self._language_picker_keyboard(
                            callback_prefix="lang:pick_ast",
                            ui_lang=lang,
                            current_lang=lang,
                            back_callback="ast:menu",
                        ),
                    )
                return

            if data == "ast:currency":
                if query.message:
                    current_currency = str(getattr(self.settings, "default_currency", "UZS") or "UZS").upper()
                    user = update.effective_user
                    if user:
                        try:
                            profile = self.db.get_user_profile(user.id)
                            current_currency = str(profile.get("currency") or current_currency).upper()
                        except Exception:
                            logger.exception("Failed to read current currency for assistant settings: user_id=%s", user.id)
                    await query.edit_message_text(
                        f"{t('assistant_currency_title', lang)}\n\n"
                        + {
                            "ru": f"Текущая валюта: {current_currency}\n",
                            "uz": f"Joriy valyuta: {current_currency}\n",
                            "en": f"Current currency: {current_currency}\n",
                        }.get(lang, f"Current currency: {current_currency}\n")
                        + t("assistant_currency_in_miniapp", lang),
                        reply_markup=self._assistant_settings_back_keyboard(lang),
                    )
                return

            if data == "ast:support":
                if query.message:
                    await query.edit_message_text(
                        t("assistant_support_title", lang),
                        reply_markup=self._assistant_support_menu_keyboard(lang),
                    )
                return

            if data in {"ast:support_message", "ast:support_bug"}:
                title = (
                    t("profile_support_dev", lang)
                    if data == "ast:support_message"
                    else t("profile_support_bug", lang)
                )
                if query.message:
                    await query.edit_message_text(
                        f"{title}\n\n"
                        + {
                            "ru": "Используйте команду /support, чтобы отправить сообщение в чат поддержки.",
                            "uz": "Yordam chatiga xabar yuborish uchun /support buyrug'idan foydalaning.",
                            "en": "Use /support to send a message to the support chat.",
                        }.get(lang, "Use /support to contact support."),
                        reply_markup=self._assistant_support_menu_keyboard(lang),
                    )
                return
        except Exception:
            logger.exception("Assistant settings callback failed: data=%s", data)
            await self._safe_query_answer(query, t("error_generic", lang), show_alert=True)

    async def profile_name_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return PROFILE_NAME

        self._touch_user_safely(user.id)
        print("PROFILE STATE:", "PROFILE_NAME")
        value = message.text.strip()
        if not value:
            await message.reply_text("Имя не должно быть пустым. Попробуйте ещё раз.")
            return PROFILE_NAME
        if self._looks_like_button_label_text(value):
            await message.reply_text("Введите имя текстом (например: Мурад).")
            return PROFILE_NAME
        if len(value) > 64:
            await message.reply_text("Имя слишком длинное. Напишите короче (до 64 символов).")
            return PROFILE_NAME

        try:
            self.db.update_user_profile_field(user.id, "display_name", value)
        except Exception:
            logger.exception("Failed to update profile display_name: user_id=%s", user.id)
            await message.reply_text("Не удалось сохранить имя. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text(
            "Шаг 2/4 — Телефон\n\nВведите номер в формате +998901234567",
            reply_markup=self._profile_wizard_cancel_keyboard(),
        )
        return PROFILE_PHONE

    async def profile_phone_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return PROFILE_PHONE

        self._touch_user_safely(user.id)
        print("PROFILE STATE:", "PROFILE_PHONE")
        value = message.text.strip()
        if not PROFILE_PHONE_RE.fullmatch(value):
            await message.reply_text("Введите номер в формате +998901234567")
            return PROFILE_PHONE

        try:
            self.db.update_user_profile_field(user.id, "phone", value)
        except Exception:
            logger.exception("Failed to update profile phone: user_id=%s", user.id)
            await message.reply_text("Не удалось сохранить телефон. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text(
            "Шаг 3/4 — Email\n\nВведите email (например: name@example.com)",
            reply_markup=self._profile_wizard_cancel_keyboard(),
        )
        return PROFILE_EMAIL

    async def profile_email_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return PROFILE_EMAIL

        self._touch_user_safely(user.id)
        print("PROFILE STATE:", "PROFILE_EMAIL")
        value = message.text.strip()
        if not PROFILE_EMAIL_RE.fullmatch(value):
            await message.reply_text("Введите корректный email, например: name@example.com")
            return PROFILE_EMAIL

        try:
            self.db.update_user_profile_field(user.id, "email", value)
        except Exception:
            logger.exception("Failed to update profile email: user_id=%s", user.id)
            await message.reply_text("Не удалось сохранить email. Попробуйте позже.")
            return ConversationHandler.END

        await message.reply_text(
            "Шаг 4/4 — Дата рождения\n\nВведите дату в формате DD-MM-YYYY (можно DD.MM.YYYY)",
            reply_markup=self._profile_wizard_cancel_keyboard(),
        )
        return PROFILE_BIRTHDATE

    async def profile_birth_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> int:
        message = update.effective_message
        user = update.effective_user
        if not message or not user or not message.text:
            return PROFILE_BIRTHDATE

        self._touch_user_safely(user.id)
        print("PROFILE STATE:", "PROFILE_BIRTHDATE")
        value = message.text.strip()
        parsed: datetime | None = None
        for fmt in ("%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(value, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            await message.reply_text("Введите дату в формате DD-MM-YYYY (или DD.MM.YYYY)")
            return PROFILE_BIRTHDATE
        stored_value = parsed.strftime("%Y-%m-%d")

        try:
            self.db.update_user_profile_field(user.id, "birth_date", stored_value)
            profile = self.db.get_user_profile(user.id)
            completed, total = self._profile_completion_from_profile(profile)
        except Exception:
            logger.exception("Failed to update profile birth_date: user_id=%s", user.id)
            await message.reply_text("Не удалось сохранить дату рождения. Попробуйте позже.")
            return ConversationHandler.END

        context.user_data.pop(PROFILE_WIZARD_KEY, None)
        try:
            await self._refresh_profile_view_message(context=context, user=user)
        except Exception:
            logger.exception("Failed to refresh profile screen after completion: user_id=%s", user.id)
        if completed >= total:
            await message.reply_text("🎉 Профиль полностью заполнен!\nСпасибо!")
        else:
            await message.reply_text(
                f"Профиль обновлён. Заполнено: {completed}/{total}\n"
                f"{self._profile_progress_dots(completed, total)}"
            )
        return ConversationHandler.END

    async def app_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return
        lang = self._user_language(user.id)

        if not await self._ensure_private_mode_selected(update, context):
            return

        resolved_workspace = self.db.resolve_workspace(user.id, chat.id, chat.type)
        bound_workspace = self.db.get_workspace_by_chat_id(chat.id)

        if chat.type == "private":
            if resolved_workspace and str(resolved_workspace.get("type") or "") == "personal":
                base_url = self._miniapp_base_url()
                if not base_url:
                    await message.reply_text(t("miniapp_not_configured", lang))
                    return
                await message.reply_text(
                    "Вы работаете в личном бюджете.",
                    reply_markup=build_miniapp_button(resolved_workspace, base_url),
                )
                return
            if bound_workspace and str(bound_workspace.get("type") or "") == "family":
                await message.reply_text("Откройте Mini App из нужной группы")
                return
            await message.reply_text(t("private_mode_required", lang))
            return

        workspace = resolved_workspace or bound_workspace
        if not workspace:
            await message.reply_text(t("group_start_required", lang))
            return

        if str(workspace.get("type") or "") != "family":
            await message.reply_text("Выберите группу для семейного бюджета.")
            return

        workspace_id = int(workspace.get("workspace_id") or workspace.get("id") or 0)
        if workspace_id > 0 and not bool(workspace.get("is_member")):
            try:
                self.db.add_workspace_member(
                    workspace_id=workspace_id,
                    telegram_user_id=user.id,
                    role="member",
                )
                self.db.register_member(
                    chat_id=chat.id,
                    telegram_user_id=user.id,
                    username=user.username,
                    full_name=_display_name(update),
                )
            except Exception:
                logger.exception(
                    "Failed to auto-join user to workspace for /app: chat_id=%s workspace_id=%s user_id=%s",
                    chat.id,
                    workspace_id,
                    user.id,
                )

        workspace_title = str(workspace.get("title") or "").strip() or "Семейный бюджет"
        await self._send_miniapp_workspace_message(
            context=context,
            chat_id=chat.id,
            workspace=workspace,
            text=f"Вы работаете в семейном бюджете: {workspace_title}",
            lang=lang,
        )

    async def new_members_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        if not message or not chat or not message.new_chat_members:
            return

        inviter = message.from_user
        if inviter and not inviter.is_bot:
            full_name = f"{inviter.first_name or ''} {inviter.last_name or ''}".strip() or (
                inviter.username or str(inviter.id)
            )
            self.db.register_member(
                chat_id=chat.id,
                telegram_user_id=inviter.id,
                username=inviter.username,
                full_name=full_name,
            )
        inviter_lang = self._user_language(inviter.id) if inviter and not inviter.is_bot else DEFAULT_LANGUAGE

        for member in message.new_chat_members:
            if member.is_bot:
                continue
            full_name = f"{member.first_name or ''} {member.last_name or ''}".strip() or (
                member.username or str(member.id)
            )
            self.db.register_member(
                chat_id=chat.id,
                telegram_user_id=member.id,
                username=member.username,
                full_name=full_name,
            )
            member_lang = self._user_language(member.id)
            await message.reply_text(
                {
                    "ru": f"{_user_handle(member)}, привет! Как к тебе обращаться?",
                    "uz": f"{_user_handle(member)}, salom! Sizga qanday murojaat qilay?",
                    "en": f"{_user_handle(member)}, hi! What should I call you?",
                }.get(member_lang, f"{_user_handle(member)}, hi! What should I call you?")
            )

        bot_added = any(member.id == context.bot.id for member in message.new_chat_members)
        if bot_added:
            is_admin = await self._is_bot_admin_in_group(chat.id, context)
            if is_admin:
                return
            onboarding = self._get_group_onboarding_state(context)
            existing_admin_prompt = onboarding.get("admin_prompt_message_id")
            if isinstance(existing_admin_prompt, int) and existing_admin_prompt > 0:
                return
            sent = await message.reply_text(
                t("start_group_admin_request", inviter_lang),
                reply_markup=self._group_admin_ready_keyboard(inviter_lang),
            )
            onboarding["admin_prompt_message_id"] = sent.message_id
            self._track_onboarding_message_id(onboarding, sent.message_id)
            onboarding["admin_confirmed"] = False

    async def onboarding_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat:
            return
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE

        data = query.data or ""
        if not data.startswith("onb:"):
            await query.answer()
            return

        if data == "onb:p_step2":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            await self._send_personal_onboarding_step2(chat.id, context, lang)
            return

        if data == "onb:p_skip_voice":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            await self._send_personal_onboarding_step4(chat.id, context, lang)
            return

        if data == "onb:p_show_report":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            await self._send_personal_onboarding_report(chat.id, context, lang)
            return

        if data == "onb:p_step5":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            await self._send_personal_onboarding_step5(chat.id, context, lang)
            return

        if data.startswith("onb:p_time:"):
            await query.answer()
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    try:
                        await query.message.edit_reply_markup(reply_markup=None)
                    except Exception:
                        pass
            hour = int(data.split(":")[-1])
            user_id = user.id if user else 0
            await self._finish_personal_onboarding(
                chat.id, user_id, context, lang, send_hour=hour,
            )
            return

        if data == "onb:how_private":
            if chat.type != "private":
                await query.answer()
                return
            add_group_url = await self._build_add_group_url(context)
            await query.answer()
            await context.bot.send_message(
                chat_id=chat.id,
                text={
                    "ru": (
                        "💡 Как работает семейный финансовый бот\n"
                        "1️⃣ Вы создаёте отдельную группу для бюджета (лучше только вдвоём).\n"
                        "2️⃣ Добавляете меня в неё.\n"
                        "3️⃣ Каждый представляется — я запоминаю, кто есть кто.\n"
                        "4️⃣ Дальше вы просто пишете расходы и доходы обычным текстом или голосом.\n\n"
                        "Примеры:\n"
                        "• кофе 20 000\n"
                        "• продукты 150 000\n"
                        "• зарплата 5 000 000\n"
                        "• перевёл жене 300 000\n\n"
                        "Я:\n"
                        "• считаю общий баланс\n"
                        "• разделяю личные и семейные расходы\n"
                        "• учитываю переводы внутри семьи\n"
                        "• делаю ежедневные и месячные отчёты\n\n"
                        "Ничего сложного 🙂\n"
                        "Готовы начать?"
                    ),
                    "uz": (
                        "💡 Oilaviy moliya boti qanday ishlaydi\n"
                        "1️⃣ Byudjet uchun alohida guruh yaratasiz (yaxshisi faqat ikkovingiz).\n"
                        "2️⃣ Meni guruhga qo'shasiz.\n"
                        "3️⃣ Har kim o'zini tanishtiradi — men kimligini eslab qolaman.\n"
                        "4️⃣ Keyin xarajat va daromadlarni oddiy matn yoki ovozli xabar bilan yozasiz.\n\n"
                        "Misollar:\n"
                        "• qahva 20 000\n"
                        "• mahsulotlar 150 000\n"
                        "• maosh 5 000 000\n"
                        "• turmush o'rtog'imga 300 000 o'tkazdim\n\n"
                        "Men:\n"
                        "• umumiy balansni hisoblayman\n"
                        "• shaxsiy va oilaviy xarajatlarni ajrataman\n"
                        "• oila ichidagi o'tkazmalarni hisobga olaman\n"
                        "• kunlik va oylik hisobotlar tayyorlayman\n\n"
                        "Hammasi juda oddiy 🙂\n"
                        "Boshlashga tayyormisiz?"
                    ),
                    "en": (
                        "💡 How the family finance bot works\n"
                        "1️⃣ You create a separate group for the budget (ideally just the two of you).\n"
                        "2️⃣ You add me to it.\n"
                        "3️⃣ Everyone introduces themselves — I remember who is who.\n"
                        "4️⃣ Then you simply send expenses and income as text or voice.\n\n"
                        "Examples:\n"
                        "• coffee 20,000\n"
                        "• groceries 150,000\n"
                        "• salary 5,000,000\n"
                        "• transferred 300,000 to my spouse\n\n"
                        "I will:\n"
                        "• calculate the total balance\n"
                        "• separate personal and family expenses\n"
                        "• track transfers inside the family\n"
                        "• create daily and monthly reports\n\n"
                        "Nothing complicated 🙂\n"
                        "Ready to start?"
                    ),
                }.get(lang),
                reply_markup=self._private_how_keyboard(add_group_url, lang=lang),
            )
            return

        if data == "onb:admin_ready":
            if chat.type not in {"group", "supergroup"}:
                await query.answer()
                return

            is_admin = await self._is_bot_admin_in_group(chat.id, context)
            if not is_admin:
                await query.answer(
                    {
                        "ru": "Пока я не админ. Дайте права администратора и нажмите «Готово» ещё раз.",
                        "uz": "Hozircha admin emasman. Administrator huquqini bering va «Tayyor»ni yana bosing.",
                        "en": "I am not an admin yet. Grant admin rights and press “Done” again.",
                    }.get(lang, "I am not an admin yet."),
                    show_alert=True,
                )
                return

            onboarding = self._get_group_onboarding_state(context)
            onboarding["admin_confirmed"] = True
            onboarding["admin_prompt_message_id"] = None
            source_message_id = query.message.message_id if query.message else 0
            source_deleted = False
            admin_message_ids = onboarding.get("admin_message_ids")
            if isinstance(admin_message_ids, list):
                for message_id in list(admin_message_ids):
                    if not isinstance(message_id, int) or message_id <= 0:
                        continue
                    try:
                        await context.bot.delete_message(chat_id=chat.id, message_id=message_id)
                        if source_message_id and source_message_id == message_id:
                            source_deleted = True
                    except Exception:
                        logger.exception("Failed to delete onboarding admin message %s", message_id)
                onboarding["admin_message_ids"] = []
            if query.message and not source_deleted:
                try:
                    await query.message.delete()
                except Exception:
                    logger.exception("Failed to delete admin-ready source message")

            await self._auto_set_member_names(chat.id, context)
            await self._send_group_family_onboarding(
                chat_id=chat.id,
                context=context,
                lang=lang,
            )
            onboarding["intro_sent"] = True
            onboarding["onboarding_step"] = "try_it"
            await query.answer(
                {
                    "ru": "Отлично, продолжаем",
                    "uz": "Ajoyib, davom etamiz",
                    "en": "Great, let's continue",
                }.get(lang, "Great, let's continue")
            )
            return

        if data == "onb:try_it":
            await query.answer()
            onboarding = self._get_group_onboarding_state(context)
            onboarding["onboarding_step"] = "waiting_expenses"
            onboarding["onboarding_expense_count"] = 0
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            prompt = {
                "ru": "Отлично! Попробуйте записать первый расход.\n\nНапишите в чат, например:\n«кофе 15 000»",
                "uz": "Ajoyib! Birinchi xarajatni yozib ko'ring.\n\nChatga yozing, masalan:\n«qahva 15 000»",
                "en": "Great! Try recording your first expense.\n\nType in chat, for example:\n\"coffee 15,000\"",
            }.get(lang, "Great! Try recording your first expense.\nType: \"coffee 15,000\"")
            await context.bot.send_message(chat_id=chat.id, text=prompt)
            return

        if data == "onb:show_report":
            await query.answer()
            onboarding = self._get_group_onboarding_state(context)
            onboarding["onboarding_step"] = "report_shown"
            if query.message:
                try:
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception:
                    pass
            try:
                state = self._default_report_state()
                payload = self._build_report_payload(chat.id, state, lang=lang)
                report_text = self._build_report_text(
                    period_title=str(payload["period_title"]),
                    scope_title=str(payload["scope_title"]),
                    breakdown=dict(payload["breakdown"]),
                    detailed=False,
                    transactions=list(payload["transactions"]),
                    balance_title=str(payload["balance_title"]),
                    show_transfer_summary=False,
                    lang=lang,
                )
            except Exception:
                logger.exception("Failed to build onboarding report")
                report_text = None
            report_intro = {
                "ru": "📊 Вот ваш первый отчёт! Так будет выглядеть статистика каждый день:",
                "uz": "📊 Mana birinchi hisobotingiz! Har kuni statistika shunday ko'rinadi:",
                "en": "📊 Here's your first report! This is what daily stats look like:",
            }.get(lang, "📊 Here's your first report!")
            await context.bot.send_message(chat_id=chat.id, text=report_intro)
            if report_text:
                await context.bot.send_message(
                    chat_id=chat.id,
                    text=report_text,
                    parse_mode=ParseMode.HTML,
                )
            app_prompt = {
                "ru": "А для полной аналитики — есть мини-приложение с графиками, категориями и прогнозами 👇",
                "uz": "To'liq tahlil uchun — grafiklar, kategoriyalar va prognozli mini-ilova bor 👇",
                "en": "For full analytics — there's a mini-app with charts, categories, and forecasts 👇",
            }.get(lang, "For full analytics — there's a mini-app 👇")
            await context.bot.send_message(chat_id=chat.id, text=app_prompt)
            await self._send_group_family_mode_activated(
                chat_id=chat.id,
                context=context,
                lang=lang,
            )
            await context.bot.send_message(
                chat_id=chat.id,
                text={
                    "ru": "Готово! Теперь просто пишите расходы и доходы в чат 🎉",
                    "uz": "Tayyor! Endi chatga xarajat va daromadlarni yozsangiz bo'ldi 🎉",
                    "en": "Done! Now just write expenses and income in the chat 🎉",
                }.get(lang, "Done! 🎉"),
                reply_markup=self._group_main_reply_keyboard(lang=lang),
            )
            onboarding["final_sent"] = True
            return

        if data == "onb:demo":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_text(
                        self._demo_report_text(),
                        reply_markup=self._group_back_keyboard(lang),
                    )
                    return
                except Exception:
                    logger.exception("Failed to render demo report in-place")
            await context.bot.send_message(
                chat_id=chat.id,
                text=self._demo_report_text(),
                reply_markup=self._group_back_keyboard(lang),
            )
            return

        if data == "onb:settings":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_text(
                        self._assistant_settings_text(lang),
                        reply_markup=self._assistant_settings_keyboard(lang=lang, back_callback="onb:back"),
                    )
                    return
                except Exception:
                    logger.exception("Failed to render onboarding settings in-place")
            await context.bot.send_message(
                chat_id=chat.id,
                text=self._assistant_settings_text(lang),
                reply_markup=self._assistant_settings_keyboard(lang=lang, back_callback="onb:back"),
            )
            return

        if data == "onb:back":
            await query.answer()
            if query.message:
                try:
                    await query.message.edit_text(
                        self._group_onboarding_final_text(lang=lang),
                        reply_markup=self._group_after_onboarding_keyboard(lang=lang),
                    )
                    return
                except Exception:
                    logger.exception("Failed to restore onboarding menu in-place")
            await context.bot.send_message(
                chat_id=chat.id,
                text=self._group_onboarding_final_text(lang=lang),
                reply_markup=self._group_after_onboarding_keyboard(lang=lang),
            )
            return

        await query.answer()

    async def main_menu_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat:
            return
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE

        data = query.data or ""
        if not data.startswith("mn:"):
            await query.answer()
            return

        if data == "mn:settings":
            await query.answer()
            await context.bot.send_message(
                chat_id=chat.id,
                text=self._assistant_settings_text(lang),
                reply_markup=self._assistant_settings_keyboard(lang=lang),
            )
            return

        if data == "mn:report":
            if not await self._ensure_private_mode_selected(update, context):
                return
            user_lang = lang
            state = self._get_report_state(context, chat.id)
            state.clear()
            state.update(self._default_report_state())
            payload = self._build_report_payload(chat.id, state, lang=user_lang)
            report_text = self._build_report_text(
                period_title=str(payload["period_title"]),
                scope_title=str(payload["scope_title"]),
                breakdown=dict(payload["breakdown"]),
                detailed=bool(state.get("detailed", False)),
                transactions=list(payload["transactions"]),
                balance_title=str(payload["balance_title"]),
                show_transfer_summary=str(payload["scope"]) == "user",
                lang=user_lang,
            )
            sent = await context.bot.send_message(
                chat_id=chat.id,
                text=report_text,
                parse_mode=ParseMode.HTML,
                reply_markup=self._report_menu_keyboard(
                    chat.id,
                    selected_scope=str(payload["scope"]),
                    selected_user_id=(
                        int(payload["selected_user_id"])
                        if isinstance(payload.get("selected_user_id"), int)
                        else None
                    ),
                    detailed=bool(state.get("detailed", False)),
                    period_mode=str(payload["period_mode"]),
                    lang=user_lang,
                ),
            )
            state["report_message_id"] = sent.message_id
            await query.answer(self._report_text("opening_report", user_lang))
            return

        if data == "mn:app":
            user = update.effective_user
            if not user:
                await query.answer()
                return
            await self.app_command(update, context)
            await self._safe_query_answer(
                query,
                {
                    "ru": "Открываю приложение",
                    "uz": "Ilovani ochaman",
                    "en": "Opening app",
                }.get(lang, "Opening app"),
            )
            return

        await query.answer()

    async def guess_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            return

        data = query.data or ""
        if not data.startswith("gc:"):
            await query.answer()
            return

        pending = context.user_data.get(PENDING_GUESS_KEY)
        if not isinstance(pending, dict):
            await query.answer("Подтверждение уже неактуально.", show_alert=True)
            return
        if int(pending.get("chat_id") or 0) != chat.id or int(pending.get("user_id") or 0) != user.id:
            await query.answer("Это подтверждение не для вас.", show_alert=True)
            return

        event_data = pending.get("event")
        if not isinstance(event_data, dict):
            context.user_data.pop(PENDING_GUESS_KEY, None)
            await query.answer("Подтверждение устарело.", show_alert=True)
            return

        event = ParsedEvent(
            kind=str(event_data.get("kind") or "expense"),
            amount=float(event_data.get("amount") or 0.0),
            currency=str(event_data.get("currency") or self.settings.default_currency),
            category=str(event_data.get("category") or "expense_other"),
            description=str(event_data.get("description") or ""),
            is_family=bool(event_data.get("is_family")),
            confidence=0.6,
        )
        source_text = str(pending.get("source_text") or event.description)
        source_message_id = int(pending.get("source_message_id") or 0)

        async def save_and_render(selected_event: ParsedEvent) -> None:
            member_name = self.db.get_member_display_name(chat.id, user.id) or (
                user.full_name or user.username or str(user.id)
            )
            transaction_id = self._store_event(
                chat_id=chat.id,
                user_id=user.id,
                member_name=member_name,
                event=selected_event,
                source_type="text",
                original_text=source_text,
                transcript=None,
                message_id=source_message_id if source_message_id > 0 else (query.message.message_id if query.message else 0),
            )
            context.user_data.pop(PENDING_GUESS_KEY, None)
            card_text = self._format_transaction_card(
                kind=selected_event.kind,
                amount=selected_event.amount,
                currency=selected_event.currency,
                category=selected_event.category,
                description=selected_event.description,
                is_updated=False,
            )
            if query.message:
                await query.message.edit_text(
                    card_text,
                    reply_markup=self._transaction_actions_keyboard(transaction_id, user.id),
                )
            else:
                await context.bot.send_message(
                    chat_id=chat.id,
                    text=card_text,
                    reply_markup=self._transaction_actions_keyboard(transaction_id, user.id),
                )

        if data == "gc:cancel":
            context.user_data.pop(PENDING_GUESS_KEY, None)
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    logger.exception("Failed to delete guess confirmation message")
            await query.answer("Отменено")
            return

        if data == "gc:change":
            emoji, group = self._category_group_view(event.category, event.kind)
            if query.message:
                await query.message.edit_text(
                    "Выберите категорию:\n"
                    f"Текущая: {emoji} {group}",
                    reply_markup=self._guess_category_picker_keyboard(),
                )
            await query.answer()
            return

        if data == "gc:back":
            emoji, group = self._category_group_view(event.category, event.kind)
            kind_text = "доход" if event.kind == "income" else "расход"
            if query.message:
                await query.message.edit_text(
                    f"Похоже, это {kind_text} на {_fmt_money(event.amount, event.currency)}.\n"
                    "Я правильно понимаю?\n"
                    f"Категория: {emoji} {group}",
                    reply_markup=self._guess_confirm_keyboard(),
                )
            await query.answer()
            return

        if data.startswith("gc:cat:"):
            category = data.split(":", 2)[2]
            is_family = category.startswith("groceries_") or category.startswith("home_")
            selected_event = ParsedEvent(
                kind=event.kind,
                amount=event.amount,
                currency=event.currency,
                category=category,
                description=event.description,
                is_family=is_family,
                confidence=0.7,
            )
            await save_and_render(selected_event)
            await query.answer("Записал")
            return

        if data == "gc:yes":
            await save_and_render(event)
            await query.answer("Записал")
            return

        await query.answer()

    def _name_greeting(self, name: str, *, changed: bool) -> str:
        if changed:
            return f"Хорошо, {name}. Буду обращаться так."
        return random.choice(NAME_GREETING_VARIANTS).format(name=name)

    async def _after_name_saved(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        name: str,
        was_changed: bool,
    ) -> None:
        message = update.effective_message
        chat = update.effective_chat
        if not message or not chat:
            return

        await message.reply_text(self._name_greeting(name, changed=was_changed))
        if chat.type not in {"group", "supergroup"}:
            return

        onboarding = self._get_group_onboarding_state(context)
        if not bool(onboarding.get("admin_confirmed")):
            return
        if not bool(onboarding.get("final_sent")):
            user_lang = self._user_language(message.from_user.id) if message.from_user else DEFAULT_LANGUAGE
            await self._send_group_onboarding_final(chat_id=chat.id, context=context, lang=user_lang)
            onboarding["final_sent"] = True
            return

        if not was_changed:
            await message.reply_text(
                (
                    {
                        "ru": (
                            f"{name}, рад знакомству! "
                            "Посмотри, пожалуйста, сообщения выше: "
                            "там финальные подсказки, пример отчёта и настройки ассистента."
                        ),
                        "uz": (
                            f"{name}, tanishganimdan xursandman! "
                            "Iltimos, yuqoridagi xabarlarga qarang: "
                            "u yerda yakuniy ko'rsatmalar, hisobot namunasi va assistent sozlamalari bor."
                        ),
                        "en": (
                            f"{name}, nice to meet you! "
                            "Please check the messages above: "
                            "they contain final tips, a sample report, and assistant settings."
                        ),
                    }.get(self._user_language(message.from_user.id) if message.from_user else DEFAULT_LANGUAGE)
                ),
                reply_markup=self._group_after_onboarding_keyboard(
                    lang=self._user_language(message.from_user.id) if message.from_user else DEFAULT_LANGUAGE
                ),
            )

    def _normalize_name(self, raw_name: str) -> str | None:
        name = " ".join(raw_name.split())
        if not name:
            return None
        if not NAME_PATTERN.match(name):
            return None

        lower = name.lower()
        if lower in NON_NAME_WORDS:
            return None
        if any(ch.isdigit() for ch in name):
            return None
        if any(marker in lower for marker in FINANCE_HINTS):
            return None

        parts = [part for part in name.split(" ") if part]
        if len(parts) == 0 or len(parts) > 3:
            return None
        banned_name_tokens = {
            "меня",
            "зовут",
            "name",
            "is",
            "поменяй",
            "измени",
            "зови",
            "на",
            "my",
            "change",
        }
        for part in parts:
            if len(part) < 2:
                return None
            if part.lower() in NON_NAME_WORDS:
                return None
            if part.lower() in banned_name_tokens:
                return None

        normalized_parts = [p[:1].upper() + p[1:].lower() if len(p) > 1 else p.upper() for p in parts]
        return " ".join(normalized_parts)

    def _extract_explicit_name_candidate(self, text: str) -> str | None:
        for pattern in NAME_EXPLICIT_PATTERNS:
            match = pattern.match(text)
            if not match:
                continue
            candidate = self._normalize_name(match.group(1).strip())
            if candidate:
                return candidate

        return None

    def _extract_plain_name_candidate(self, text: str) -> str | None:
        cleaned = text.strip().strip(".,!?")
        if "?" in text:
            return None
        return self._normalize_name(cleaned)

    async def name_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return
        await self._remember_member(update)

        raw_name = " ".join(context.args).strip()
        parsed_name = self._normalize_name(raw_name)
        if not parsed_name:
            await message.reply_text("Напиши, пожалуйста, так: /name Мурад")
            return

        previous_name = self.db.get_custom_name(chat.id, user.id)
        self.db.set_custom_name(chat.id, user.id, parsed_name)
        self.name_prompted.discard((chat.id, user.id))
        await self._after_name_saved(
            update=update,
            context=context,
            name=parsed_name,
            was_changed=bool(previous_name and previous_name != parsed_name),
        )

    async def register_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        if not message:
            return

        if context.args and context.args[0].lower() in {"husband", "wife", "муж", "жена"}:
            await message.reply_text(
                "Роли больше не нужны. Просто напиши имя, как тебе удобно."
            )
            return

        await self.name_command(update, context)

    async def whoami_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        chat = update.effective_chat
        user = update.effective_user
        message = update.effective_message
        if not chat or not user or not message:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return
        await self._remember_member(update)

        custom_name = self.db.get_custom_name(chat.id, user.id)
        if not custom_name:
            await message.reply_text("Я пока не знаю твое имя. Напиши его одним сообщением.")
            return

        await message.reply_text(f"У меня ты записан как {custom_name}.")

    async def members_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        chat = update.effective_chat
        message = update.effective_message
        if not chat or not message:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return

        await self._warn_if_privacy_mode_blocks_messages(
            chat.id, chat.type, message, context
        )
        members = self.db.list_members(chat.id)
        if not members:
            await message.reply_text("Пока никого не вижу в списке.")
            return

        lines = ["Сейчас у меня так:"]
        for member in members:
            shown_name = str(member.get("custom_name") or "еще не представился")
            lines.append(f"- {_member_handle(member)}: {shown_name}")
        missing = [m for m in members if not m.get("custom_name")]
        if missing:
            lines.append("")
            for member in missing:
                lines.append(f"{_member_handle(member)}, как к вам обращаться?")
        await message.reply_text("\n".join(lines))

    @staticmethod
    def _transaction_actions_keyboard(
        transaction_id: int, owner_user_id: int
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "🗑 Удалить",
                        callback_data=f"tx:delete:{transaction_id}:{owner_user_id}",
                    ),
                    InlineKeyboardButton(
                        "✏️ Редактировать",
                        callback_data=f"tx:edit:{transaction_id}:{owner_user_id}",
                    ),
                ]
            ]
        )

    @staticmethod
    def _transfer_actions_keyboard(
        transaction_id: int, owner_user_id: int
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "🗑 Удалить",
                        callback_data=f"tx:delete:{transaction_id}:{owner_user_id}",
                    ),
                    InlineKeyboardButton(
                        "✏️ Редактировать",
                        callback_data=f"tx:edit:{transaction_id}:{owner_user_id}",
                    ),
                ]
            ]
        )

    @staticmethod
    def _transaction_edit_keyboard(
        transaction_id: int, owner_user_id: int
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "💰 Сумма",
                        callback_data=f"tx:set_amount:{transaction_id}:{owner_user_id}",
                    ),
                    InlineKeyboardButton(
                        "📝 Описание",
                        callback_data=f"tx:set_desc:{transaction_id}:{owner_user_id}",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        "🗑 Удалить",
                        callback_data=f"tx:delete:{transaction_id}:{owner_user_id}",
                    )
                ],
            ]
        )

    @staticmethod
    def _transfer_edit_keyboard(
        transaction_id: int, owner_user_id: int
    ) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "💰 Сумма",
                        callback_data=f"tx:set_transfer_amount:{transaction_id}:{owner_user_id}",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        "🗑 Удалить",
                        callback_data=f"tx:delete:{transaction_id}:{owner_user_id}",
                    )
                ],
            ]
        )

    @staticmethod
    def _format_transaction_card(
        *,
        kind: str,
        amount: float,
        currency: str,
        category: str,
        description: str,
        is_updated: bool = False,
    ) -> str:
        title = "Запись обновил ✅" if is_updated else "Ваш доход успешно добавлен ✅"
        if kind == "expense" and not is_updated:
            title = "Ваш расход успешно добавлен ✅"
        if kind == "expense" and is_updated:
            title = "Запись обновил ✅"

        return (
            f"{title}\n\n"
            f"💰 Сумма: {_fmt_money(amount, currency)}\n"
            f"🏷 Категория: {_ru_category_label(category, kind)}\n"
            f"🧾 Описание: {description}"
        )

    @staticmethod
    def _is_internal_transfer_category(category: str) -> bool:
        return str(category or "") in {TRANSFER_OUT_CATEGORY, TRANSFER_IN_CATEGORY}

    @staticmethod
    def _format_transfer_saved_text(
        *,
        amount: float,
        currency: str,
        sender_name: str,
        recipient_name: str,
        is_updated: bool = False,
    ) -> str:
        prefix = "✅ Перевод обновил:" if is_updated else "✅ Перевод записан:"
        return (
            f"{prefix} {_fmt_money(amount, currency)} от {sender_name} → {recipient_name}. "
            "(На семейный баланс не влияет)"
        )

    def _get_internal_transfer_pair(
        self, chat_id: int, transaction_id: int
    ) -> tuple[dict[str, object] | None, dict[str, object] | None]:
        tx = self.db.get_transaction(chat_id, transaction_id)
        if not tx or not self._is_internal_transfer_category(str(tx.get("category") or "")):
            return None, None
        source_message_id = int(tx.get("message_id") or 0)
        rows = self.db.get_transfer_transactions_by_message_id(chat_id, source_message_id)
        out_tx = next(
            (row for row in rows if str(row.get("category") or "") == TRANSFER_OUT_CATEGORY),
            None,
        )
        in_tx = next(
            (row for row in rows if str(row.get("category") or "") == TRANSFER_IN_CATEGORY),
            None,
        )
        return out_tx, in_tx

    @staticmethod
    def _is_transaction_owner(
        transaction: dict[str, object] | None, telegram_user_id: int
    ) -> bool:
        if not transaction:
            return False
        owner = transaction.get("telegram_user_id")
        return isinstance(owner, int) and owner == telegram_user_id

    @staticmethod
    def _set_pending_edit(
        context: ContextTypes.DEFAULT_TYPE,
        *,
        chat_id: int,
        transaction_id: int,
        field: str,
        bot_message_id: int,
        prompt_message_id: int | None,
    ) -> None:
        context.user_data[EDIT_STATE_KEY] = {
            "chat_id": chat_id,
            "transaction_id": transaction_id,
            "field": field,
            "bot_message_id": bot_message_id,
            "prompt_message_id": prompt_message_id,
        }

    @staticmethod
    def _clear_pending_edit(context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop(EDIT_STATE_KEY, None)

    async def transaction_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat or not user:
            return
        callback_message = query.message

        data = query.data or ""
        parts = data.split(":")
        if len(parts) not in {3, 4} or parts[0] != "tx":
            await query.answer()
            return

        action = parts[1]
        try:
            transaction_id = int(parts[2])
        except ValueError:
            await query.answer("Некорректная операция.", show_alert=True)
            return
        owner_from_callback: int | None = None
        if len(parts) == 4:
            try:
                owner_from_callback = int(parts[3])
            except ValueError:
                owner_from_callback = None

        transaction = self.db.get_transaction(chat.id, transaction_id)
        if not transaction:
            await query.answer("Эта запись уже удалена.", show_alert=True)
            if callback_message:
                try:
                    await callback_message.delete()
                except Exception:
                    logger.exception("Failed to delete stale transaction message")
            return

        if owner_from_callback is not None and owner_from_callback != user.id:
            await query.answer(
                "Эту запись может менять только тот, кто ее добавил.",
                show_alert=True,
            )
            return

        if not self._is_transaction_owner(transaction, user.id):
            await query.answer(
                "Эту запись может редактировать только тот, кто ее добавил.",
                show_alert=True,
            )
            return

        is_internal_transfer = self._is_internal_transfer_category(str(transaction["category"]))
        transfer_out_tx = None
        transfer_in_tx = None
        if is_internal_transfer:
            transfer_out_tx, transfer_in_tx = self._get_internal_transfer_pair(chat.id, transaction_id)
            if not transfer_out_tx or not transfer_in_tx:
                await query.answer(
                    "Не нашел связанную пару перевода. Попробуйте удалить и записать заново.",
                    show_alert=True,
                )
                return
            if int(transfer_out_tx["telegram_user_id"]) != user.id:
                await query.answer(
                    "Перевод может редактировать только тот, кто его отправил.",
                    show_alert=True,
                )
                return

        if action == "edit":
            await query.answer("Что поменяем?")
            if callback_message:
                try:
                    reply_markup = (
                        self._transfer_edit_keyboard(
                            int(transfer_out_tx["id"]), int(transfer_out_tx["telegram_user_id"])
                        )
                        if is_internal_transfer and transfer_out_tx
                        else self._transaction_edit_keyboard(
                            transaction_id, int(transaction["telegram_user_id"])
                        )
                    )
                    await callback_message.edit_reply_markup(reply_markup=reply_markup)
                except Exception:
                    logger.exception("Failed to show edit keyboard for transaction")
            return

        if action in {"set_amount", "set_desc", "set_transfer_amount"}:
            if is_internal_transfer and action == "set_desc":
                await query.answer("Для перевода пока можно менять только сумму.", show_alert=True)
                return
            field = "amount" if action == "set_amount" else "description"
            if action == "set_transfer_amount":
                field = "transfer_amount"
            if not callback_message:
                await query.answer("Не вижу сообщение для редактирования.", show_alert=True)
                return
            prompt = None
            if field in {"amount", "transfer_amount"}:
                prompt = await callback_message.reply_text(
                    "Напиши новую сумму одним сообщением. Например: 5400000"
                )
            else:
                prompt = await callback_message.reply_text("Напиши новое описание.")
            self._set_pending_edit(
                context,
                chat_id=chat.id,
                transaction_id=transaction_id,
                field=field,
                bot_message_id=callback_message.message_id,
                prompt_message_id=prompt.message_id if prompt else None,
            )
            await query.answer()
            return

        if action == "delete":
            source_message_id = int(transaction["message_id"])
            deleted_ids = {transaction_id}
            if is_internal_transfer and transfer_out_tx and transfer_in_tx:
                self.db.delete_transaction(chat.id, int(transfer_out_tx["id"]))
                self.db.delete_transaction(chat.id, int(transfer_in_tx["id"]))
                deleted_ids = {int(transfer_out_tx["id"]), int(transfer_in_tx["id"])}
            else:
                self.db.delete_transaction(chat.id, transaction_id)
            state = context.user_data.get(EDIT_STATE_KEY)
            if isinstance(state, dict):
                state_tx = state.get("transaction_id")
                if isinstance(state_tx, int) and state_tx in deleted_ids:
                    prompt_message_id = state.get("prompt_message_id")
                    if isinstance(prompt_message_id, int):
                        try:
                            await context.bot.delete_message(
                                chat_id=chat.id, message_id=prompt_message_id
                            )
                        except Exception:
                            logger.exception("Failed to delete pending edit prompt message")
                    self._clear_pending_edit(context)

            try:
                await context.bot.delete_message(
                    chat_id=chat.id, message_id=source_message_id
                )
            except Exception:
                logger.exception("Failed to delete original user message")

            if callback_message:
                try:
                    await callback_message.delete()
                except Exception:
                    logger.exception("Failed to delete transaction card message")

            await query.answer("Удалил.")
            return

        await query.answer()

    async def _try_apply_pending_edit(
        self,
        *,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: str,
    ) -> bool:
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return False

        state = context.user_data.get(EDIT_STATE_KEY)
        if not isinstance(state, dict):
            return False

        state_chat_id = state.get("chat_id")
        if not isinstance(state_chat_id, int) or state_chat_id != chat.id:
            return False

        transaction_id = state.get("transaction_id")
        field = state.get("field")
        bot_message_id = state.get("bot_message_id")
        prompt_message_id = state.get("prompt_message_id")
        if not isinstance(transaction_id, int) or field not in {"amount", "description", "transfer_amount"}:
            self._clear_pending_edit(context)
            return False

        transaction = self.db.get_transaction(chat.id, transaction_id)
        if not self._is_transaction_owner(transaction, user.id):
            self._clear_pending_edit(context)
            await message.reply_text("Не нашел твою запись для редактирования.")
            return True

        old_amount = float(transaction["amount"])
        old_description = str(transaction["description"])
        new_amount = old_amount
        new_description = old_description

        if field in {"amount", "transfer_amount"}:
            parsed_amount = _extract_amount_value(text)
            if parsed_amount is None:
                await message.reply_text("Не вижу сумму. Напиши только число, например 250000.")
                return True
            new_amount = parsed_amount
        else:
            cleaned = " ".join(text.split()).strip()
            if not cleaned:
                await message.reply_text("Описание пустое. Напиши, пожалуйста, нормальным текстом.")
                return True
            new_description = cleaned[:200]

        if field == "transfer_amount" or self._is_internal_transfer_category(str(transaction["category"])):
            out_tx, in_tx = self._get_internal_transfer_pair(chat.id, transaction_id)
            if not out_tx or not in_tx:
                self._clear_pending_edit(context)
                await message.reply_text("Не нашел связанную пару перевода для редактирования.")
                return True
            if int(out_tx["telegram_user_id"]) != user.id:
                self._clear_pending_edit(context)
                await message.reply_text("Этот перевод может менять только отправитель.")
                return True

            updated_out = self.db.update_transaction(
                chat_id=chat.id,
                transaction_id=int(out_tx["id"]),
                amount=new_amount,
                description=str(out_tx["description"]),
                category=str(out_tx["category"]),
            )
            updated_in = self.db.update_transaction(
                chat_id=chat.id,
                transaction_id=int(in_tx["id"]),
                amount=new_amount,
                description=str(in_tx["description"]),
                category=str(in_tx["category"]),
            )
            self._clear_pending_edit(context)
            if not (updated_out and updated_in):
                await message.reply_text("Не получилось обновить сумму перевода. Попробуй еще раз.")
                return True

            out_tx = self.db.get_transaction(chat.id, int(out_tx["id"]))
            in_tx = self.db.get_transaction(chat.id, int(in_tx["id"]))
            if not out_tx or not in_tx:
                await message.reply_text("Не получилось прочитать обновленный перевод.")
                return True

            card_text = self._format_transfer_saved_text(
                amount=float(out_tx["amount"]),
                currency=str(out_tx["currency"]),
                sender_name=str(out_tx["member_name"]),
                recipient_name=str(in_tx["member_name"]),
                is_updated=True,
            )

            if isinstance(bot_message_id, int):
                try:
                    await context.bot.edit_message_text(
                        chat_id=chat.id,
                        message_id=bot_message_id,
                        text=card_text,
                        reply_markup=self._transfer_actions_keyboard(
                            int(out_tx["id"]), int(out_tx["telegram_user_id"])
                        ),
                    )
                except Exception:
                    logger.exception("Failed to refresh transfer card after edit")
            if isinstance(prompt_message_id, int):
                try:
                    await context.bot.delete_message(
                        chat_id=chat.id, message_id=prompt_message_id
                    )
                except Exception:
                    logger.exception("Failed to delete transfer edit prompt message")
            try:
                await context.bot.delete_message(chat_id=chat.id, message_id=message.message_id)
            except Exception:
                logger.exception("Failed to delete user transfer edit message")
            return True

        try:
            new_category = await self.ai.infer_category(
                kind=str(transaction["kind"]),
                description=new_description,
                amount=new_amount,
                currency=str(transaction["currency"]),
                is_family=bool(transaction["is_family"]),
            )
        except Exception:
            logger.exception("Failed to infer category for edited transaction")
            new_category = str(transaction["category"])

        updated = self.db.update_transaction(
            chat_id=chat.id,
            transaction_id=transaction_id,
            amount=new_amount,
            description=new_description,
            category=new_category,
        )
        self._clear_pending_edit(context)
        if not updated:
            await message.reply_text("Не получилось обновить запись. Попробуй еще раз.")
            return True

        updated_tx = self.db.get_transaction(chat.id, transaction_id)
        if not updated_tx:
            await message.reply_text("Не получилось прочитать обновленную запись.")
            return True

        card_text = self._format_transaction_card(
            kind=str(updated_tx["kind"]),
            amount=float(updated_tx["amount"]),
            currency=str(updated_tx["currency"]),
            category=str(updated_tx["category"]),
            description=str(updated_tx["description"]),
            is_updated=True,
        )

        if isinstance(bot_message_id, int):
            try:
                await context.bot.edit_message_text(
                    chat_id=chat.id,
                    message_id=bot_message_id,
                    text=card_text,
                    reply_markup=self._transaction_actions_keyboard(
                        transaction_id, int(updated_tx["telegram_user_id"])
                    ),
                )
            except Exception:
                logger.exception("Failed to refresh transaction card after edit")
        if isinstance(prompt_message_id, int):
            try:
                await context.bot.delete_message(
                    chat_id=chat.id, message_id=prompt_message_id
                )
            except Exception:
                logger.exception("Failed to delete edit prompt message")

        try:
            await context.bot.delete_message(chat_id=chat.id, message_id=message.message_id)
        except Exception:
            logger.exception("Failed to delete user edit message")
        return True

    async def handle_text(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user or user.is_bot:
            return

        try:
            self.db.touch_user(user.id)
        except Exception:
            logger.exception("Failed to touch user on text message: user_id=%s", user.id)

        if not await self._ensure_private_mode_selected(update, context):
            return
        await self._remember_member(update)
        text = (message.text or "").strip()
        if not text:
            return
        text_l = text.lower()
        lang = self._user_language(user.id)

        if chat.type in {"group", "supergroup"}:
            if text_l in {
                MAIN_MENU_REPORT_TEXT.lower(),
                "📊 отчеты",
                "📊 отчёты",
                "отчеты",
                "отчёты",
                t("main_menu_reports", "uz").lower(),
                t("main_menu_reports", "en").lower(),
                "hisobotlar",
                "reports",
            }:
                await self.report_command(update, context)
                return
            if text_l in {
                MAIN_MENU_SETTINGS_TEXT.lower(),
                "⚙️ настройки ассистента",
                "настройки ассистента",
                t("main_menu_settings", "uz").lower(),
                t("main_menu_settings", "en").lower(),
                "assistant settings",
                "assistent sozlamalari",
            }:
                sent = await message.reply_text(
                    self._assistant_settings_text(lang),
                    reply_markup=self._assistant_settings_keyboard(lang=lang),
                )
                self._remember_assistant_settings_trigger(
                    context=context,
                    settings_message_id=sent.message_id,
                    user_message_id=message.message_id,
                )
                return
            if text_l in {
                MAIN_MENU_APP_TEXT.lower(),
                "📱 приложение",
                "приложение",
                "миниапп",
                "mini app",
                "miniapp",
                t("main_menu_app", "uz").lower(),
                t("main_menu_app", "en").lower(),
                "ilova",
                "app",
            }:
                await self.app_command(update, context)
                return

        if await self._try_apply_pending_edit(update=update, context=context, text=text):
            return

        if await self._try_apply_pending_report_filter(
            update=update,
            context=context,
            text=text,
        ):
            return

        if await self._try_apply_pending_transfer_input(
            update=update,
            context=context,
            text=text,
        ):
            return

        if await self._try_handle_transfer_message(
            update=update,
            context=context,
            text=text,
        ):
            return

        custom_name = self.db.get_custom_name(chat.id, user.id)
        is_reply_to_bot = bool(
            message.reply_to_message
            and message.reply_to_message.from_user
            and message.reply_to_message.from_user.id == context.bot.id
        )

        explicit_name_candidate = self._extract_explicit_name_candidate(text)
        if explicit_name_candidate:
            was_changed = bool(custom_name and custom_name != explicit_name_candidate)
            self.db.set_custom_name(chat.id, user.id, explicit_name_candidate)
            self.name_prompted.discard((chat.id, user.id))
            await self._after_name_saved(
                update=update,
                context=context,
                name=explicit_name_candidate,
                was_changed=was_changed,
                )
            return

        if not custom_name:
            name_candidate = self._extract_plain_name_candidate(text)
            if name_candidate:
                self.db.set_custom_name(chat.id, user.id, name_candidate)
                self.name_prompted.discard((chat.id, user.id))
                await self._after_name_saved(
                    update=update,
                    context=context,
                    name=name_candidate,
                    was_changed=False,
                )
                return

            if is_reply_to_bot or (chat.type in {"group", "supergroup"}):
                await self._prompt_name_if_needed(update, context)

        # Ignore date filters/report dumps when they are sent as plain chat text
        # outside report input mode.
        if self._looks_like_date_or_range_text(text) or self._looks_like_report_dump_text(text):
            return

        processing_message = None
        if _has_amount_and_text(text):
            try:
                processing_message = await message.reply_text(
                    "⏳ Секунду, анализирую запись и подбираю категорию..."
                )
            except Exception:
                logger.exception("Failed to send processing placeholder message")

        events: list[ParsedEvent] = []
        try:
            events = await self.ai.parse_finance_events(text)
        except Exception:
            logger.exception("Failed to parse text message with AI.")

        if not events and _has_amount_and_text(text) and not _looks_like_math_expression(text):
            guess_event = await self._build_guess_event(text=text)
            if guess_event is not None:
                events = [guess_event]

        if processing_message is not None:
            try:
                await processing_message.delete()
            except Exception:
                logger.exception("Failed to delete processing placeholder message")

        if not events:
            return

        member_name = self.db.get_member_display_name(chat.id, user.id) or _display_name(update)
        for event in events:
            transaction_id = self._store_event(
                chat_id=chat.id,
                user_id=user.id,
                member_name=member_name,
                event=event,
                source_type="text",
                original_text=text,
                transcript=None,
                message_id=message.message_id,
            )
            await message.reply_text(
                self._format_transaction_card(
                    kind=event.kind,
                    amount=event.amount,
                    currency=event.currency,
                    category=event.category,
                    description=event.description,
                    is_updated=False,
                ),
                reply_markup=self._transaction_actions_keyboard(transaction_id, user.id),
            )

        if events:
            onb_step = self._get_personal_onboarding_step(context)
            if onb_step == 2:
                await self._send_personal_onboarding_step3(chat.id, context, lang)
            elif onb_step == 3:
                await self._send_personal_onboarding_step4(chat.id, context, lang)

            if chat.type in {"group", "supergroup"}:
                onboarding = self._get_group_onboarding_state(context)
                if onboarding.get("onboarding_step") == "waiting_expenses":
                    count = int(onboarding.get("onboarding_expense_count") or 0) + len(events)
                    onboarding["onboarding_expense_count"] = count
                    if count >= 2:
                        onboarding["onboarding_step"] = "ready_for_report"
                        report_btn = {
                            "ru": "📊 Посмотреть отчёт",
                            "uz": "📊 Hisobotni ko'rish",
                            "en": "📊 View report",
                        }.get(lang, "📊 View report")
                        await context.bot.send_message(
                            chat_id=chat.id,
                            text={
                                "ru": "Супер! Записи сохранены. Давайте посмотрим как выглядит отчёт?",
                                "uz": "Ajoyib! Yozuvlar saqlandi. Hisobot qanday ko'rinishini ko'ramizmi?",
                                "en": "Great! Records saved. Want to see what the report looks like?",
                            }.get(lang, "Great! Records saved."),
                            reply_markup=InlineKeyboardMarkup([[
                                InlineKeyboardButton(report_btn, callback_data="onb:show_report"),
                            ]]),
                        )
                    elif count == 1:
                        await context.bot.send_message(
                            chat_id=chat.id,
                            text={
                                "ru": "Отлично, первая запись! 👍 Попробуйте ещё одну, например:\n«продукты 80 000»",
                                "uz": "Ajoyib, birinchi yozuv! 👍 Yana birini sinab ko'ring:\n«oziq-ovqat 80 000»",
                                "en": "Great, first record! 👍 Try one more, like:\n\"groceries 80,000\"",
                            }.get(lang, "Great, first record! Try one more."),
                        )

    async def handle_voice(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user or user.is_bot or not message.voice:
            return

        try:
            self.db.touch_user(user.id)
        except Exception:
            logger.exception("Failed to touch user on voice message: user_id=%s", user.id)

        if not await self._ensure_private_mode_selected(update, context):
            return
        await self._remember_member(update)
        await self._prompt_name_if_needed(update, context)

        await message.reply_text("Слушаю, сейчас обработаю голосовое сообщение.")

        temp_path = ""
        transcript = ""
        try:
            tg_file = await context.bot.get_file(message.voice.file_id)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".ogg") as tmp_file:
                temp_path = tmp_file.name
            await tg_file.download_to_drive(custom_path=temp_path)

            try:
                transcript = await self.ai.transcribe_voice(temp_path)
            except Exception:
                logger.exception("Failed to transcribe voice message.")
                await message.reply_text(
                    "Пока не смог распознать голосовое. Давай еще раз, пожалуйста."
                )
                return
            if not transcript:
                await message.reply_text("Не расслышал речь в голосовом сообщении.")
                return

            # Voice messages should also support internal family transfers
            # (e.g. "перевёл жене 1 миллион сум") before generic AI categorization.
            if await self._try_handle_transfer_message(
                update=update,
                context=context,
                text=transcript,
            ):
                return

            events: list[ParsedEvent] = []
            try:
                events = await self.ai.parse_finance_events(transcript)
            except Exception:
                logger.exception("Failed to parse transcribed voice message with AI.")

            if (
                not events
                and _has_amount_and_text(transcript)
                and not _looks_like_math_expression(transcript)
            ):
                guess_event = await self._build_guess_event(text=transcript)
                if guess_event is not None:
                    events = [guess_event]

            if not events:
                short_transcript = transcript[:300]
                if len(transcript) > 300:
                    short_transcript += "..."
                await message.reply_text(
                    f"Я услышал: {short_transcript}\nПока не вижу здесь финансовой операции."
                )
                return

            member_name = self.db.get_member_display_name(chat.id, user.id) or _display_name(update)
            for event in events:
                transaction_id = self._store_event(
                    chat_id=chat.id,
                    user_id=user.id,
                    member_name=member_name,
                    event=event,
                    source_type="voice",
                    original_text="[VOICE MESSAGE]",
                    transcript=transcript,
                    message_id=message.message_id,
                )
                await message.reply_text(
                    self._format_transaction_card(
                        kind=event.kind,
                        amount=event.amount,
                        currency=event.currency,
                        category=event.category,
                        description=event.description,
                        is_updated=False,
                    ),
                    reply_markup=self._transaction_actions_keyboard(
                        transaction_id, user.id
                    ),
                )

            if events and self._get_personal_onboarding_step(context) == 3:
                lang = self._user_language(user.id)
                await self._send_personal_onboarding_step4(chat.id, context, lang)
        finally:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)

    async def handle_photo(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user or user.is_bot:
            return

        is_photo_message = bool(message.photo)
        is_image_document = bool(
            message.document
            and (message.document.mime_type or "").lower().startswith("image/")
        )
        if not is_photo_message and not is_image_document:
            return

        if not await self._ensure_private_mode_selected(update, context):
            return
        await self._remember_member(update)
        await self._prompt_name_if_needed(update, context)

        processing_message = None
        try:
            processing_message = await message.reply_text(
                "⏳ Секунду, анализирую изображение и подбираю категорию..."
            )
        except Exception:
            logger.exception("Failed to send image processing placeholder message")

        try:
            if is_photo_message:
                file_id = message.photo[-1].file_id
                mime_type = "image/jpeg"
            else:
                doc = message.document
                if not doc:
                    return
                file_id = doc.file_id
                mime_type = (doc.mime_type or "image/jpeg").strip().lower()

            tg_file = await context.bot.get_file(file_id)
            image_data = await tg_file.download_as_bytearray()
            image_bytes = bytes(image_data)
            caption = (message.caption or "").strip()

            events: list[ParsedEvent] = []
            try:
                events = await self.ai.parse_finance_events_from_image(
                    image_bytes,
                    mime_type=mime_type,
                    caption_text=caption,
                )
            except Exception:
                logger.exception("Failed to parse image message with AI.")

            if not events and caption:
                try:
                    events = await self.ai.parse_finance_events(caption)
                except Exception:
                    logger.exception("Failed to parse image caption with AI.")

            if (
                not events
                and caption
                and _has_amount_and_text(caption)
                and not _looks_like_math_expression(caption)
            ):
                guess_event = await self._build_guess_event(text=caption)
                if guess_event is not None:
                    events = [guess_event]

            if not events:
                await message.reply_text(
                    "Не вижу на изображении явной записи дохода/расхода. "
                    "Можешь продублировать коротко текстом: «кофе 20 000»."
                )
                return

            member_name = self.db.get_member_display_name(chat.id, user.id) or _display_name(update)
            original_text = caption or "[IMAGE MESSAGE]"
            for event in events:
                transaction_id = self._store_event(
                    chat_id=chat.id,
                    user_id=user.id,
                    member_name=member_name,
                    event=event,
                    source_type="text",
                    original_text=original_text,
                    transcript=None,
                    message_id=message.message_id,
                )
                await message.reply_text(
                    self._format_transaction_card(
                        kind=event.kind,
                        amount=event.amount,
                        currency=event.currency,
                        category=event.category,
                        description=event.description,
                        is_updated=False,
                    ),
                    reply_markup=self._transaction_actions_keyboard(
                        transaction_id, user.id
                    ),
                )
        finally:
            if processing_message is not None:
                try:
                    await processing_message.delete()
                except Exception:
                    logger.exception("Failed to delete image processing placeholder")

    def _store_event(
        self,
        chat_id: int,
        user_id: int,
        member_name: str,
        event: ParsedEvent,
        source_type: str,
        original_text: str,
        transcript: str | None,
        message_id: int,
    ) -> int:
        transaction_id = self.db.add_transaction(
            FinanceTransaction(
                chat_id=chat_id,
                telegram_user_id=user_id,
                member_name=member_name,
                kind=event.kind,
                amount=event.amount,
                currency=event.currency or self.settings.default_currency,
                category=event.category,
                description=event.description,
                is_family=event.is_family,
                source_type=source_type,
                original_text=original_text,
                transcript=transcript,
                message_id=message_id,
            )
        )
        return transaction_id

    async def stats_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        if not message or not chat:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return

        await self._warn_if_privacy_mode_blocks_messages(
            chat.id, chat.type, message, context
        )
        start, end, label = _current_month_period()
        users_summary = self.db.get_group_user_summaries(chat.id, start, end)
        family_summary = self.db.get_family_expense_summary(chat.id, start, end)

        lines = [f"Статистика за {label}", "", "По участникам:"]
        if not users_summary:
            lines.append("- Пока операций нет.")
        else:
            for item in users_summary:
                name = str(item.get("name") or "Участник")
                totals = item.get("totals")
                transfer_totals = item.get("transfer_totals")
                if not isinstance(totals, dict):
                    totals = {}
                lines.append(f"- {name}:")
                currencies = set(totals.keys())
                if isinstance(transfer_totals, dict):
                    currencies |= set(transfer_totals.keys())
                if not currencies:
                    lines.append("  пока операций нет")
                    continue
                for currency in sorted(currencies):
                    stat = totals.get(currency, {}) if isinstance(totals, dict) else {}
                    if not isinstance(stat, dict):
                        stat = {}
                    income = float(stat.get("income", 0.0))
                    expense = float(stat.get("expense", 0.0))
                    transfer_stat = (
                        transfer_totals.get(currency, {}) if isinstance(transfer_totals, dict) else {}
                    )
                    sent = (
                        float(transfer_stat.get("sent", 0.0))
                        if isinstance(transfer_stat, dict)
                        else 0.0
                    )
                    received = (
                        float(transfer_stat.get("received", 0.0))
                        if isinstance(transfer_stat, dict)
                        else 0.0
                    )
                    balance = income - expense + received - sent
                    lines.append(
                        f"  доход {_fmt_money(income, currency)}, "
                        f"расход {_fmt_money(expense, currency)}, "
                        f"баланс {_fmt_money(balance, currency, signed=True)}"
                    )
                    if sent > 0 or received > 0:
                        lines.append(
                            f"  ⬅ переводы отправлено {_fmt_money(sent, currency)}, "
                            f"➡ переводы получено {_fmt_money(received, currency)}"
                        )

        lines.append("")
        lines.extend(self._build_family_section(family_summary))

        await message.reply_text("\n".join(lines))

    async def mystats_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat or not user:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return

        await self._warn_if_privacy_mode_blocks_messages(
            chat.id, chat.type, message, context
        )
        await self._remember_member(update)

        start, end, label = _current_month_period()
        summary = self.db.get_user_summary(chat.id, user.id, start, end)
        member_name = self.db.get_member_display_name(chat.id, user.id) or _display_name(update)

        lines = [f"Личная статистика за {label} ({member_name})", ""]
        lines.extend(self._build_user_section(summary))
        await message.reply_text("\n".join(lines))

    async def familystats_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        if not message or not chat:
            return
        if not await self._ensure_private_mode_selected(update, context):
            return

        await self._warn_if_privacy_mode_blocks_messages(
            chat.id, chat.type, message, context
        )
        start, end, label = _current_month_period()
        summary = self.db.get_family_expense_summary(chat.id, start, end)
        lines = [f"Семейные расходы за {label}", ""]
        lines.extend(self._build_family_section(summary))
        await message.reply_text("\n".join(lines))

    async def report_command(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user
        if not message or not chat:
            return
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE
        if not await self._ensure_private_mode_selected(update, context):
            return

        await self._warn_if_privacy_mode_blocks_messages(
            chat.id, chat.type, message, context
        )
        context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
        state = self._get_report_state(context, chat.id)
        state.clear()
        state.update(self._default_report_state())
        payload = self._build_report_payload(chat.id, state, lang=lang)
        report_text = self._build_report_text(
            period_title=str(payload["period_title"]),
            scope_title=str(payload["scope_title"]),
            breakdown=dict(payload["breakdown"]),
            detailed=bool(state.get("detailed", False)),
            transactions=list(payload["transactions"]),
            balance_title=str(payload["balance_title"]),
            show_transfer_summary=str(payload["scope"]) == "user",
            lang=lang,
        )
        sent = await message.reply_text(
            report_text,
            parse_mode=ParseMode.HTML,
            reply_markup=self._report_menu_keyboard(
                chat.id,
                selected_scope=str(payload["scope"]),
                selected_user_id=(
                    int(payload["selected_user_id"])
                    if isinstance(payload.get("selected_user_id"), int)
                    else None
                ),
                detailed=bool(state.get("detailed", False)),
                period_mode=str(payload["period_mode"]),
                lang=lang,
            ),
        )
        state["report_message_id"] = sent.message_id

    async def report_callback_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        self.track_user_activity(update)
        query = update.callback_query
        chat = update.effective_chat
        user = update.effective_user
        if not query or not chat:
            return
        lang = self._user_language(user.id) if user else DEFAULT_LANGUAGE

        data = query.data or ""
        if not data.startswith("rp:"):
            await query.answer()
            return

        state = self._get_report_state(context, chat.id)
        if data in {"rp:p1d:y", "rp:p1d:n"}:
            confirm = context.user_data.get(REPORT_SINGLE_DATE_CONFIRM_KEY)
            if not isinstance(confirm, dict) or int(confirm.get("chat_id") or 0) != chat.id:
                await query.answer(
                    self._report_text("single_date_confirm_outdated", lang),
                    show_alert=True,
                )
                return
            if data == "rp:p1d:n":
                context.user_data.pop(REPORT_SINGLE_DATE_CONFIRM_KEY, None)
                if query.message:
                    try:
                        await query.message.delete()
                    except Exception:
                        logger.exception("Failed to delete single-date confirm message")
                # Keep report period input mode active, user will enter period again.
                if not isinstance(context.user_data.get(REPORT_FILTER_INPUT_KEY), dict):
                    context.user_data[REPORT_FILTER_INPUT_KEY] = {
                        "chat_id": chat.id,
                        "mode": "date_or_range",
                    }
                await context.bot.send_message(
                    chat_id=chat.id,
                    text=self._report_text("reenter_period_prompt", lang),
                    reply_markup=self._period_input_cancel_keyboard(lang),
                )
                await query.answer(self._report_text("waiting_period_answer", lang))
                return

            confirmed_date = confirm.get("date")
            if not isinstance(confirmed_date, date):
                await query.answer(
                    self._report_text("single_date_confirm_outdated", lang),
                    show_alert=True,
                )
                return

            context.user_data.pop(REPORT_SINGLE_DATE_CONFIRM_KEY, None)
            context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
            state["period_mode"] = "date"
            state["custom_date"] = confirmed_date
            state["range_from"] = None
            state["range_to"] = None

            payload = self._build_report_payload(chat.id, state, lang=lang)
            report_text = self._build_report_text(
                period_title=str(payload["period_title"]),
                scope_title=str(payload["scope_title"]),
                breakdown=dict(payload["breakdown"]),
                detailed=bool(state.get("detailed", False)),
                transactions=list(payload["transactions"]),
                balance_title=str(payload["balance_title"]),
                show_transfer_summary=str(payload["scope"]) == "user",
                lang=lang,
            )
            keyboard = self._report_menu_keyboard(
                chat.id,
                selected_scope=str(payload["scope"]),
                selected_user_id=(
                    int(payload["selected_user_id"])
                    if isinstance(payload.get("selected_user_id"), int)
                    else None
                ),
                detailed=bool(state.get("detailed", False)),
                period_mode=str(payload["period_mode"]),
                lang=lang,
            )
            sent = await context.bot.send_message(
                chat_id=chat.id,
                text=report_text,
                parse_mode=ParseMode.HTML,
                reply_markup=keyboard,
            )
            state["report_message_id"] = sent.message_id
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    logger.exception("Failed to delete single-date confirm prompt")
            await query.answer(self._report_text("done", lang))
            return

        if data == "rp:pd":
            state["period_mode"] = "day"
            state["custom_date"] = None
            state["range_from"] = None
            state["range_to"] = None
        elif data == "rp:pw":
            state["period_mode"] = "week"
            state["custom_date"] = None
            state["range_from"] = None
            state["range_to"] = None
        elif data == "rp:pm":
            state["period_mode"] = "month"
            state["custom_date"] = None
            state["range_from"] = None
            state["range_to"] = None
        elif data == "rp:pcd":
            context.user_data[REPORT_FILTER_INPUT_KEY] = {
                "chat_id": chat.id,
                "mode": "date_or_range",
            }
            await query.answer()
            if query.message:
                await query.message.reply_text(
                    self._report_text("ask_date_or_period", lang),
                    reply_markup=self._period_input_cancel_keyboard(lang),
                )
            return
        elif data == "rp:pf":
            context.user_data[REPORT_FILTER_INPUT_KEY] = {
                "chat_id": chat.id,
                "mode": "date_or_range",
            }
            await query.answer()
            if query.message:
                await query.message.reply_text(
                    self._report_text("ask_date_or_period", lang),
                    reply_markup=self._period_input_cancel_keyboard(lang),
                )
            return
        elif data == "rp:pcr":
            context.user_data[REPORT_FILTER_INPUT_KEY] = {
                "chat_id": chat.id,
                "mode": "range",
            }
            await query.answer()
            if query.message:
                await query.message.reply_text(
                    self._report_text("ask_period_only", lang),
                    reply_markup=self._period_input_cancel_keyboard(lang),
                )
            return
        elif data == "rp:pcancel":
            context.user_data.pop(REPORT_FILTER_INPUT_KEY, None)
            context.user_data.pop(REPORT_SINGLE_DATE_CONFIRM_KEY, None)
            if query.message:
                try:
                    await query.message.delete()
                except Exception:
                    logger.exception("Failed to delete report filter input message")
                    try:
                        await query.message.edit_reply_markup(reply_markup=None)
                    except Exception:
                        logger.exception(
                            "Failed to clear keyboard after report filter cancel fallback"
                        )
            await query.answer(self._report_text("cancelled", lang))
            return
        elif data == "rp:x":
            payload = self._build_report_payload(chat.id, state, lang=lang)
            await query.answer(self._report_text("collecting_excel", lang))
            await self._send_report_excel(
                chat_id=chat.id,
                context=context,
                period_title=str(payload["period_title"]),
                scope_title=str(payload["scope_title"]),
                transactions=list(payload["transactions"]),
                lang=lang,
            )
            return

        parsed = self._parse_report_callback_data(data)
        if parsed is not None:
            scope, selected_user_id, detailed = parsed
            state["scope"] = scope
            state["selected_user_id"] = selected_user_id
            state["detailed"] = detailed
        else:
            await query.answer()
            return

        payload = self._build_report_payload(chat.id, state, lang=lang)
        if str(payload["scope"]) == "all" and scope == "user":
            # Selected user no longer available.
            await query.answer(self._report_text("member_not_in_report_menu", lang), show_alert=True)
            state["scope"] = "all"
            state["selected_user_id"] = None
            state["detailed"] = detailed
            payload = self._build_report_payload(chat.id, state, lang=lang)

        report_text = self._build_report_text(
            period_title=str(payload["period_title"]),
            scope_title=str(payload["scope_title"]),
            breakdown=dict(payload["breakdown"]),
            detailed=bool(state.get("detailed", False)),
            transactions=list(payload["transactions"]),
            balance_title=str(payload["balance_title"]),
            show_transfer_summary=str(payload["scope"]) == "user",
            lang=lang,
        )
        keyboard = self._report_menu_keyboard(
            chat.id,
            selected_scope=str(payload["scope"]),
            selected_user_id=(
                int(payload["selected_user_id"])
                if isinstance(payload.get("selected_user_id"), int)
                else None
            ),
            detailed=bool(state.get("detailed", False)),
            period_mode=str(payload["period_mode"]),
            lang=lang,
        )

        try:
            await query.edit_message_text(
                text=report_text,
                parse_mode=ParseMode.HTML,
                reply_markup=keyboard,
            )
        except Exception as exc:
            if "message is not modified" in str(exc).lower():
                await query.answer()
                return
            logger.exception("Failed to update report by callback")
            await query.answer(self._report_text("cant_open_report", lang), show_alert=True)
            return

        if query.message:
            state["report_message_id"] = query.message.message_id
        await query.answer()

    @staticmethod
    def _bold_html(text: str) -> str:
        return f"<b>{html.escape(text)}</b>"

    @staticmethod
    def _parse_report_callback_data(data: str) -> tuple[str, int | None, bool] | None:
        parts = data.split(":")
        if not parts or parts[0] != "rp":
            return None

        # Backward compatibility:
        # rp:f / rp:a / rp:u:<id>
        if len(parts) == 2 and parts[1] in {"f", "a"}:
            return ("family", None, False) if parts[1] == "f" else ("all", None, False)
        if len(parts) == 3 and parts[1] == "u":
            try:
                return "user", int(parts[2]), False
            except ValueError:
                return None

        # New format:
        # rp:f:s|d
        # rp:a:s|d
        # rp:u:<id>:s|d
        if len(parts) == 3 and parts[1] in {"f", "a"}:
            detailed = parts[2] == "d"
            return ("family", None, detailed) if parts[1] == "f" else ("all", None, detailed)
        if len(parts) == 4 and parts[1] == "u":
            try:
                user_id = int(parts[2])
            except ValueError:
                return None
            detailed = parts[3] == "d"
            return "user", user_id, detailed
        return None

    def _report_member_items(
        self, chat_id: int, *, lang: str = DEFAULT_LANGUAGE
    ) -> list[tuple[int, str]]:
        items: list[tuple[int, str]] = []
        members = self.db.list_members(chat_id)
        for member in members[:2]:
            user_id = int(member.get("telegram_user_id") or 0)
            raw_name = str(
                member.get("custom_name")
                or member.get("full_name")
                or self._report_text("member_fallback", lang, id=user_id)
            )
            compact_name = " ".join(raw_name.split()).strip() or self._report_text(
                "member_fallback", lang, id=user_id
            )
            if len(compact_name) > 16:
                compact_name = f"{compact_name[:15]}…"
            items.append((user_id, compact_name))
        return items

    def _report_menu_keyboard(
        self,
        chat_id: int,
        *,
        selected_scope: str | None = None,
        selected_user_id: int | None = None,
        detailed: bool = False,
        period_mode: str = "day",
        lang: str = DEFAULT_LANGUAGE,
    ) -> InlineKeyboardMarkup:
        ui_lang = normalize_language(lang)
        _ = period_mode
        rows: list[list[InlineKeyboardButton]] = []

        if selected_scope == "user" and selected_user_id is not None:
            toggle_callback = (
                f"rp:u:{selected_user_id}:s" if detailed else f"rp:u:{selected_user_id}:d"
            )
        elif selected_scope == "family":
            toggle_callback = "rp:f:s" if detailed else "rp:f:d"
        else:
            toggle_callback = "rp:a:s" if detailed else "rp:a:d"
        toggle_text = (
            self._report_text("toggle_hide_details", ui_lang)
            if detailed
            else self._report_text("toggle_show_details", ui_lang)
        )
        rows.append([InlineKeyboardButton(toggle_text, callback_data=toggle_callback)])

        scope_suffix = "d" if detailed else "s"
        scope_row: list[InlineKeyboardButton] = []
        for user_id, name in self._report_member_items(chat_id, lang=ui_lang):
            active_prefix = "✅ " if selected_scope == "user" and selected_user_id == user_id else ""
            scope_row.append(
                InlineKeyboardButton(
                    f"{active_prefix}👤 {name}",
                    callback_data=f"rp:u:{user_id}:{scope_suffix}",
                )
            )
        family_prefix = "✅ " if selected_scope == "family" else ""
        all_prefix = "✅ " if selected_scope == "all" else ""
        scope_row.append(
            InlineKeyboardButton(
                f"{family_prefix}👨‍👩‍👧‍👦 {self._report_text('scope_family', ui_lang)}",
                callback_data=f"rp:f:{scope_suffix}",
            )
        )
        scope_row.append(
            InlineKeyboardButton(
                f"{all_prefix}📊 {self._report_text('scope_all', ui_lang)}",
                callback_data=f"rp:a:{scope_suffix}",
            )
        )
        rows.append(scope_row)
        rows.append(
            [
                InlineKeyboardButton(
                    self._report_text("pick_period", ui_lang),
                    callback_data="rp:pf",
                ),
                InlineKeyboardButton(self._report_text("download_excel", ui_lang), callback_data="rp:x"),
            ]
        )
        return InlineKeyboardMarkup(rows)

    def _build_report_text(
        self,
        *,
        period_title: str,
        scope_title: str,
        breakdown: dict[str, dict[str, dict[str, object]]],
        detailed: bool,
        transactions: list[dict[str, object]],
        balance_title: str,
        show_transfer_summary: bool = False,
        lang: str = DEFAULT_LANGUAGE,
    ) -> str:
        ui_lang = normalize_language(lang)
        lines: list[str] = [
            self._bold_html(f"📅 {period_title}"),
            "",
            self._bold_html(scope_title),
            "",
            "━━━━━━━━━━━━━━━━━━",
            "",
        ]
        lines.extend(
            self._build_day_balance_section(
                breakdown,
                bold=True,
                balance_title=balance_title,
                lang=ui_lang,
            )
        )
        if show_transfer_summary:
            transfer_lines = self._build_transfer_summary_lines(transactions, bold=True, lang=ui_lang)
            if transfer_lines:
                lines.extend([""])
                lines.extend(transfer_lines)
        lines.extend(["", "━━━━━━━━━━━━━━━━━━", ""])
        if detailed:
            lines.extend(
                self._build_daily_detailed_section(
                    breakdown=breakdown,
                    transactions=transactions,
                    lang=ui_lang,
                )
            )
        else:
            lines.extend(self._build_daily_breakdown_section(breakdown, bold=True, lang=ui_lang))

        while lines and not lines[-1].strip():
            lines.pop()
        text = "\n".join(lines)
        if len(text) > 3900:
            text = text[:3800].rstrip()
            text += (
                "\n\n<i>"
                + html.escape(self._report_text("truncated_report_note", ui_lang))
                + "</i>"
            )
        return text

    def _build_transfer_summary_lines(
        self,
        transactions: list[dict[str, object]],
        *,
        bold: bool = False,
        lang: str = DEFAULT_LANGUAGE,
    ) -> list[str]:
        ui_lang = normalize_language(lang)
        sent: dict[str, float] = {}
        received: dict[str, float] = {}
        for tx in transactions:
            category = str(tx.get("category") or "").strip().lower()
            if category not in {TRANSFER_OUT_CATEGORY, TRANSFER_IN_CATEGORY}:
                continue
            currency = str(tx.get("currency") or "").upper() or self.settings.default_currency
            amount = float(tx.get("amount") or 0.0)
            if amount <= 0:
                continue
            if category == TRANSFER_OUT_CATEGORY:
                sent[currency] = float(sent.get(currency, 0.0)) + amount
            else:
                received[currency] = float(received.get(currency, 0.0)) + amount

        currencies = sorted(set(sent.keys()) | set(received.keys()))
        if not currencies:
            return []

        lines: list[str] = []
        for currency in currencies:
            sent_value = float(sent.get(currency, 0.0))
            recv_value = float(received.get(currency, 0.0))
            line_sent = self._report_text(
                "transfers_sent",
                ui_lang,
                amount=_fmt_money(sent_value, currency),
            )
            line_recv = self._report_text(
                "transfers_received",
                ui_lang,
                amount=_fmt_money(recv_value, currency),
            )
            lines.append(self._bold_html(line_sent) if bold else line_sent)
            lines.append(self._bold_html(line_recv) if bold else line_recv)
        return lines

    @staticmethod
    def _subscription_markers(text_l: str) -> bool:
        markers = (
            "подписк",
            "subscription",
            "chatgpt",
            "tg premium",
            "telegram premium",
            "youtube premium",
            "spotify",
            "netflix",
            "apple one",
            "icloud",
            "yandex plus",
            "яндекс плюс",
        )
        return any(marker in text_l for marker in markers)

    def _normalize_report_category(
        self,
        *,
        kind: str,
        category: str,
        description: str,
    ) -> str:
        raw_key = (category or "").strip().lower().replace("-", "_").replace(" ", "_")
        if kind != "expense":
            return raw_key

        desc_l = (description or "").strip().lower()
        if raw_key in {"work_subscriptions", "leisure_subscriptions"}:
            return "leisure_subscriptions"
        if raw_key == "leisure_games" and self._subscription_markers(desc_l):
            return "leisure_subscriptions"
        if self._subscription_markers(desc_l):
            return "leisure_subscriptions"
        return raw_key

    def _build_breakdown_from_transactions(
        self, transactions: list[dict[str, object]]
    ) -> dict[str, dict[str, dict[str, object]]]:
        result: dict[str, dict[str, dict[str, object]]] = {"expense": {}, "income": {}}
        category_sums: dict[str, dict[str, dict[str, float]]] = {"expense": {}, "income": {}}

        for tx in transactions:
            category_raw = str(tx.get("category") or "").strip().lower()
            if category_raw in {TRANSFER_OUT_CATEGORY, TRANSFER_IN_CATEGORY}:
                continue
            kind = str(tx.get("kind") or "")
            if kind not in {"expense", "income"}:
                continue
            currency = str(tx.get("currency") or "").upper()
            if not currency:
                continue
            amount = float(tx.get("amount") or 0.0)
            if amount <= 0:
                continue

            category = self._normalize_report_category(
                kind=kind,
                category=category_raw,
                description=str(tx.get("description") or ""),
            )

            by_kind = result.setdefault(kind, {})
            by_currency = by_kind.setdefault(currency, {"total": 0.0, "categories": []})
            by_currency["total"] = float(by_currency.get("total", 0.0)) + amount

            sums_kind = category_sums.setdefault(kind, {})
            sums_currency = sums_kind.setdefault(currency, {})
            sums_currency[category] = float(sums_currency.get(category, 0.0)) + amount

        for kind, by_currency in category_sums.items():
            result_kind = result.setdefault(kind, {})
            for currency, sums in by_currency.items():
                bucket = result_kind.setdefault(currency, {"total": 0.0, "categories": []})
                bucket["categories"] = sorted(
                    [(cat, total) for cat, total in sums.items()],
                    key=lambda item: float(item[1]),
                    reverse=True,
                )

        return result

    def _build_day_balance_section(
        self,
        breakdown: dict[str, dict[str, dict[str, object]]],
        *,
        bold: bool = False,
        balance_title: str = "💰 Баланс дня",
        lang: str = DEFAULT_LANGUAGE,
    ) -> list[str]:
        ui_lang = normalize_language(lang)
        totals: dict[str, dict[str, float]] = {}
        for kind in ("income", "expense"):
            by_currency = breakdown.get(kind, {})
            if not isinstance(by_currency, dict):
                continue
            for currency, bucket in by_currency.items():
                if not isinstance(bucket, dict):
                    continue
                total = float(bucket.get("total", 0.0))
                stats = totals.setdefault(str(currency), {"income": 0.0, "expense": 0.0})
                stats[kind] = total

        if not totals:
            totals[self.settings.default_currency] = {"income": 0.0, "expense": 0.0}

        lines = [self._bold_html(balance_title) if bold else balance_title]
        if len(totals) == 1:
            currency = sorted(totals.keys())[0]
            income = float(totals[currency].get("income", 0.0))
            expense = float(totals[currency].get("expense", 0.0))
            balance = income - expense
            lines.append(f"{self._report_text('income', ui_lang)}: {_fmt_money(income, currency)}")
            lines.append(f"{self._report_text('expense', ui_lang)}: {_fmt_money(expense, currency)}")
            total_line = f"{self._report_text('total', ui_lang)}: {_fmt_money(balance, currency, signed=True)}"
            lines.append(self._bold_html(total_line) if bold else total_line)
            return lines

        for currency in sorted(totals.keys()):
            income = float(totals[currency].get("income", 0.0))
            expense = float(totals[currency].get("expense", 0.0))
            balance = income - expense
            lines.append(
                f"{self._report_text('income', ui_lang)} ({str(currency).upper()}): {_fmt_money(income, currency)}"
            )
            lines.append(
                f"{self._report_text('expense', ui_lang)} ({str(currency).upper()}): {_fmt_money(expense, currency)}"
            )
            total_line = (
                f"{self._report_text('total', ui_lang)} ({str(currency).upper()}): "
                f"{_fmt_money(balance, currency, signed=True)}"
            )
            lines.append(self._bold_html(total_line) if bold else total_line)
            lines.append("")

        while lines and not lines[-1].strip():
            lines.pop()
        return lines

    @staticmethod
    def _tx_local_datetime(value: object) -> datetime | None:
        raw = str(value or "").strip()
        if not raw:
            return None
        try:
            dt_utc = datetime.strptime(raw, "%Y-%m-%d %H:%M:%S").replace(
                tzinfo=timezone.utc
            )
            return dt_utc.astimezone()
        except ValueError:
            return None

    @classmethod
    def _tx_time_hhmm(cls, value: object) -> str:
        dt_local = cls._tx_local_datetime(value)
        if dt_local is None:
            return "--:--"
        return dt_local.strftime("%H:%M")

    @staticmethod
    def _fmt_detail_amount(amount: float, currency: str) -> str:
        curr = str(currency).upper()
        if curr == "UZS":
            value = float(amount)
            normalized = math.ceil(value / 100.0) * 100.0 if value > 0 else 0.0
            return _fmt_money(normalized, curr)
        return _fmt_money(float(amount), curr)

    def _build_daily_detailed_section(
        self,
        *,
        breakdown: dict[str, dict[str, dict[str, object]]],
        transactions: list[dict[str, object]],
        lang: str = DEFAULT_LANGUAGE,
    ) -> list[str]:
        ui_lang = normalize_language(lang)
        if not transactions:
            return [self._report_text("no_ops_period", ui_lang)]

        lines: list[str] = []
        by_kind_currency_category: dict[str, dict[str, dict[str, list[dict[str, object]]]]] = {}
        for tx in transactions:
            kind = str(tx.get("kind") or "")
            currency = str(tx.get("currency") or "").upper()
            category = self._normalize_report_category(
                kind=kind,
                category=str(tx.get("category") or ""),
                description=str(tx.get("description") or ""),
            )
            if kind not in {"income", "expense"} or not currency:
                continue
            by_kind = by_kind_currency_category.setdefault(kind, {})
            by_currency = by_kind.setdefault(currency, {})
            by_category = by_currency.setdefault(category, [])
            by_category.append(tx)

        for kind in ("expense", "income"):
            kind_bucket = by_kind_currency_category.get(kind, {})
            if not isinstance(kind_bucket, dict) or not kind_bucket:
                continue

            summary_by_currency = breakdown.get(kind, {})
            if not isinstance(summary_by_currency, dict):
                continue

            for currency in sorted(summary_by_currency.keys()):
                currency_summary = summary_by_currency.get(currency)
                if not isinstance(currency_summary, dict):
                    continue
                total = float(currency_summary.get("total", 0.0))
                if total <= 0:
                    continue

                kind_title = (
                    self._report_text("section_expense", ui_lang)
                    if kind == "expense"
                    else self._report_text("section_income", ui_lang)
                )
                lines.append(self._bold_html(f"{kind_title} ({_fmt_money(total, currency)})"))
                lines.append("")

                category_order = [
                    str(item[0])
                    for item in currency_summary.get("categories", [])
                    if isinstance(item, (tuple, list)) and len(item) >= 1
                ]
                if not category_order:
                    category_order = list(kind_bucket.get(currency, {}).keys())

                for category in category_order:
                    tx_list = kind_bucket.get(currency, {}).get(category, [])
                    if not tx_list:
                        continue
                    tx_list = sorted(
                        tx_list,
                        key=lambda item: str(item.get("created_at") or ""),
                    )
                    emoji, label = self._report_category_view(category, kind)
                    lines.append(f"{emoji} {html.escape(label)}")
                    for tx in tx_list:
                        amount = float(tx.get("amount") or 0.0)
                        description = html.escape(str(tx.get("description") or ""))
                        time_part = self._tx_time_hhmm(tx.get("created_at"))
                        lines.append(
                            f"[{time_part}] — {self._fmt_detail_amount(amount, currency)} — {description}"
                        )
                    lines.append("")

        while lines and not lines[-1].strip():
            lines.pop()
        return lines

    def _build_report_excel_file(
        self,
        *,
        period_title: str,
        scope_title: str,
        transactions: list[dict[str, object]],
        lang: str = DEFAULT_LANGUAGE,
    ) -> str:
        ui_lang = normalize_language(lang)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except Exception as exc:
            raise RuntimeError(
                "Для Excel-выгрузки установите пакет openpyxl: pip install openpyxl"
            ) from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._report_text("excel_sheet_title", ui_lang)

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", fgColor="EFEFEF")
        header_font = Font(bold=True)
        income_font = Font(color="008000")
        expense_font = Font(color="C00000")
        transfer_font = Font(color="1F4E79")
        total_font = Font(bold=True)

        sheet.merge_cells("A1:G1")
        sheet["A1"] = f"{self._report_text('excel_title_prefix', ui_lang)}: {period_title} | {scope_title}"
        sheet["A1"].font = Font(bold=True, size=12)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

        headers = [
            self._report_text("excel_col_date", ui_lang),
            self._report_text("excel_col_time", ui_lang),
            self._report_text("excel_col_type", ui_lang),
            self._report_text("excel_col_category", ui_lang),
            self._report_text("excel_col_amount", ui_lang),
            self._report_text("excel_col_currency", ui_lang),
            self._report_text("excel_col_description", ui_lang),
        ]
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

        totals: dict[str, dict[str, float]] = {}
        transfer_totals: dict[str, float] = {}
        ordered_transactions = sorted(
            transactions,
            key=lambda item: str(item.get("created_at") or ""),
        )
        row = header_row + 1
        for tx in ordered_transactions:
            kind = str(tx.get("kind") or "")
            if kind not in {"income", "expense"}:
                continue
            amount = float(tx.get("amount") or 0.0)
            if amount <= 0:
                continue
            currency = str(tx.get("currency") or "").upper() or self.settings.default_currency
            category_raw = str(tx.get("category") or "").strip().lower()
            is_transfer = category_raw in {TRANSFER_OUT_CATEGORY, TRANSFER_IN_CATEGORY}
            description_raw = str(tx.get("description") or "")
            description = " ".join(description_raw.split()).strip()
            if len(description) > 180:
                description = f"{description[:177]}..."

            dt_local = self._tx_local_datetime(tx.get("created_at"))
            date_part = dt_local.strftime("%d/%m/%Y") if dt_local else "--/--/----"
            time_part = dt_local.strftime("%H:%M") if dt_local else "--:--"
            if is_transfer:
                type_label = self._report_text("excel_type_transfer", ui_lang)
                category_full = self._report_text("excel_family_transfers", ui_lang)
            else:
                type_label = (
                    self._report_text("excel_type_income", ui_lang)
                    if kind == "income"
                    else self._report_text("excel_type_expense", ui_lang)
                )
                category_key = self._normalize_report_category(
                    kind=kind,
                    category=category_raw,
                    description=description,
                )
                emoji, category_label = self._report_category_view(category_key, kind)
                category_full = f"{emoji} {category_label}"

            sheet.cell(row=row, column=1, value=date_part)
            sheet.cell(row=row, column=2, value=time_part)
            type_cell = sheet.cell(row=row, column=3, value=type_label)
            sheet.cell(row=row, column=4, value=category_full)
            amount_cell = sheet.cell(row=row, column=5, value=int(round(amount)))
            sheet.cell(row=row, column=6, value=currency)
            sheet.cell(row=row, column=7, value=description)

            if is_transfer:
                type_cell.font = transfer_font
            else:
                type_cell.font = income_font if kind == "income" else expense_font
            amount_cell.number_format = "#,##0"

            for col in range(1, 8):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row += 1

            if is_transfer:
                transfer_totals[currency] = float(transfer_totals.get(currency, 0.0)) + amount
            else:
                stats = totals.setdefault(currency, {"income": 0.0, "expense": 0.0})
                stats[kind] = float(stats.get(kind, 0.0)) + amount

        if not totals:
            totals[self.settings.default_currency] = {"income": 0.0, "expense": 0.0}

        row += 1
        for currency in sorted(totals.keys()):
            income = float(totals[currency].get("income", 0.0))
            expense = float(totals[currency].get("expense", 0.0))
            balance = income - expense
            for label, value in (
                (self._report_text("excel_total_income", ui_lang), income),
                (self._report_text("excel_total_expense", ui_lang), expense),
                (self._report_text("excel_total_balance", ui_lang), balance),
            ):
                sheet.cell(row=row, column=3, value=label).font = total_font
                amount_total = sheet.cell(row=row, column=5, value=int(round(value)))
                amount_total.number_format = "#,##0"
                amount_total.font = total_font
                sheet.cell(row=row, column=6, value=currency).font = total_font
                for col in range(1, 8):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1
            transfer_total = float(transfer_totals.get(currency, 0.0))
            if transfer_total > 0:
                sheet.cell(row=row, column=3, value=self._report_text("excel_total_transfers", ui_lang)).font = total_font
                transfer_total_cell = sheet.cell(
                    row=row,
                    column=5,
                    value=int(round(transfer_total)),
                )
                transfer_total_cell.number_format = "#,##0"
                transfer_total_cell.font = total_font
                sheet.cell(row=row, column=6, value=currency).font = total_font
                for col in range(1, 8):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

        sheet.column_dimensions["A"].width = 13
        sheet.column_dimensions["B"].width = 8
        sheet.column_dimensions["C"].width = 17
        sheet.column_dimensions["D"].width = 30
        sheet.column_dimensions["E"].width = 14
        sheet.column_dimensions["F"].width = 10
        sheet.column_dimensions["G"].width = 40
        sheet.auto_filter.ref = f"A{header_row}:G{max(header_row, row - 1)}"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            output_path = tmp.name
        workbook.save(output_path)
        return output_path

    async def _send_report_excel(
        self,
        *,
        chat_id: int,
        context: ContextTypes.DEFAULT_TYPE,
        period_title: str,
        scope_title: str,
        transactions: list[dict[str, object]],
        lang: str = DEFAULT_LANGUAGE,
    ) -> None:
        try:
            file_path = self._build_report_excel_file(
                period_title=period_title,
                scope_title=scope_title,
                transactions=transactions,
                lang=lang,
            )
        except RuntimeError as exc:
            await context.bot.send_message(chat_id=chat_id, text=str(exc))
            return
        except Exception:
            logger.exception("Failed to generate report excel file")
            await context.bot.send_message(
                chat_id=chat_id,
                text=self._report_text("excel_generate_failed", lang),
            )
            return

        file_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            with open(file_path, "rb") as report_file:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=report_file,
                    filename=file_name,
                    caption=f"{period_title}\n{scope_title}",
                )
        finally:
            try:
                os.unlink(file_path)
            except OSError:
                logger.exception("Failed to remove temporary report excel file")

    @staticmethod
    def _report_category_view(category: str, kind: str) -> tuple[str, str]:
        key = category.lower().replace("-", "_").replace(" ", "_")
        if key in {"leisure_subscriptions", "work_subscriptions"}:
            return "🧾", "Подписки"
        label = _ru_category_label(category, kind)

        income_emoji = {
            "salary": "💼",
            "bonus": "🎁",
            "windfall": "🍀",
            "profit": "📈",
            "cashback": "💳",
            "gift": "🎁",
            "income_other": "📦",
            "finance_debt_received": "💳",
        }
        if kind == "income":
            return income_emoji.get(key, "📦"), label

        group = label
        item = label
        if " / " in label:
            group, item = label.split(" / ", 1)

        expense_group_emoji = {
            "Жильё и дом": "🏠",
            "Продукты и быт": "🛒",
            "Кафе и рестораны": "🍽",
            "Транспорт": "🚗",
            "Работа и бизнес": "💼",
            "Образование": "🎓",
            "Здоровье": "🏥",
            "Одежда и уход": "👕",
            "Дети": "👶",
            "Развлечения": "🎉",
            "Финансы": "💳",
            "Животные": "🐾",
            "Прочие расходы": "📦",
        }
        return expense_group_emoji.get(group, "📦"), item

    def _build_daily_breakdown_section(
        self,
        breakdown: dict[str, dict[str, dict[str, object]]],
        *,
        bold: bool = False,
        max_categories: int | None = 4,
        lang: str = DEFAULT_LANGUAGE,
    ) -> list[str]:
        ui_lang = normalize_language(lang)
        lines: list[str] = []
        has_any = False
        kind_titles = {
            "expense": self._report_text("section_expense", ui_lang),
            "income": self._report_text("section_income", ui_lang),
        }

        for kind in ("expense", "income"):
            by_currency = breakdown.get(kind, {})
            if not isinstance(by_currency, dict) or not by_currency:
                continue

            for currency in sorted(by_currency.keys()):
                bucket = by_currency[currency]
                if not isinstance(bucket, dict):
                    continue
                total = float(bucket.get("total", 0.0))
                if total <= 0:
                    continue
                categories = bucket.get("categories", [])

                has_any = True
                section_title = kind_titles[kind]
                lines.append(self._bold_html(section_title) if bold else section_title)
                lines.append("")
                if isinstance(categories, list):
                    visible_categories = categories
                    if isinstance(max_categories, int) and max_categories > 0:
                        visible_categories = categories[:max_categories]
                    for category, amount in visible_categories:
                        amount_value = float(amount)
                        percent = int(round((amount_value / total) * 100)) if total else 0
                        emoji, label = self._report_category_view(str(category), kind)
                        safe_label = html.escape(label)
                        lines.append(
                            f"{emoji} {safe_label} — {_fmt_money(amount_value, str(currency))} ({percent}%)"
                        )
                    hidden_count = len(categories) - len(visible_categories)
                    if hidden_count > 0:
                        lines.append(self._report_text("more_categories", ui_lang, count=hidden_count))
                lines.append("")
                total_line = f"{self._report_text('total', ui_lang)}: {_fmt_money(total, str(currency))}"
                lines.append(self._bold_html(total_line) if bold else total_line)
                lines.append("")

        if not has_any:
            return [self._report_text("no_ops_period", ui_lang)]

        while lines and not lines[-1].strip():
            lines.pop()
        return lines

    def _build_user_section(self, summary: dict[str, object]) -> list[str]:
        totals = summary.get("totals", {})
        categories = summary.get("top_expense_categories", [])
        transfer_totals = summary.get("transfer_totals", {})
        lines: list[str] = []

        if not isinstance(totals, dict) or not totals:
            totals = {}

        currencies = set(totals.keys()) if isinstance(totals, dict) else set()
        if isinstance(transfer_totals, dict):
            currencies |= set(transfer_totals.keys())
        if not currencies:
            lines.append("- Пока операций нет.")
            return lines

        for currency in sorted(currencies):
            stat = totals.get(currency, {}) if isinstance(totals, dict) else {}
            if not isinstance(stat, dict):
                stat = {}
            income = float(stat.get("income", 0.0))
            expense = float(stat.get("expense", 0.0))
            transfer_stat = (
                transfer_totals.get(currency, {}) if isinstance(transfer_totals, dict) else {}
            )
            sent = (
                float(transfer_stat.get("sent", 0.0))
                if isinstance(transfer_stat, dict)
                else 0.0
            )
            received = (
                float(transfer_stat.get("received", 0.0))
                if isinstance(transfer_stat, dict)
                else 0.0
            )
            balance = income - expense + received - sent
            lines.append(
                f"- доход {_fmt_money(income, currency)}, "
                f"расход {_fmt_money(expense, currency)}, "
                f"баланс {_fmt_money(balance, currency, signed=True)}"
            )
            if sent > 0 or received > 0:
                lines.append(f"  ⬅ Переводы отправлено: {_fmt_money(sent, currency)}")
                lines.append(f"  ➡ Переводы получено: {_fmt_money(received, currency)}")

        if isinstance(categories, list) and categories:
            lines.append("- Топ категорий расходов:")
            for category, currency, amount in categories[:5]:
                lines.append(
                    f"- {_ru_category_label(str(category), 'expense')}: "
                    f"{_fmt_money(float(amount), currency)}"
                )
        return lines

    def _build_family_section(self, summary: dict[str, object]) -> list[str]:
        totals = summary.get("totals", {})
        categories = summary.get("top_categories", [])
        lines = ["Семейные траты:"]

        if not isinstance(totals, dict) or not totals:
            lines.append("- Пока операций нет.")
            return lines

        for currency in sorted(totals.keys()):
            total = float(totals[currency])
            lines.append(f"- {_fmt_money(total, currency)}")

        if isinstance(categories, list) and categories:
            lines.append("- Топ семейных категорий:")
            for category, currency, amount in categories[:5]:
                lines.append(
                    f"- {_ru_category_label(str(category), 'expense')}: "
                    f"{_fmt_money(float(amount), currency)}"
                )
        return lines
